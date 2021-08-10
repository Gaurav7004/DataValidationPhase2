import re, os, sys, itertools, string, time
import numpy as np
import pandas as pd
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import QTimer, Qt
from PyQt5.QtWidgets import QFileDialog, QLabel,QMessageBox, QSizePolicy, QTabWidget, QWidget
import matplotlib.pyplot as plt
from matplotlib.pyplot import axes, figure
from openpyxl.styles import PatternFill, Alignment, Font
from HealthSubCentreVal import *
from PrimaryHealthCentre import *
from SubDistrictHospitalVal import *
from DistrictHospitalVal import *
from CommunityHealthCentreVal import *
from pandas.io.formats import style
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from PyQt5.QtCore import QThread, pyqtSignal, QObject, pyqtSlot
from PyQt5.QtWidgets import QWidget, QProgressBar, QVBoxLayout


class Worker(QObject):
    finished = pyqtSignal()
    intReady = pyqtSignal(int)
    
    @pyqtSlot()
    def proc_counter(self):  # A slot takes no params
        def decimal_range(start, stop, increment):
            while start < stop: # and not math.isclose(start, stop): Py>3.5
                yield start
                start += increment

        for i in decimal_range(1, 100, 1):
            time.sleep(0.1)
            self.intReady.emit(i)

        self.finished.emit()


class PopUpProgressB(QWidget):

    def __init__(self):
        super().__init__()
        self.pbar = QProgressBar(self)
        self.label_1 = QLabel(self)
        self.pbar.setGeometry(30, 40, 500, 75)
        self.label_1.setGeometry(40, 50, 500, 75)
        self.label_1.setText("This is working")
        self.layout = QVBoxLayout()
        self.layout.addWidget(self.label_1)
        self.layout.addWidget(self.pbar)
        self.setLayout(self.layout)
        self.setGeometry(700, 500, 500, 200)
        self.setWindowTitle('File is saving, please wait!')
        # self.setWindowTitle.setStyleSheet("color: rgb(0, 0, 127;\n""font: 75 12pt;\n""")
        
        # self.show()

        self.obj = Worker()
        self.thread = QThread()
        self.obj.intReady.connect(self.on_count_changed)
        self.obj.moveToThread(self.thread)
        self.obj.finished.connect(self.thread.quit)
        self.obj.finished.connect(self.hide)  # To hide the progress bar after the progress is completed
        self.thread.started.connect(self.obj.proc_counter)
        # self.thread.start()  # This was moved to start_progress

    def start_progress(self):  # To restart the progress every time
        self.show()
        self.thread.start()

    def on_count_changed(self, value):
        self.pbar.setValue(value)


# ---------------------------------------------------- UI -----------------------------------------------
class Ui_TabWidget(object):
    def setupUi(self, TabWidget):
        TabWidget.setObjectName("TabWidget")
        TabWidget.resize(1619, 860)
        self.tab = QtWidgets.QWidget()
        self.tab.setObjectName("tab")
        self.frame = QtWidgets.QFrame(self.tab)
        self.frame.setGeometry(QtCore.QRect(10, 10, 1601, 811))
        self.frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame.setObjectName("frame")
        self.horizontalLayoutWidget = QtWidgets.QWidget(self.frame)
        self.horizontalLayoutWidget.setGeometry(QtCore.QRect(20, 9, 1551, 91))
        self.horizontalLayoutWidget.setObjectName("horizontalLayoutWidget")
        self.horizontalLayout = QtWidgets.QHBoxLayout(self.horizontalLayoutWidget)
        self.horizontalLayout.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.label = QtWidgets.QLabel(self.horizontalLayoutWidget)
        font = QtGui.QFont()
        font.setPointSize(28)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setStyleSheet("QLabel{background-color: #003679; color : white;}\n"
"\n"
"")
        self.label.setAlignment(QtCore.Qt.AlignCenter)
        self.label.setObjectName("label")
        self.horizontalLayout.addWidget(self.label)
        self.frame_2 = QtWidgets.QFrame(self.frame)
        self.frame_2.setGeometry(QtCore.QRect(10, 110, 1581, 691))
        self.frame_2.setStyleSheet("QFrame{background-color: rgb(255, 255, 255);}")
        self.frame_2.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_2.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_2.setObjectName("frame_2")
        self.pushButton_6 = QtWidgets.QPushButton(self.frame_2)
        self.popup = PopUpProgressB()
        self.pushButton_6.clicked.connect(self.upload)
        self.pushButton_6.setGeometry(QtCore.QRect(10, 10, 281, 61))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_6.setFont(font)
        self.pushButton_6.setStyleSheet("background-color: #73C067;\n"
"color: white;\n"
"")
        self.pushButton_6.setObjectName("pushButton_6")
        self.lineEdit_4 = QtWidgets.QLineEdit(self.frame_2)
        self.lineEdit_4.setGeometry(QtCore.QRect(360, 10, 651, 61))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_4.setFont(font)
        self.lineEdit_4.setStyleSheet("background-color: white;\n"
"color: #656565;\n"
"")
        self.lineEdit_4.setObjectName("lineEdit_4")
        self.label_21 = QtWidgets.QLabel(self.frame_2)
        self.label_21.setGeometry(QtCore.QRect(10, 80, 1001, 21))
        font = QtGui.QFont()
        font.setFamily("MS Sans Serif")
        font.setPointSize(8)
        font.setBold(False)
        font.setItalic(True)
        font.setWeight(50)
        self.label_21.setFont(font)
        self.label_21.setStyleSheet("QLabel{color : red;}")
        self.label_21.setObjectName("label_21")
        self.lineEdit_5 = QtWidgets.QLineEdit(self.frame_2)
        self.lineEdit_5.setGeometry(QtCore.QRect(10, 120, 451, 61))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_5.setFont(font)
        self.lineEdit_5.setStyleSheet("background-color: white;\n"
"color: #656565;\n"
"")
        self.lineEdit_5.setObjectName("lineEdit_5")
        self.lineEdit_6 = QtWidgets.QLineEdit(self.frame_2)
        self.lineEdit_6.setGeometry(QtCore.QRect(480, 120, 201, 61))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_6.setFont(font)
        self.lineEdit_6.setStyleSheet("background-color: white;\n"
"color: #656565;\n"
"")
        self.lineEdit_6.setText("")
        self.lineEdit_6.setObjectName("lineEdit_6")
        self.pushButton_12 = QtWidgets.QPushButton(self.frame_2)
        self.pushButton_12.clicked.connect(self.VerifyFType)
        self.pushButton_12.setGeometry(QtCore.QRect(700, 120, 311, 61))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_12.setFont(font)
        self.pushButton_12.setStyleSheet("background-color: #00A5DB;color: white;\n"
"")
        self.pushButton_12.setObjectName("pushButton_12")
        self.pushButton_19 = QtWidgets.QPushButton(self.frame_2)
        self.pushButton_19.clicked.connect(self.methodology)
        self.pushButton_19.setGeometry(QtCore.QRect(1350, 70, 211, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_19.setFont(font)
        self.pushButton_19.setStyleSheet("background-color: #F47B1F;\n"
"color: white;\n"
"")
        self.pushButton_19.setObjectName("pushButton_19")
        self.label_22 = QtWidgets.QLabel(self.frame_2)
        self.label_22.setGeometry(QtCore.QRect(700, 190, 591, 31))
        font = QtGui.QFont()
        font.setFamily("MS Sans Serif")
        font.setPointSize(8)
        font.setBold(False)
        font.setItalic(True)
        font.setWeight(50)
        self.label_22.setFont(font)
        self.label_22.setStyleSheet("QLabel{color : red;}")
        self.label_22.setObjectName("label_22")
        self.pushButton_20 = QtWidgets.QPushButton(self.frame_2)
        self.pushButton_20.clicked.connect(self.UserManual)
        self.pushButton_20.setGeometry(QtCore.QRect(1350, 130, 211, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_20.setFont(font)
        self.pushButton_20.setStyleSheet("background-color: #F47B1F;\n"
"color: white;\n"
"")
        self.pushButton_20.setObjectName("pushButton_20")
        self.frame_5 = QtWidgets.QFrame(self.frame_2)
        self.frame_5.setGeometry(QtCore.QRect(20, 240, 1041, 441))
        self.frame_5.setStyleSheet("QLabel{background-color: rgb(185, 255, 238);}\n"
"\n"
"")
        self.frame_5.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_5.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_5.setObjectName("frame_5")
        self.pushButton_13 = QtWidgets.QPushButton(self.frame_5)
        self.pushButton_13.installEventFilter(TabWidget)
        self.pushButton_13.clicked.connect(self.onSelectDistrict)
        self.pushButton_13.setGeometry(QtCore.QRect(180, 190, 271, 51))
        self.pushButton_13.setStyleSheet("\n"
"background-color: rgb(222, 222, 222);\n"
"")
        self.pushButton_13.setObjectName("pushButton_13")
        self.pushButton_14 = QtWidgets.QPushButton(self.frame_5)
        self.pushButton_14.installEventFilter(TabWidget)
        self.pushButton_14.clicked.connect(self.onSelectSubDistrict)
        self.pushButton_14.setGeometry(QtCore.QRect(180, 280, 271, 51))
        self.pushButton_14.setStyleSheet("\n"
"background-color: rgb(222, 222, 222);")
        self.pushButton_14.setObjectName("pushButton_14")
        self.pushButton_15 = QtWidgets.QPushButton(self.frame_5)
        self.pushButton_15.installEventFilter(TabWidget)
        self.pushButton_15.clicked.connect(self.onSelectBlock)
        self.pushButton_15.setGeometry(QtCore.QRect(180, 370, 271, 51))
        self.pushButton_15.setStyleSheet("\n"
"background-color: rgb(222, 222, 222);\n"
"")
        self.pushButton_15.setObjectName("pushButton_15")
        self.label_12 = QtWidgets.QLabel(self.frame_5)
        self.label_12.setGeometry(QtCore.QRect(10, 190, 151, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_12.setFont(font)
        self.label_12.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_12.setAlignment(QtCore.Qt.AlignCenter)
        self.label_12.setObjectName("label_12")
        self.label_13 = QtWidgets.QLabel(self.frame_5)
        self.label_13.setGeometry(QtCore.QRect(10, 280, 151, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_13.setFont(font)
        self.label_13.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_13.setAlignment(QtCore.Qt.AlignCenter)
        self.label_13.setObjectName("label_13")
        self.label_14 = QtWidgets.QLabel(self.frame_5)
        self.label_14.setGeometry(QtCore.QRect(10, 370, 151, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_14.setFont(font)
        self.label_14.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_14.setAlignment(QtCore.Qt.AlignCenter)
        self.label_14.setObjectName("label_14")
        self.label_15 = QtWidgets.QLabel(self.frame_5)
        self.label_15.setGeometry(QtCore.QRect(530, 100, 151, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_15.setFont(font)
        self.label_15.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_15.setAlignment(QtCore.Qt.AlignCenter)
        self.label_15.setObjectName("label_15")
        self.label_16 = QtWidgets.QLabel(self.frame_5)
        self.label_16.setGeometry(QtCore.QRect(290, 10, 471, 31))
        font = QtGui.QFont()
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_16.setFont(font)
        self.label_16.setStyleSheet("background-color: white;\n"
"color: grey;")
        self.label_16.setAlignment(QtCore.Qt.AlignCenter)
        self.label_16.setObjectName("label_16")
        self.label_17 = QtWidgets.QLabel(self.frame_5)
        self.label_17.setGeometry(QtCore.QRect(10, 100, 151, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_17.setFont(font)
        self.label_17.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_17.setAlignment(QtCore.Qt.AlignCenter)
        self.label_17.setObjectName("label_17")
        self.pushButton_16 = QtWidgets.QPushButton(self.frame_5)
        self.pushButton_16.installEventFilter(TabWidget)
        self.pushButton_16.clicked.connect(self.onSelectState)
        self.pushButton_16.setGeometry(QtCore.QRect(180, 100, 271, 51))
        self.pushButton_16.setStyleSheet("\n"
"background-color: rgb(222, 222, 222);\n"
"")
        self.pushButton_16.setObjectName("pushButton_16")
        self.label_18 = QtWidgets.QLabel(self.frame_5)
        self.label_18.setGeometry(QtCore.QRect(530, 190, 161, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_18.setFont(font)
        self.label_18.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_18.setAlignment(QtCore.Qt.AlignCenter)
        self.label_18.setObjectName("label_18")
        self.label_20 = QtWidgets.QLabel(self.frame_5)
        self.label_20.setGeometry(QtCore.QRect(530, 280, 151, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_20.setFont(font)
        self.label_20.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_20.setAlignment(QtCore.Qt.AlignCenter)
        self.label_20.setObjectName("label_20")
        self.label_24 = QtWidgets.QLabel(self.frame_5)
        self.label_24.setGeometry(QtCore.QRect(520, 370, 171, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_24.setFont(font)
        self.label_24.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_24.setAlignment(QtCore.Qt.AlignCenter)
        self.label_24.setObjectName("label_24")
        self.pushButton_27 = QtWidgets.QPushButton(self.frame_5)
        self.pushButton_27.installEventFilter(TabWidget)
        self.pushButton_27.clicked.connect(self.onSelectFacilityName)
        self.pushButton_27.setGeometry(QtCore.QRect(720, 370, 271, 51))
        self.pushButton_27.setStyleSheet("\n"
"background-color: rgb(222, 222, 222);\n"
"")
        self.pushButton_27.setObjectName("pushButton_27")
        self.pushButton_28 = QtWidgets.QPushButton(self.frame_5)
        self.pushButton_28.installEventFilter(TabWidget)
        self.pushButton_28.clicked.connect(self.onSelectOwnership)
        self.pushButton_28.setGeometry(QtCore.QRect(720, 280, 271, 51))
        self.pushButton_28.setStyleSheet("\n"
"background-color: rgb(222, 222, 222);\n"
"")
        self.pushButton_28.setObjectName("pushButton_28")
        self.pushButton_29 = QtWidgets.QPushButton(self.frame_5)
        self.pushButton_29.installEventFilter(TabWidget)
        self.pushButton_29.clicked.connect(self.onSelectRuralUrban)
        self.pushButton_29.setGeometry(QtCore.QRect(720, 190, 271, 51))
        self.pushButton_29.setStyleSheet("\n"
"background-color: rgb(222, 222, 222);\n"
"")
        self.pushButton_29.setObjectName("pushButton_29")
        self.pushButton_30 = QtWidgets.QPushButton(self.frame_5)
        self.pushButton_30.installEventFilter(TabWidget)
        self.pushButton_30.clicked.connect(self.onSelectHealthBlock)
        self.pushButton_30.setGeometry(QtCore.QRect(720, 100, 271, 51))
        self.pushButton_30.setStyleSheet("\n"
"background-color: rgb(222, 222, 222);\n"
"")
        self.pushButton_30.setObjectName("pushButton_30")
        self.label_5 = QtWidgets.QLabel(self.frame_5)
        self.label_5.setGeometry(QtCore.QRect(630, 10, 91, 31))
        self.label_5.setStyleSheet("QLabel{background-color: rgb(255, 189, 167);}")
        self.label_5.setAlignment(QtCore.Qt.AlignCenter)
        self.label_5.setObjectName("label_5")
        self.pushButton_17 = QtWidgets.QPushButton(self.frame_2)
        self.pushButton_17.clicked.connect(self.export)
        self.pushButton_17.setGeometry(QtCore.QRect(1150, 350, 411, 71))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_17.setFont(font)
        self.pushButton_17.setStyleSheet("background-color: #73C067;color: white;")
        self.pushButton_17.setObjectName("pushButton_17")
        self.label_25 = QtWidgets.QLabel(self.frame_2)
        self.label_25.setGeometry(QtCore.QRect(1150, 430, 411, 31))
        font = QtGui.QFont()
        font.setFamily("MS Sans Serif")
        font.setPointSize(8)
        font.setBold(False)
        font.setItalic(True)
        font.setWeight(50)
        self.label_25.setFont(font)
        self.label_25.setStyleSheet("QLabel{color : red;}")
        self.label_25.setObjectName("label_25")
        self.pushButton_18 = QtWidgets.QPushButton(self.frame_2)
        self.pushButton_18.clicked.connect(self.reset)
        self.pushButton_18.setGeometry(QtCore.QRect(1150, 530, 411, 71))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_18.setFont(font)
        self.pushButton_18.setStyleSheet("background-color: #B00020;color: white;")
        self.pushButton_18.setObjectName("pushButton_18")
        self.label_3 = QtWidgets.QLabel(self.frame_2)
        self.label_3.setGeometry(QtCore.QRect(300, 20, 41, 31))
        self.label_3.setStyleSheet("QLabel{background-color: rgb(255, 189, 167);}")
        self.label_3.setAlignment(QtCore.Qt.AlignCenter)
        self.label_3.setObjectName("label_3")
        self.label_4 = QtWidgets.QLabel(self.frame_2)
        self.label_4.setGeometry(QtCore.QRect(1020, 130, 41, 31))
        self.label_4.setStyleSheet("QLabel{background-color: rgb(255, 189, 167);}")
        self.label_4.setAlignment(QtCore.Qt.AlignCenter)
        self.label_4.setObjectName("label_4")
        self.label_6 = QtWidgets.QLabel(self.frame_2)
        self.label_6.setGeometry(QtCore.QRect(1520, 310, 41, 31))
        self.label_6.setStyleSheet("QLabel{background-color: rgb(255, 189, 167);}")
        self.label_6.setAlignment(QtCore.Qt.AlignCenter)
        self.label_6.setObjectName("label_6")
        self.label_7 = QtWidgets.QLabel(self.frame_2)
        self.label_7.setGeometry(QtCore.QRect(1520, 490, 41, 31))
        self.label_7.setStyleSheet("QLabel{background-color: rgb(255, 189, 167);}")
        self.label_7.setAlignment(QtCore.Qt.AlignCenter)
        self.label_7.setObjectName("label_7")
        self.pushButton_38 = QtWidgets.QPushButton(self.frame_2)
        self.pushButton_38.clicked.connect(self.data_extraction)
        self.pushButton_38.setGeometry(QtCore.QRect(1350, 10, 211, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_38.setFont(font)
        self.pushButton_38.setStyleSheet("background-color: #F47B1F;\n"
"color: white;\n"
"")
        self.pushButton_38.setObjectName("pushButton_38")
        self.label_38 = QtWidgets.QLabel(self.frame_2)
        self.label_38.setGeometry(QtCore.QRect(1450, 650, 111, 21))
        self.label_38.setStyleSheet("QLabel{background-color: rgb(255, 189, 167);}")
        self.label_38.setAlignment(QtCore.Qt.AlignCenter)
        self.label_38.setObjectName("label_38")
        TabWidget.addTab(self.tab, "")
        self.tab1 = QtWidgets.QWidget()
        self.tab1.setObjectName("tab1")
        self.frame_3 = QtWidgets.QFrame(self.tab1)
        self.frame_3.setGeometry(QtCore.QRect(10, 10, 1601, 811))
        self.frame_3.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_3.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_3.setObjectName("frame_3")
        self.label_2 = QtWidgets.QLabel(self.frame_3)
        self.label_2.setGeometry(QtCore.QRect(20, 10, 1551, 101))
        font = QtGui.QFont()
        font.setPointSize(28)
        font.setBold(True)
        font.setWeight(75)
        self.label_2.setFont(font)
        self.label_2.setStyleSheet("QLabel{background-color: #003679; color : white;}\n"
"\n"
"")
        self.label_2.setAlignment(QtCore.Qt.AlignCenter)
        self.label_2.setObjectName("label_2")
        self.frame_4 = QtWidgets.QFrame(self.frame_3)
        self.frame_4.setGeometry(QtCore.QRect(10, 120, 1581, 681))
        self.frame_4.setStyleSheet("QFrame{background-color: rgb(255, 255, 255);}")
        self.frame_4.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_4.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_4.setObjectName("frame_4")
        self.pushButton_7 = QtWidgets.QPushButton(self.frame_4)
        self.pushButton_7.clicked.connect(self.upload)
        self.pushButton_7.setGeometry(QtCore.QRect(10, 10, 281, 61))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_7.setFont(font)
        self.pushButton_7.setStyleSheet("background-color: #73C067;\n"
"color: white;\n"
"")
        self.pushButton_7.setObjectName("pushButton_7")
        self.lineEdit_7 = QtWidgets.QLineEdit(self.frame_4)
        self.lineEdit_7.setGeometry(QtCore.QRect(360, 10, 641, 61))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_7.setFont(font)
        self.lineEdit_7.setStyleSheet("background-color: white;\n"
"color: #656565;\n"
"")
        self.lineEdit_7.setObjectName("lineEdit_7")
        self.label_33 = QtWidgets.QLabel(self.frame_4)
        self.label_33.setGeometry(QtCore.QRect(10, 80, 991, 21))
        font = QtGui.QFont()
        font.setFamily("MS Sans Serif")
        font.setPointSize(8)
        font.setBold(False)
        font.setItalic(True)
        font.setWeight(50)
        self.label_33.setFont(font)
        self.label_33.setStyleSheet("QLabel{color : red;}")
        self.label_33.setObjectName("label_33")
        self.pushButton_36 = QtWidgets.QPushButton(self.frame_4)
        self.pushButton_36.clicked.connect(self.methodology)
        self.pushButton_36.setGeometry(QtCore.QRect(1330, 70, 231, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_36.setFont(font)
        self.pushButton_36.setStyleSheet("background-color: #F47B1F;\n"
"color: white;\n"
"")
        self.pushButton_36.setObjectName("pushButton_36")
        self.pushButton_37 = QtWidgets.QPushButton(self.frame_4)
        self.pushButton_37.clicked.connect(self.UserManual)
        self.pushButton_37.setGeometry(QtCore.QRect(1330, 130, 231, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_37.setFont(font)
        self.pushButton_37.setStyleSheet("background-color: #F47B1F;\n"
"color: white;\n"
"")
        self.pushButton_37.setObjectName("pushButton_37")
        self.lineEdit_8 = QtWidgets.QLineEdit(self.frame_4)
        self.lineEdit_8.setGeometry(QtCore.QRect(10, 120, 451, 61))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_8.setFont(font)
        self.lineEdit_8.setStyleSheet("background-color: white;\n"
"color: #656565;\n"
"")
        self.lineEdit_8.setObjectName("lineEdit_8")
        self.lineEdit_9 = QtWidgets.QLineEdit(self.frame_4)
        self.lineEdit_9.setGeometry(QtCore.QRect(480, 120, 201, 61))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_9.setFont(font)
        self.lineEdit_9.setStyleSheet("background-color: white;\n"
"color: #656565;\n"
"")
        self.lineEdit_9.setText("")
        self.lineEdit_9.setObjectName("lineEdit_9")
        self.pushButton_21 = QtWidgets.QPushButton(self.frame_4)
        self.pushButton_21.clicked.connect(self.VerifyFType)
        self.pushButton_21.setGeometry(QtCore.QRect(700, 120, 301, 61))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_21.setFont(font)
        self.pushButton_21.setStyleSheet("background-color: #00A5DB;color: white;\n"
"")
        self.pushButton_21.setObjectName("pushButton_21")
        self.label_34 = QtWidgets.QLabel(self.frame_4)
        self.label_34.setGeometry(QtCore.QRect(700, 190, 591, 31))
        font = QtGui.QFont()
        font.setFamily("MS Sans Serif")
        font.setPointSize(8)
        font.setBold(False)
        font.setItalic(True)
        font.setWeight(50)
        self.label_34.setFont(font)
        self.label_34.setStyleSheet("QLabel{color : red;}")
        self.label_34.setObjectName("label_34")
        self.frame_6 = QtWidgets.QFrame(self.frame_4)
        self.frame_6.setGeometry(QtCore.QRect(10, 230, 1041, 441))
        self.frame_6.setStyleSheet("QLabel{background-color: rgb(185, 255, 238);}\n"
"\n"
"")
        self.frame_6.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_6.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_6.setObjectName("frame_6")
        self.pushButton_22 = QtWidgets.QPushButton(self.frame_6)
        self.pushButton_22.clicked.connect(self.onSelectDistrict)
        self.pushButton_22.setGeometry(QtCore.QRect(180, 190, 271, 51))
        self.pushButton_22.setStyleSheet("\n"
"background-color: rgb(222, 222, 222);\n"
"")
        self.pushButton_22.setObjectName("pushButton_22")
        self.pushButton_23 = QtWidgets.QPushButton(self.frame_6)
        self.pushButton_23.clicked.connect(self.onSelectSubDistrict)
        self.pushButton_23.setGeometry(QtCore.QRect(180, 280, 271, 51))
        self.pushButton_23.setStyleSheet("\n"
"background-color: rgb(222, 222, 222);")
        self.pushButton_23.setObjectName("pushButton_23")
        self.pushButton_24 = QtWidgets.QPushButton(self.frame_6)
        self.pushButton_24.clicked.connect(self.onSelectBlock)
        self.pushButton_24.setGeometry(QtCore.QRect(180, 370, 271, 51))
        self.pushButton_24.setStyleSheet("\n"
"background-color: rgb(222, 222, 222);\n"
"")
        self.pushButton_24.setObjectName("pushButton_24")
        self.label_19 = QtWidgets.QLabel(self.frame_6)
        self.label_19.setGeometry(QtCore.QRect(10, 190, 151, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_19.setFont(font)
        self.label_19.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_19.setAlignment(QtCore.Qt.AlignCenter)
        self.label_19.setObjectName("label_19")
        self.label_23 = QtWidgets.QLabel(self.frame_6)
        self.label_23.setGeometry(QtCore.QRect(10, 280, 151, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_23.setFont(font)
        self.label_23.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_23.setAlignment(QtCore.Qt.AlignCenter)
        self.label_23.setObjectName("label_23")
        self.label_26 = QtWidgets.QLabel(self.frame_6)
        self.label_26.setGeometry(QtCore.QRect(10, 370, 151, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_26.setFont(font)
        self.label_26.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_26.setAlignment(QtCore.Qt.AlignCenter)
        self.label_26.setObjectName("label_26")
        self.label_27 = QtWidgets.QLabel(self.frame_6)
        self.label_27.setGeometry(QtCore.QRect(530, 100, 151, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_27.setFont(font)
        self.label_27.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_27.setAlignment(QtCore.Qt.AlignCenter)
        self.label_27.setObjectName("label_27")
        self.label_28 = QtWidgets.QLabel(self.frame_6)
        self.label_28.setGeometry(QtCore.QRect(290, 10, 471, 31))
        font = QtGui.QFont()
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_28.setFont(font)
        self.label_28.setStyleSheet("background-color: white;\n"
"color: grey;")
        self.label_28.setAlignment(QtCore.Qt.AlignCenter)
        self.label_28.setObjectName("label_28")
        self.label_29 = QtWidgets.QLabel(self.frame_6)
        self.label_29.setGeometry(QtCore.QRect(10, 100, 151, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_29.setFont(font)
        self.label_29.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_29.setAlignment(QtCore.Qt.AlignCenter)
        self.label_29.setObjectName("label_29")
        self.pushButton_25 = QtWidgets.QPushButton(self.frame_6)
        self.pushButton_25.clicked.connect(self.onSelectState)
        self.pushButton_25.setGeometry(QtCore.QRect(180, 100, 271, 51))
        self.pushButton_25.setStyleSheet("\n"
"background-color: rgb(222, 222, 222);\n"
"")
        self.pushButton_25.setObjectName("pushButton_25")
        self.label_30 = QtWidgets.QLabel(self.frame_6)
        self.label_30.setGeometry(QtCore.QRect(530, 190, 161, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_30.setFont(font)
        self.label_30.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_30.setAlignment(QtCore.Qt.AlignCenter)
        self.label_30.setObjectName("label_30")
        self.label_31 = QtWidgets.QLabel(self.frame_6)
        self.label_31.setGeometry(QtCore.QRect(530, 280, 151, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_31.setFont(font)
        self.label_31.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_31.setAlignment(QtCore.Qt.AlignCenter)
        self.label_31.setObjectName("label_31")
        self.label_32 = QtWidgets.QLabel(self.frame_6)
        self.label_32.setGeometry(QtCore.QRect(510, 370, 191, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_32.setFont(font)
        self.label_32.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_32.setAlignment(QtCore.Qt.AlignCenter)
        self.label_32.setObjectName("label_32")
        self.pushButton_31 = QtWidgets.QPushButton(self.frame_6)
        self.pushButton_31.clicked.connect(self.onSelectFacilityName)
        self.pushButton_31.setGeometry(QtCore.QRect(720, 370, 271, 51))
        self.pushButton_31.setStyleSheet("\n"
"background-color: rgb(222, 222, 222);\n"
"")
        self.pushButton_31.setObjectName("pushButton_31")
        self.pushButton_32 = QtWidgets.QPushButton(self.frame_6)
        self.pushButton_32.clicked.connect(self.onSelectOwnership)
        self.pushButton_32.setGeometry(QtCore.QRect(720, 280, 271, 51))
        self.pushButton_32.setStyleSheet("\n"
"background-color: rgb(222, 222, 222);\n"
"")
        self.pushButton_32.setObjectName("pushButton_32")
        self.pushButton_33 = QtWidgets.QPushButton(self.frame_6)
        self.pushButton_33.clicked.connect(self.onSelectRuralUrban)
        self.pushButton_33.setGeometry(QtCore.QRect(720, 190, 271, 51))
        self.pushButton_33.setStyleSheet("\n"
"background-color: rgb(222, 222, 222);\n"
"")
        self.pushButton_33.setObjectName("pushButton_33")
        self.pushButton_34 = QtWidgets.QPushButton(self.frame_6)
        self.pushButton_34.clicked.connect(self.onSelectHealthBlock)
        self.pushButton_34.setGeometry(QtCore.QRect(720, 100, 271, 51))
        self.pushButton_34.setStyleSheet("\n"
"background-color: rgb(222, 222, 222);\n"
"")
        self.pushButton_34.setObjectName("pushButton_34")
        self.label_36 = QtWidgets.QLabel(self.frame_6)
        self.label_36.setGeometry(QtCore.QRect(650, 10, 91, 31))
        self.label_36.setStyleSheet("QLabel{background-color: rgb(255, 189, 167);}")
        self.label_36.setAlignment(QtCore.Qt.AlignCenter)
        self.label_36.setObjectName("label_36")
        self.pushButton_26 = QtWidgets.QPushButton(self.frame_4)
        self.pushButton_26.clicked.connect(self.export)
        self.pushButton_26.setGeometry(QtCore.QRect(1150, 350, 411, 71))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_26.setFont(font)
        self.pushButton_26.setStyleSheet("background-color: #73C067;color: white;")
        self.pushButton_26.setObjectName("pushButton_26")
        self.label_35 = QtWidgets.QLabel(self.frame_4)
        self.label_35.setGeometry(QtCore.QRect(1150, 430, 391, 31))
        font = QtGui.QFont()
        font.setFamily("MS Sans Serif")
        font.setPointSize(8)
        font.setBold(False)
        font.setItalic(True)
        font.setWeight(50)
        self.label_35.setFont(font)
        self.label_35.setStyleSheet("QLabel{color : red;}")
        self.label_35.setObjectName("label_35")
        self.pushButton_35 = QtWidgets.QPushButton(self.frame_4)
        self.pushButton_35.clicked.connect(self.reset)
        self.pushButton_35.setGeometry(QtCore.QRect(1150, 530, 411, 71))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_35.setFont(font)
        self.pushButton_35.setStyleSheet("background-color: #B00020;color: white;")
        self.pushButton_35.setObjectName("pushButton_35")
        self.label_8 = QtWidgets.QLabel(self.frame_4)
        self.label_8.setGeometry(QtCore.QRect(300, 20, 41, 31))
        self.label_8.setStyleSheet("QLabel{background-color: rgb(255, 189, 167);}")
        self.label_8.setAlignment(QtCore.Qt.AlignCenter)
        self.label_8.setObjectName("label_8")
        self.label_9 = QtWidgets.QLabel(self.frame_4)
        self.label_9.setGeometry(QtCore.QRect(1020, 130, 41, 31))
        self.label_9.setStyleSheet("QLabel{background-color: rgb(255, 189, 167);}")
        self.label_9.setAlignment(QtCore.Qt.AlignCenter)
        self.label_9.setObjectName("label_9")
        self.label_10 = QtWidgets.QLabel(self.frame_4)
        self.label_10.setGeometry(QtCore.QRect(1520, 310, 41, 31))
        self.label_10.setStyleSheet("QLabel{background-color: rgb(255, 189, 167);}")
        self.label_10.setAlignment(QtCore.Qt.AlignCenter)
        self.label_10.setObjectName("label_10")
        self.label_11 = QtWidgets.QLabel(self.frame_4)
        self.label_11.setGeometry(QtCore.QRect(1520, 490, 41, 31))
        self.label_11.setStyleSheet("QLabel{background-color: rgb(255, 189, 167);}")
        self.label_11.setAlignment(QtCore.Qt.AlignCenter)
        self.label_11.setObjectName("label_11")
        self.pushButton_39 = QtWidgets.QPushButton(self.frame_4)
        self.pushButton_39.clicked.connect(self.data_extraction)
        self.pushButton_39.setGeometry(QtCore.QRect(1330, 10, 231, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_39.setFont(font)
        self.pushButton_39.setStyleSheet("background-color: #F47B1F;\n"
"color: white;\n"
"")
        self.pushButton_39.setObjectName("pushButton_39")
        self.label_37 = QtWidgets.QLabel(self.frame_4)
        self.label_37.setGeometry(QtCore.QRect(1430, 640, 131, 21))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_37.setFont(font)
        self.label_37.setStyleSheet("QLabel{background-color: rgb(255, 189, 167);}")
        self.label_37.setAlignment(QtCore.Qt.AlignCenter)
        self.label_37.setObjectName("label_37")
        TabWidget.addTab(self.tab1, "")

        self.retranslateUi(TabWidget)
        TabWidget.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(TabWidget)

    def retranslateUi(self, TabWidget):
        _translate = QtCore.QCoreApplication.translate
        TabWidget.setWindowTitle(_translate("TabWidget", "Data Validation Tool"))
        self.label_38.setText(_translate("TabWidget", "** Mandatory"))
        self.label_37.setText(_translate("TabWidget", "** अनिवार्य है"))
        self.label.setText(_translate("TabWidget", "Data Validation Tool"))
        self.pushButton_6.setText(_translate("TabWidget", "Upload"))
        self.lineEdit_4.setPlaceholderText(_translate("TabWidget", "  Your uploaded file name will display here ..."))
        self.label_21.setText(_translate("TabWidget", "* Upload data in .xls / .xlsx format for one month and one facility type only."))
        self.lineEdit_5.setPlaceholderText(_translate("TabWidget", "  Facility Type selected will display here ..."))
        self.lineEdit_6.setPlaceholderText(_translate("TabWidget", " Month, Year"))
        self.pushButton_12.setText(_translate("TabWidget", "Validate"))
        self.pushButton_19.setText(_translate("TabWidget", "Methodology"))
        self.label_22.setText(_translate("TabWidget", "* Press Validate button to perform validation check on your data"))
        self.pushButton_20.setText(_translate("TabWidget", "User Manual"))
        self.pushButton_13.setText(_translate("TabWidget", "-- All Selected --"))
        self.pushButton_14.setText(_translate("TabWidget", "-- All Selected --"))
        self.pushButton_15.setText(_translate("TabWidget", "-- All Selected --"))
        self.label_12.setText(_translate("TabWidget", "District"))
        self.label_13.setText(_translate("TabWidget", "Sub-district"))
        self.label_14.setText(_translate("TabWidget", "Block"))
        self.label_15.setText(_translate("TabWidget", "Health Block"))
        self.label_16.setText(_translate("TabWidget", "Select Filters"))
        self.label_17.setText(_translate("TabWidget", "State"))
        self.pushButton_16.setText(_translate("TabWidget", "-- All Selected --"))
        self.label_18.setText(_translate("TabWidget", "Rural / Urban"))
        self.label_20.setText(_translate("TabWidget", "Ownership"))
        self.label_24.setText(_translate("TabWidget", "Facility Name"))
        self.pushButton_27.setText(_translate("TabWidget", "-- All Selected --"))
        self.pushButton_28.setText(_translate("TabWidget", "-- All Selected --"))
        self.pushButton_29.setText(_translate("TabWidget", "-- All Selected --"))
        self.pushButton_30.setText(_translate("TabWidget", "-- All Selected --"))
        self.label_5.setText(_translate("TabWidget", "3 (optional)"))
        self.pushButton_17.setText(_translate("TabWidget", "Download File"))
        self.label_25.setText(_translate("TabWidget", "* Press Download button to download your validated data."))
        self.pushButton_18.setText(_translate("TabWidget", "Reset"))
        self.label_3.setText(_translate("TabWidget", "1**"))
        self.label_4.setText(_translate("TabWidget", "2**"))
        self.label_6.setText(_translate("TabWidget", "4**"))
        self.label_7.setText(_translate("TabWidget", "5**"))
        self.pushButton_38.setText(_translate("TabWidget", "Data Extraction"))
        TabWidget.setTabText(TabWidget.indexOf(self.tab), _translate("TabWidget", "English Version of Tool"))
        self.label_2.setText(_translate("TabWidget", "डाटा वेलिडेशन टूल"))
        self.pushButton_7.setText(_translate("TabWidget", "अपलोड करें"))
        self.lineEdit_7.setPlaceholderText(_translate("TabWidget", "आपकी अपलोड की गई फ़ाइल का नाम यहां प्रदर्शित होगा..."))
        self.label_33.setText(_translate("TabWidget", "* केवल एक प्रकार के स्वास्थ्य केंद्र और केवल एक महीने  के लिए .xls / .xlsx प्रारूप में डेटा अपलोड करें।"))
        self.pushButton_36.setText(_translate("TabWidget", "प्रक्रिया"))
        self.pushButton_37.setText(_translate("TabWidget", "मैन्युअल"))
        self.lineEdit_8.setPlaceholderText(_translate("TabWidget", "चयनित स्वास्थ्य केंद्र प्रकार यहां प्रदर्शित होगा..."))
        self.lineEdit_9.setPlaceholderText(_translate("TabWidget", "महीना,  वर्ष"))
        self.pushButton_21.setText(_translate("TabWidget", "डेटा मान्य करें"))
        self.label_34.setText(_translate("TabWidget", "* अपने डेटा पर सत्यापन जांच करने के लिए मान्य बटन दबाएं"))
        self.pushButton_22.setText(_translate("TabWidget", "-- सभी चयनित --"))
        self.pushButton_23.setText(_translate("TabWidget", "-- सभी चयनित --"))
        self.pushButton_24.setText(_translate("TabWidget", "-- सभी चयनित --"))
        self.label_19.setText(_translate("TabWidget", "जिला"))
        self.label_23.setText(_translate("TabWidget", "उप - जिला"))
        self.label_26.setText(_translate("TabWidget", "ब्लॉक"))
        self.label_27.setText(_translate("TabWidget", "हेल्थ ब्लॉक"))
        self.label_28.setText(_translate("TabWidget", "फ़िल्टर का चयन करें"))
        self.label_29.setText(_translate("TabWidget", "राज्य"))
        self.pushButton_25.setText(_translate("TabWidget", "-- सभी चयनित --"))
        self.label_30.setText(_translate("TabWidget", "शहरी / ग्रामीण"))
        self.label_31.setText(_translate("TabWidget", "ओनरशिप"))
        self.label_32.setText(_translate("TabWidget", "स्वास्थ्य केंद्र का नाम"))
        self.pushButton_31.setText(_translate("TabWidget", "-- सभी चयनित --"))
        self.pushButton_32.setText(_translate("TabWidget", "-- सभी चयनित --"))
        self.pushButton_33.setText(_translate("TabWidget", "-- सभी चयनित --"))
        self.pushButton_34.setText(_translate("TabWidget", "-- सभी चयनित --"))
        self.label_36.setText(_translate("TabWidget", "3 (वैकल्पिक)"))
        self.pushButton_26.setText(_translate("TabWidget", "डाउनलोड फाइल"))
        self.label_35.setText(_translate("TabWidget", "* अपने मान्य डेटा को निर्यात करने के लिए निर्यात बटन दबाएं।"))
        self.pushButton_35.setText(_translate("TabWidget", "डेटा रीसेट करें"))
        self.label_8.setText(_translate("TabWidget", "1**"))
        self.label_9.setText(_translate("TabWidget", "2**"))
        self.label_10.setText(_translate("TabWidget", "4**"))
        self.label_11.setText(_translate("TabWidget", "5**"))
        self.pushButton_39.setText(_translate("TabWidget", "डेटा एक्सट्रैक्शन"))
        TabWidget.setTabText(TabWidget.indexOf(self.tab1), _translate("TabWidget", "Hindi Version of Tool"))


    ''' Upload Function '''
    # =====================

    def upload(self):
        global df_, res_dict

        # Validation for uploaded valid excel file
        try:
            # Upload file by opening filedialog
            fileName,_ = QFileDialog.getOpenFileName(TabWidget, "Open Excel", (QtCore.QDir.homePath()), "Excel (*.xls *.xlsx)")
            
            # Read uploaded excel file
            df_ = pd.read_excel(fileName)

            # Converted again to csv file
            df_.to_csv("FileName.csv")

            # Read converted csv file
            df_ = pd.read_csv("FileName.csv")
        except:
            msg = QMessageBox()
            msg.setWindowTitle("Uploaded File Error Message / अपलोड की गई फ़ाइल त्रुटि संदेश")
            msg.setIcon(QMessageBox.Critical)
            msg.setText(
                "The file which you have uploaded is not in the valid format of excel, Please upload valid excel file \n\n आपके द्वारा अपलोड की गई फ़ाइल एक्सेल के मान्य प्रारूप में नहीं है, कृपया मान्य एक्सेल फ़ाइल अपलोड करें")
            msg.exec()

            try:
                # Upload file by opening filedialog
                fileName, _ = QFileDialog.getOpenFileName(
                    TabWidget, "Open Excel", (QtCore.QDir.homePath()), "Excel (*.xls *.xlsx)")

                # Read uploaded excel file
                df_ = pd.read_excel(fileName)

                # Converted again to csv file
                df_.to_csv("FileName.csv")

                # Read converted csv file
                df_ = pd.read_csv("FileName.csv")
            except:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Critical)
                msg.setWindowTitle("Uploaded File Error Message / अपलोड की गई फ़ाइल त्रुटि संदेश")
                msg.setText("Please upload valid excel file, Try again! \n\n कृपया मान्य एक्सेल फ़ाइल अपलोड करें, पुनः प्रयास करें!")
                msg.exec()

        self.popup.start_progress()

        self.lineEdit_4.setText(fileName)
        self.lineEdit_7.setText(fileName)

        # Dropping last two rows
        df_.drop(df_.index[[-1, -2]], inplace=True)

        # Extracting string from 1st cell of dataframe
        str_to_extr_MonthYear = str(df_.iloc[0])
        print(str_to_extr_MonthYear)

        # grab the first row for the header
        new_header = df_.iloc[1]

        # #take the data less the header row
        df_ = df_[1:]

        # set the header row as the df header
        df_.columns = new_header

        # Extracting Month , Year from string
        results = re.findall(
            r"[abceglnoprtuvyADFJMNOS|]{3}[\s-]\d{2,4}", str_to_extr_MonthYear)
        print(results)

        # Splitting Month and Year
        MYList = results[0].split('-')

        # Partial list of headers
        lst1 = df_.columns[:21].values

        # Picking row items after 18th row to merge with lst1
        lst2 = df_.iloc[1, 21:].to_numpy()

        # Merging both lists
        lst3 = np.concatenate((lst1, lst2))

        # Assign lst3 as new column header
        df_.columns = lst3

        # Taking DataFrame from second row
        df_ = df_[2:]

        # Insering Month and Year to the orignal dataframe
        df_.insert(1, 'Month', MYList[0])
        df_.insert(2, 'Year', MYList[1])

        # Removing A column named as # coming from orignal data
        df_ = df_.loc[:, df_.columns != '#']

        # Reindexing dataframe
        df_ = df_.reset_index(drop=True)

        df_ = df_.iloc[:, 1:]

        # Temporary column to verify modified checks
        temp_columns = ['col_' + str(index)
                        for index in range(1, len(df_.columns)+1)]

        # Merging and converting temp_columns to orignal header to dictionary
        res_dict = {temp_columns[i]: df_.columns[i]
                    for i in range(len(temp_columns))}

        # Picking the temporary column names and renaming column headers with it
        df_.columns = [i for i in res_dict.keys()]

        # Orignal Header
        df_OrgHeaders = [i for i in res_dict.values()]

        # convert the set to the list and fill inside comboBox to select facility type
        list_set = df_['col_12'].tolist()
        unique_list = set(list_set)

        self.lineEdit_5.setText(["{0}".format(col) for col in unique_list][0])
        self.lineEdit_6.setText(results[0])
        self.lineEdit_8.setText(["{0}".format(col) for col in unique_list][0])
        self.lineEdit_9.setText(results[0])

        # Disabling upload Button
        self.pushButton_6.setDisabled(True)
        self.pushButton_7.setDisabled(True)

        self.popup.close()

        # Create the messagebox object
        self.msg = QMessageBox()
        # Set the information icon
        self.msg.setWindowIcon(QtGui.QIcon('checked.png'))
        self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
        # Set the main message
        self.msg.setText("The file has been uploaded successfully. \n\n फ़ाइल सफलतापूर्वक अपलोड कर दी गई है।")
        # Set the title of the window
        self.msg.setWindowTitle(" ")
        # Display the message box
        self.msg.show()
        return df_


    # Upload file button functionality
    # ================================
    def loadFile(self, df_):
        return df_


    # Filtering Facility Type
    # =======================
    def VerifyFType(self):
        global df, FType
        self.popup.start_progress()

        FType = self.lineEdit_5.text()

        if (FType == 'Primary Health Centre'):
            # Signaling PHC_Validate function i.e function where validation checks are present
            df = PHC_Validate(self, df_)

        elif (FType == 'Health Sub Centre' ):
            # Signaling HSC_Validate function i.e function where validation checks are present
            df = HSC_Validate(self, df_)

        elif (FType == 'District Hospital'):
            df = DH_Validate(self, df_)

        elif (FType == 'Sub District Hospital'):
            df = SDH_Validate(self, df_)

        elif (FType == 'Community Health Centre'):
            df = CHC_Validate(self, df_)

        else:
            raise Exception('Facility Type Name is not matching')

        self.pushButton_12.setEnabled(False)
        self.pushButton_21.setEnabled(False)
        self.popup.close()

    '''
    # Filter to decide which filter button user clicked
    # =================================================
    '''
    def eventFilter(self, target, event):
        if target == self.pushButton_16 and event.type() == QtCore.QEvent.MouseButtonPress:
            self.pushButton_16.clicked.connect(self.onSelectState)
            return True

        elif target == self.pushButton_13 and event.type() == QtCore.QEvent.MouseButtonPress:
            self.pushButton_13.clicked.connect(self.onSelectDistrict)
            return True

        elif target == self.pushButton_27 and event.type() == QtCore.QEvent.MouseButtonPress:
            self.pushButton_27.clicked.connect(self.onSelectFacilityName)
            return True

        elif target == self.pushButton_29 and event.type() == QtCore.QEvent.MouseButtonPress:
            self.pushButton_29.clicked.connect(self.onSelectRuralUrban)
            return True

        elif target == self.pushButton_28 and event.type() == QtCore.QEvent.MouseButtonPress:
            self.pushButton_28.clicked.connect(self.onSelectOwnership)
            return True

        elif target == self.pushButton_14 and event.type() == QtCore.QEvent.MouseButtonPress:
            self.pushButton_14.clicked.connect(self.onSelectSubDistrict)
            return True

        elif target == self.pushButton_15 and event.type() == QtCore.QEvent.MouseButtonPress:
            self.pushButton_15.clicked.connect(self.onSelectBlock)
            return True

        elif target == self.pushButton_30 and event.type() == QtCore.QEvent.MouseButtonPress:
            self.pushButton_30.clicked.connect(self.onSelectHealthBlock)
            return True

        return False


    ################################################################################
    # Filter State Functionality
    
    def onSelectState(self, index):
        self.keywords = dict([(i, []) for i in range(df.shape[0])])
        print(self.keywords)
        self.menu = QtWidgets.QMenu(TabWidget)
        self.menu.setStyleSheet('QMenu { menu-scrollable: true; width: 200 }')
        font = self.menu.font()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.menu.setFont(font)

        index = 3
        self.col = index

        data_unique = []

        self.checkBoxs = []

        # Selectall added into Dropdown
        checkBox = QtWidgets.QCheckBox("Select all / Deselect all", self.menu)
        checkBox.setStyleSheet("color: red; spacing: 5px; font-size:15px;")
        # All the checkboxes are enabled to check
        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(checkBox)
        self.menu.addAction(checkableAction)
        checkBox.setChecked(True)
        checkBox.stateChanged.connect(self.slotSelectState)

        # list storing state data
        df['col_3'].fillna('Blank', inplace = True)
        list_set = df['col_3'].to_list()

        item = list_set
        item = sorted(list_set, key=str.upper)

        # looping to fill checkboxes, initially all checkboxes will be checked
        for i in range(len(item)):
            if item[i] not in data_unique:
                data_unique.append(item[i])
                checkBox = QtWidgets.QCheckBox(item[i], self.menu)
                checkBox.setChecked(True)
                checkableAction = QtWidgets.QWidgetAction(self.menu)
                checkableAction.setDefaultWidget(checkBox)
                self.menu.addAction(checkableAction)
                self.checkBoxs.append(checkBox)

        # Ok, cancel button
        btn = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel,
                                            QtCore.Qt.Vertical, self.menu)

        # ok selected
        btn.accepted.connect(self.menuCloseState)
        # rejected , nothing selected
        btn.rejected.connect(self.menu.close)

        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(btn)
        self.menu.addAction(checkableAction)
        self.pushButton_16.setMenu(self.menu)
        self.pushButton_25.setMenu(self.menu)

    # method to check -> uncheck and vice versa
    def slotSelectState(self, state):
        for checkbox in self.checkBoxs:
            checkbox.setChecked(QtCore.Qt.Checked == state)

    # after ok selected
    def menuCloseState(self):
        self.keywords[self.col] = []
        for element in self.checkBoxs:
            if element.isChecked():
                self.keywords[self.col].append(element.text())
        
        self.filterdataState()
        self.pushButton_16.setText('Selected')
        self.pushButton_16.setEnabled(False)
        self.pushButton_25.setText('चयनित')
        self.pushButton_25.setEnabled(False)
        self.menu.close()

    # Filter data columnwise
    def filterdataState(self):
        global df
        #keywords = dict([(i, []) for i in range(self.filterall.columnCount())])
        columnsShow = dict([(i, True) for i in range(df['col_3'].shape[0])])

        j = 0
        for j in range(df['col_3'].shape[0]):
            item = df['col_3'].to_list()

            # if self.keywords[self.col]:
            if item[j] not in self.keywords[self.col]:
                columnsShow[j] = False

        # for key, value in columnsShow.items():
        final_lst = [i for i in columnsShow.values()]
        df = df[final_lst]
        return df



    ################################################################################
    # Filter District Functionality

    def onSelectDistrict(self, index):
        self.keywords = dict([(i, []) for i in range(df.shape[0])])
        print(self.keywords)
        self.menu = QtWidgets.QMenu(TabWidget)
        self.menu.setStyleSheet('QMenu { menu-scrollable: true; width: 300 }')
        font = self.menu.font()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.menu.setFont(font)

        index = 5
        self.col = index

        data_unique = []

        self.checkBoxs = []

        # Selectall/deselectall added into Dropdown
        checkBox = QtWidgets.QCheckBox("Select all / Deselect all", self.menu)
        checkBox.setStyleSheet("color: red; spacing: 5px; font-size:15px;")
        # All the checkboxes are enabled to check
        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(checkBox)
        self.menu.addAction(checkableAction)
        checkBox.setChecked(True)
        checkBox.stateChanged.connect(self.slotSelectDistrict)

        # list storing state data
        df['col_5'].fillna('Blank', inplace = True)
        list_set = df['col_5'].to_list()

        item = list_set
        item = sorted(list_set, key=str.upper)

        # looping to fill checkboxes, initially all checkboxes will be checked
        for i in range(len(item)):
            if item[i] not in data_unique:
                data_unique.append(item[i])
                checkBox = QtWidgets.QCheckBox(item[i], self.menu)
                checkBox.setChecked(True)
                checkableAction = QtWidgets.QWidgetAction(self.menu)
                checkableAction.setDefaultWidget(checkBox)
                self.menu.addAction(checkableAction)
                self.checkBoxs.append(checkBox)

        # Ok, cancel button
        btn = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel,
                                            QtCore.Qt.Vertical, self.menu)

        # ok selected
        btn.accepted.connect(self.menuCloseDistrict)
        # rejected , nothing selected
        btn.rejected.connect(self.menu.close)

        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(btn)
        self.menu.addAction(checkableAction)
        self.pushButton_13.setMenu(self.menu)
        self.pushButton_22.setMenu(self.menu)

    # method to check -> uncheck and vice versa
    def slotSelectDistrict(self, state):
        for checkbox in self.checkBoxs:
            checkbox.setChecked(QtCore.Qt.Checked == state)


    # after ok selected
    def menuCloseDistrict(self):
        self.keywords[self.col] = []
        for element in self.checkBoxs:
            if element.isChecked():
                self.keywords[self.col].append(element.text())
        print(self.keywords[self.col])
        self.filterdataDistrict()
        self.pushButton_13.setText('Selected')
        self.pushButton_13.setEnabled(False)
        self.pushButton_22.setText('चयनित')
        self.pushButton_22.setEnabled(False)
        self.menu.close()

    # Filter data columnwise
    def filterdataDistrict(self):
        global df
        #keywords = dict([(i, []) for i in range(self.filterall.columnCount())])
        columnsShow = dict([(i, True) for i in range(df['col_5'].shape[0])])
        print(columnsShow)

        j = 0
        for j in range(df['col_5'].shape[0]):
            item = df['col_5'].to_list()

            # if self.keywords[self.col]:
            if item[j] not in self.keywords[self.col]:
                columnsShow[j] = False

        # for key, value in columnsShow.items():
        final_lst = [i for i in columnsShow.values()]
        df = df[final_lst]
        return df


    ################################################################################
    # Filter Sub District Functionality

    def onSelectSubDistrict(self, index):
        self.keywords = dict([(i, []) for i in range(df.shape[0])])
        print(self.keywords)
        self.menu = QtWidgets.QMenu(TabWidget)
        self.menu.setStyleSheet('QMenu { menu-scrollable: true; width: 300 }')
        font = self.menu.font()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.menu.setFont(font)

        index = 7
        self.col = index

        data_unique = []

        self.checkBoxs = []

        # Selectall added into Dropdown
        checkBox = QtWidgets.QCheckBox("Select all / Deselect all", self.menu)
        checkBox.setStyleSheet("color: red; spacing: 5px; font-size:15px;")
        # All the checkboxes are enabled to check
        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(checkBox)
        self.menu.addAction(checkableAction)
        checkBox.setChecked(True)
        checkBox.stateChanged.connect(self.slotSelectSubDistrict)

        # list storing state data
        df['col_7'].fillna('Blank', inplace = True)
        list_set = df['col_7'].to_list()

        item = list_set
        item = sorted(list_set, key=str.upper)

        # looping to fill checkboxes, initially all checkboxes will be checked
        for i in range(len(item)):
            if item[i] not in data_unique:
                data_unique.append(item[i])
                checkBox = QtWidgets.QCheckBox(item[i], self.menu)
                checkBox.setChecked(True)
                checkableAction = QtWidgets.QWidgetAction(self.menu)
                checkableAction.setDefaultWidget(checkBox)
                self.menu.addAction(checkableAction)
                self.checkBoxs.append(checkBox)

        # Ok, cancel button
        btn = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel,
                                            QtCore.Qt.Vertical, self.menu)

        # ok selected
        btn.accepted.connect(self.menuCloseSubDistrict)
        # rejected , nothing selected
        btn.rejected.connect(self.menu.close)

        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(btn)
        self.menu.addAction(checkableAction)
        self.pushButton_23.setMenu(self.menu)
        self.pushButton_14.setMenu(self.menu)

    # method to check -> uncheck and vice versa
    def slotSelectSubDistrict(self, state):
        for checkbox in self.checkBoxs:
            checkbox.setChecked(QtCore.Qt.Checked == state)

    # after ok selected
    def menuCloseSubDistrict(self):
        self.keywords[self.col] = []
        for element in self.checkBoxs:
            if element.isChecked():
                self.keywords[self.col].append(element.text())
        print(self.keywords[self.col])
        self.filterdataSubDistrict()
        self.pushButton_14.setText('Selected')
        self.pushButton_14.setEnabled(False)
        self.pushButton_23.setText('चयनित')
        self.pushButton_23.setEnabled(False)
        self.menu.close()

    # Filter data columnwise
    def filterdataSubDistrict(self):
        global df
        #keywords = dict([(i, []) for i in range(self.filterall.columnCount())])
        columnsShow = dict([(i, True) for i in range(df['col_7'].shape[0])])
        print(columnsShow)

        j = 0
        for j in range(df['col_7'].shape[0]):
            item = df['col_7'].to_list()

            # if self.keywords[self.col]:
            if item[j] not in self.keywords[self.col]:
                columnsShow[j] = False

        # for key, value in columnsShow.items():
        final_lst = [i for i in columnsShow.values()]
        df = df[final_lst]
        return df


    ################################################################################
    # Filter Block Functionality

    def onSelectBlock(self, index):
        self.keywords = dict([(i, []) for i in range(df.shape[0])])
        print(self.keywords)
        self.menu = QtWidgets.QMenu(TabWidget)
        self.menu.setStyleSheet('QMenu { menu-scrollable: true; width: 200 }')
        font = self.menu.font()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.menu.setFont(font)

        index = 8
        self.col = index

        data_unique = []

        self.checkBoxs = []

        # Selectall added into Dropdown
        checkBox = QtWidgets.QCheckBox("Select all / Deselect all", self.menu)
        checkBox.setStyleSheet("color: red; spacing: 5px; font-size:15px;")
        # All the checkboxes are enabled to check
        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(checkBox)
        self.menu.addAction(checkableAction)
        checkBox.setChecked(True)
        checkBox.stateChanged.connect(self.slotSelectBlock)

        # list storing state data
        df['col_8'].fillna('Blank', inplace = True)
        list_set = df['col_8'].to_list()

        item = list_set
        item = sorted(list_set, key=str.upper)

        # looping to fill checkboxes, initially all checkboxes will be checked
        for i in range(len(item)):
            if item[i] not in data_unique:
                data_unique.append(item[i])
                checkBox = QtWidgets.QCheckBox(item[i], self.menu)
                checkBox.setChecked(True)
                checkableAction = QtWidgets.QWidgetAction(self.menu)
                checkableAction.setDefaultWidget(checkBox)
                self.menu.addAction(checkableAction)
                self.checkBoxs.append(checkBox)

        # Ok, cancel button
        btn = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel,
                                            QtCore.Qt.Vertical, self.menu)

        # ok selected
        btn.accepted.connect(self.menuCloseBlock)
        # rejected , nothing selected
        btn.rejected.connect(self.menu.close)

        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(btn)
        self.menu.addAction(checkableAction)
        self.pushButton_15.setMenu(self.menu)
        self.pushButton_24.setMenu(self.menu)

    # method to check -> uncheck and vice versa
    def slotSelectBlock(self, state):
        for checkbox in self.checkBoxs:
            checkbox.setChecked(QtCore.Qt.Checked == state)

    # after ok selected
    def menuCloseBlock(self):
        self.keywords[self.col] = []
        for element in self.checkBoxs:
            if element.isChecked():
                self.keywords[self.col].append(element.text())
        print(self.keywords[self.col])
        self.filterdataBlock()
        self.pushButton_15.setText('Selected')
        self.pushButton_15.setEnabled(False)
        self.pushButton_24.setText('चयनित')
        self.pushButton_24.setEnabled(False)
        self.menu.close()

    # Filter data columnwise
    def filterdataBlock(self):
        global df
        #keywords = dict([(i, []) for i in range(self.filterall.columnCount())])
        columnsShow = dict([(i, True) for i in range(df['col_8'].shape[0])])
        print(columnsShow)

        j = 0
        for j in range(df['col_8'].shape[0]):
            item = df['col_8'].to_list()

            # if self.keywords[self.col]:
            if item[j] not in self.keywords[self.col]:
                columnsShow[j] = False

        # for key, value in columnsShow.items():
        final_lst = [i for i in columnsShow.values()]
        df = df[final_lst]
        return df


    ################################################################################
    # Filter Health Block Functionality

    def onSelectHealthBlock(self, index):
        self.keywords = dict([(i, []) for i in range(df.shape[0])])
        print(self.keywords)
        self.menu = QtWidgets.QMenu(TabWidget)
        self.menu.setStyleSheet('QMenu { menu-scrollable: true; width: 200 }')
        font = self.menu.font()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.menu.setFont(font)

        index = 9
        self.col = index

        data_unique = []

        self.checkBoxs = []

        # Selectall added into Dropdown
        checkBox = QtWidgets.QCheckBox("Select all / Deselect all", self.menu)
        checkBox.setStyleSheet("color: red; spacing: 5px; font-size:15px;")
        # All the checkboxes are enabled to check
        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(checkBox)
        self.menu.addAction(checkableAction)
        checkBox.setChecked(True)
        checkBox.stateChanged.connect(self.slotSelectHealthBlock)

        # list storing state data
        df['col_9'].fillna('Blank', inplace = True)
        list_set = df['col_9'].to_list()

        item = list_set
        item = sorted(list_set, key=str.upper)

        # looping to fill checkboxes, initially all checkboxes will be checked
        for i in range(len(item)):
            if item[i] not in data_unique:
                data_unique.append(item[i])
                checkBox = QtWidgets.QCheckBox(item[i], self.menu)
                checkBox.setChecked(True)
                checkableAction = QtWidgets.QWidgetAction(self.menu)
                checkableAction.setDefaultWidget(checkBox)
                self.menu.addAction(checkableAction)
                self.checkBoxs.append(checkBox)

        # Ok, cancel button
        btn = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel,
                                            QtCore.Qt.Vertical, self.menu)

        # ok selected
        btn.accepted.connect(self.menuCloseHealthBlock)
        # rejected , nothing selected
        btn.rejected.connect(self.menu.close)

        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(btn)
        self.menu.addAction(checkableAction)
        self.pushButton_30.setMenu(self.menu)
        self.pushButton_34.setMenu(self.menu)

    # method to check -> uncheck and vice versa
    def slotSelectHealthBlock(self, state):
        for checkbox in self.checkBoxs:
            checkbox.setChecked(QtCore.Qt.Checked == state)

    # after ok selected
    def menuCloseHealthBlock(self):
        self.keywords[self.col] = []
        for element in self.checkBoxs:
            if element.isChecked():
                self.keywords[self.col].append(element.text())
        print(self.keywords[self.col])
        self.filterdataHealthBlock()
        self.pushButton_30.setText('Selected')
        self.pushButton_30.setEnabled(False)
        self.pushButton_34.setText('चयनित')
        self.pushButton_34.setEnabled(False)
        self.menu.close()

    # Filter data columnwise
    def filterdataHealthBlock(self):
        global df
        #keywords = dict([(i, []) for i in range(self.filterall.columnCount())])
        columnsShow = dict([(i, True) for i in range(df['col_9'].shape[0])])
        print(columnsShow)

        j = 0
        for j in range(df['col_9'].shape[0]):
            item = df['col_9'].to_list()

            # if self.keywords[self.col]:
            if item[j] not in self.keywords[self.col]:
                columnsShow[j] = False

        # for key, value in columnsShow.items():
        final_lst = [i for i in columnsShow.values()]
        df = df[final_lst]
        return df



    ################################################################################
    # Filter Facility Name

    # Filter FacilityName Functionality

    def onSelectFacilityName(self, index):
        self.keywords = dict([(i, []) for i in range(df.shape[0])])
        print(self.keywords)
        self.menu = QtWidgets.QMenu(TabWidget)
        self.menu.setStyleSheet('QMenu { menu-scrollable: 1; width: 400 }')
        font = self.menu.font()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.menu.setFont(font)

        index = 14
        self.col = index

        data_unique = []

        self.checkBoxs = []

        # Selectall added into Dropdown
        checkBox = QtWidgets.QCheckBox("Select all/ Deselect all", self.menu)
        checkBox.setStyleSheet("color: red; spacing: 5px; font-size:15px;")
        # All the checkboxes are enabled to check
        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(checkBox)
        self.menu.addAction(checkableAction)
        checkBox.setChecked(True)
        checkBox.stateChanged.connect(self.slotSelectFacilityName)

        # list storing Facility Name data
        df['col_14'].fillna('Blank', inplace = True)
        list_set = df['col_14'].to_list()

        item = sorted(list_set, key=str.upper)

        # looping to fill checkboxes, initially all checkboxes will be checked
        for i in range(len(item)):
            if item[i] not in data_unique:
                data_unique.append(item[i])
                checkBox = QtWidgets.QCheckBox(item[i], self.menu)
                checkBox.setChecked(True)
                checkableAction = QtWidgets.QWidgetAction(self.menu)
                checkableAction.setDefaultWidget(checkBox)
                self.menu.addAction(checkableAction)
                self.checkBoxs.append(checkBox)

        # Ok, cancel button
        btn = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel,
                                            QtCore.Qt.Vertical, self.menu)

        # ok selected
        btn.accepted.connect(self.menuCloseFacilityName)
        # rejected , nothing selected
        btn.rejected.connect(self.menu.close)

        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(btn)
        self.menu.addAction(checkableAction)

        ############# Always set Pushbutton ####################
        self.pushButton_27.setMenu(self.menu)
        self.pushButton_31.setMenu(self.menu)

    # method to check -> uncheck and vice versa
    def slotSelectFacilityName(self, state):
        for checkbox in self.checkBoxs:
            checkbox.setChecked(QtCore.Qt.Checked == state)

    # after ok selected
    def menuCloseFacilityName(self):
        self.keywords[self.col] = []
        for element in self.checkBoxs:
            if element.isChecked():
                self.keywords[self.col].append(element.text())
        print(self.keywords[self.col])
        self.filterdataFacilityName()
        self.pushButton_27.setText('Selected')
        self.pushButton_27.setEnabled(False)
        self.pushButton_31.setText('चयनित')
        self.pushButton_31.setEnabled(False)
        self.menu.close()

    # Filter data columnwise
    def filterdataFacilityName(self):
        global df
        #keywords = dict([(i, []) for i in range(self.filterall.columnCount())])
        columnsShow = dict([(i, True) for i in range(df['col_14'].shape[0])])
        print(columnsShow)

        j = 0
        for j in range(df['col_14'].shape[0]):
            item = df['col_14'].to_list()

            # if self.keywords[self.col]:
            if item[j] not in self.keywords[self.col]:
                columnsShow[j] = False

        # for key, value in columnsShow.items():
        final_lst = [i for i in columnsShow.values()]
        print(final_lst, 'this is final list of Select District')

        # matching list of facility type with col of dataframe returned by onSelectDistrict fun
        df = df[final_lst]
        return df


    ################################################################################
    # Filter Rural/Urban

    def onSelectRuralUrban(self, index):
        self.keywords = dict([(i, []) for i in range(df.shape[0])])
        print(self.keywords)
        self.menu = QtWidgets.QMenu(TabWidget)
        self.menu.setStyleSheet('QMenu { menu-scrollable: true; width: 300 }')
        font = self.menu.font()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.menu.setFont(font)

        index = 18
        self.col = index

        data_unique = []

        self.checkBoxs = []

        # Selectall added into Dropdown
        checkBox = QtWidgets.QCheckBox("Select all/ Deselect all", self.menu)
        checkBox.setStyleSheet("color: red; spacing: 5px; font-size:15px;")
        # All the checkboxes are enabled to check
        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(checkBox)
        self.menu.addAction(checkableAction)
        checkBox.setChecked(True)
        checkBox.stateChanged.connect(self.slotSelectRuralUrban)

        # list storing Facility Name data
        df['col_18'].fillna('Blank', inplace = True)
        list_set = df['col_18'].to_list()

        item = list_set
        item = sorted(list_set, key=str.upper)

        # looping to fill checkboxes, initially all checkboxes will be checked
        for i in range(len(item)):
            if item[i] not in data_unique:
                data_unique.append(item[i])
                checkBox = QtWidgets.QCheckBox(item[i], self.menu)
                checkBox.setChecked(True)
                checkableAction = QtWidgets.QWidgetAction(self.menu)
                checkableAction.setDefaultWidget(checkBox)
                self.menu.addAction(checkableAction)
                self.checkBoxs.append(checkBox)

        # Ok, cancel button
        btn = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel,
                                            QtCore.Qt.Vertical, self.menu)

        # ok selected
        btn.accepted.connect(self.menuCloseRuralUrban)
        # rejected , nothing selected
        btn.rejected.connect(self.menu.close)

        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(btn)
        self.menu.addAction(checkableAction)

        ############# Always set Pushbutton ####################
        self.pushButton_29.setMenu(self.menu)
        self.pushButton_33.setMenu(self.menu)

    # method to check -> uncheck and vice versa
    def slotSelectRuralUrban(self, state):
        for checkbox in self.checkBoxs:
            checkbox.setChecked(QtCore.Qt.Checked == state)

    # after ok selected
    def menuCloseRuralUrban(self):
        self.keywords[self.col] = []
        for element in self.checkBoxs:
            if element.isChecked():
                self.keywords[self.col].append(element.text())
        print(self.keywords[self.col])
        self.filterdataRuralUrban()
        self.pushButton_29.setText('Selected')
        self.pushButton_29.setEnabled(False)
        self.pushButton_33.setText('चयनित')
        self.pushButton_33.setEnabled(False)
        self.menu.close()

    # Filter data columnwise
    def filterdataRuralUrban(self):
        global df
        #keywords = dict([(i, []) for i in range(self.filterall.columnCount())])
        columnsShow = dict([(i, True) for i in range(df['col_18'].shape[0])])
        print(columnsShow)

        j = 0
        for j in range(df['col_18'].shape[0]):
            item = df['col_18'].to_list()

            # if self.keywords[self.col]:
            if item[j] not in self.keywords[self.col]:
                columnsShow[j] = False

        # for key, value in columnsShow.items():
        final_lst = [i for i in columnsShow.values()]
        print(final_lst, 'this is final list of Select Month')

        # matching list of facility type with col of dataframe returned by onSelectDistrict fun
        df = df[final_lst]
        return df


    ################################################################################
    # Select Ownership

    # Select Ownership Filter
    def onSelectOwnership(self, index):
        self.keywords = dict([(i, []) for i in range(df.shape[0])])
        print(self.keywords)
        self.menu = QtWidgets.QMenu(TabWidget)
        self.menu.setStyleSheet('QMenu { menu-scrollable: true; width: 300 }')
        font = self.menu.font()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.menu.setFont(font)

        index = 19
        self.col = index

        data_unique = []

        self.checkBoxs = []

        # Selectall added into Dropdown
        checkBox = QtWidgets.QCheckBox("Select all/ Deselect all", self.menu)
        checkBox.setStyleSheet("color: red; spacing: 5px; font-size:15px;")
        # All the checkboxes are enabled to check
        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(checkBox)
        self.menu.addAction(checkableAction)
        checkBox.setChecked(True)
        checkBox.stateChanged.connect(self.slotSelectOwnership)

        # list storing Facility Name data
        df['col_19'].fillna('blank', inplace = True)
        list_set = df['col_19'].to_list()

        item = list_set
        item = sorted(list_set, key=str.upper)

        # looping to fill checkboxes, initially all checkboxes will be checked
        for i in range(len(item)):
            if item[i] not in data_unique:
                data_unique.append(item[i])
                checkBox = QtWidgets.QCheckBox(item[i], self.menu)
                checkBox.setChecked(True)
                checkableAction = QtWidgets.QWidgetAction(self.menu)
                checkableAction.setDefaultWidget(checkBox)
                self.menu.addAction(checkableAction)
                self.checkBoxs.append(checkBox)

        # Ok, cancel button
        btn = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel,
                                            QtCore.Qt.Vertical, self.menu)

        # ok selected
        btn.accepted.connect(self.menuCloseOwnership)
        # rejected , nothing selected
        btn.rejected.connect(self.menu.close)

        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(btn)
        self.menu.addAction(checkableAction)

        ############# Always set Pushbutton ####################
        self.pushButton_28.setMenu(self.menu)
        self.pushButton_32.setMenu(self.menu)

    # method to check -> uncheck and vice versa
    def slotSelectOwnership(self, state):
        for checkbox in self.checkBoxs:
            checkbox.setChecked(QtCore.Qt.Checked == state)

    # after ok selected
    def menuCloseOwnership(self):
        self.keywords[self.col] = []
        for element in self.checkBoxs:
            if element.isChecked():
                self.keywords[self.col].append(element.text())
        print(self.keywords[self.col])
        self.filterdataOwnership()
        self.pushButton_28.setText('Selected')
        self.pushButton_28.setEnabled(False)
        self.pushButton_32.setText('चयनित')
        self.pushButton_32.setEnabled(False)
        self.menu.close()

    # Filter data columnwise
    def filterdataOwnership(self):
        global df
        #keywords = dict([(i, []) for i in range(self.filterall.columnCount())])
        columnsShow = dict([(i, True) for i in range(df['col_19'].shape[0])])
        print(columnsShow)

        j = 0
        for j in range(df['col_19'].shape[0]):
            item = df['col_19'].to_list()

            # if self.keywords[self.col]:
            if item[j] not in self.keywords[self.col]:
                columnsShow[j] = False

        # for key, value in columnsShow.items():
        final_lst = [i for i in columnsShow.values()]

        # matching list of facility type with col of dataframe returned by onSelectDistrict fun
        df = df[final_lst]
        return df


    # To count summary of the Modified Checks
    # =======================================
    def summaryReport(self, df):
        global final_result_summ1, final_result_summ2, col_sum, dft_ARFacilityWise, dft_ARCheckWiseInc, FList1, dft_ARCheckWisePRE, FList2, dft_FacilityWisePRE, dataframeForSheet6, dataframeForSheet7, FList3, FList4, dataframeForSheet4, dataframeForSheet5
        FType = self.lineEdit_5.text()

        # For Health Sub Centre
        if FType == 'Health Sub Centre': 
            df_SummReport = df.iloc[:, 200:]     ## Taking columns after 200th
            val_Description = [
                                'Number of mothers provided full course of 180 IFA tablets after delivery <= Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) + Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+ Number of Institutional Deliveries conducted',
                                'Out of the total ANC registered, number registered within 1st trimester (within 12 weeks)<=Total number of pregnant women registered for ANC', 
                                'Out of the new cases of PW with hypertension detected, cases managed at institution <=New cases of PW with hypertension detected',
                                'Number of PW received 4 or more ANC check ups<=Total number of pregnant women registered for ANC',
                                'Number of PW tested using POC test for Syphilis<=Total number of pregnant women registered for ANC',
                                    'Out of the above, Number of PW found sero positive for Syphilis<=Number of PW tested using POC test for Syphilis',
                                    'Number of PW given Tablet Misoprostol during home delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)',
                                    'Number of newborns received 7 Home Based Newborn Care (HBNC) visits in case of Home delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)',
                                    'Number of newborns received 6 HBNC  visits after Institutional Delivery<=Number of Institutional Deliveries conducted',
                                        'Number of mothers provided 360 Calcium tablets after delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted',
                                        'Child immunisation - Vitamin K1 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                        'Child immunisation - OPV 0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                        'Child immunisation - Hepatitis-B0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                        'Out of total institutional deliveries number of women discharged within 48 hours of delivery<=Number of Institutional Deliveries conducted',
                                        'Number of Pre term newborns ( < 37 weeks of pregnancy)<=Live Birth - Male+Live Birth - Female',
                                        'Number of newborns weighed at birth<=Live Birth - Male+Live Birth - Female',
                                            'Number of newborns having weight less than 2.5 kg<=Number of newborns weighed at birth',
                                            'Number of Newborns breast fed within 1 hour of birth<=Live Birth - Male+Live Birth - Female',
                                            'Women receiving 1st post partum checkup within 48 hours of home delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)',
                                            'Number of Post Partum (within 48 hours of delivery) IUCD insertions<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Children aged between 9 and 11 months fully immunized- Male+Children aged between 9 and 11 months fully immunized - Female<=Child immunisation (9-11months) - Measles & Rubella (MR) 1st dose  & Child immunisation (9-11months) - Measles 1st dose',
                                            'Number of cases of AEFI - Abscess<=Number of children immunized (6.1.1+6.1.2+6.1.3+6.1.4+6.1.5+6.1.6+6.1.7+6.1.8+6.1.13+6.1.14+6.1.15+6.1.16+6.1.17+6.1.18+6.1.19+6.1.20+6.1.21+6.2.1+6.2.2+6.2.3+6.3.1+6.3.2+6.3.3+6.4.1+6.4.2+6.4.3+6.4.5+6.4.6+6.5.1+6.5.2+6.5.3+6.5.4)',
                                                'Number of cases of AEFI - Death<=Number of children immunized (6.1.1+6.1.2+6.1.3+6.1.4+6.1.5+6.1.6+6.1.7+6.1.8+6.1.13+6.1.14+6.1.15+6.1.16+6.1.17+6.1.18+6.1.19+6.1.20+6.1.21+6.2.1+6.2.2+6.2.3+6.3.1+6.3.2+6.3.3+6.4.1+6.4.2+6.4.3+6.4.5+6.4.6+6.5.1+6.5.2+6.5.3+6.5.4)',
                                                'Number of cases of AEFI - Others<=Number of children immunized (6.1.1+6.1.2+6.1.3+6.1.4+6.1.5+6.1.6+6.1.7+6.1.8+6.1.13+6.1.14+6.1.15+6.1.16+6.1.17+6.1.18+6.1.19+6.1.20+6.1.21+6.2.1+6.2.2+6.2.3+6.3.1+6.3.2+6.3.3+6.4.1+6.4.2+6.4.3+6.4.5+6.4.6+6.5.1+6.5.2+6.5.3+6.5.4)',
                                                'Number of Immunisation sessions where ASHAs were present<=Immunisation sessions held',
                                                'Out of the total number of Hb tests done, Number having Hb < 7 mg <=Number of Hb tests conducted',
                                                'out of the above, Number screened positive<=Number of Pregnant Women screened for HIV',
                                                    'Live Birth - Male+Live Birth - Female+Still Birth>=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted',
                                                    'Malaria (RDT) - Plamodium Falciparum test positive<=RDT conducted for Malaria',
                                                    'Allopathic- Outpatient attendance+Ayush - Outpatient attendance >= 9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8',
                                                    'Malaria (RDT) - Plasmodium Vivax test positive<=RDT conducted for Malaria' ]

        # For Primary Health Centre
        elif FType == 'Primary Health Centre':
            df_SummReport = df.iloc[:, 305:]     ## Taking columns after 305th
            val_Description = [
                                'Child immunisation - Vitamin K (Birth Dose) <= Live Birth - Male+Live Birth - Female',
                                'Out of the total ANC registered, number registered within 1st trimester (within 12 weeks)<=Total number of pregnant women registered for ANC',
                                'Number of PW given 180 Iron Folic Acid (IFA) tablets <=Total number of pregnant women registered for ANC', 
                                'Number of PW given 360 Calcium tablets <=Total number of pregnant women registered for ANC', 
                                'Number of PW received 4 or more ANC check ups<=Total number of pregnant women registered for ANC', 
                                'Number of PW given Tablet Misoprostol during home delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)',
                                'Number of newborns received 7 Home Based Newborn Care (HBNC) visits in case of Home delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)',
                                'Out of total institutional deliveries number of women discharged within 48 hours of delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                'Number of Eclampsia cases managed during delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'No. of PW having severe anaemia (Hb<7) treated could be greater than No. of PW having severe anaemia (Hb<7)  tested cases',
                                    'Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)<=Total number of pregnant women registered for ANC',
                                    'Number of PW tested positive for GDM<=Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)',
                                    'Number of PW given insulin out of total tested positive for GDM<=Number of PW tested positive for GDM',
                                    'Number of Pregnant women tested for Syphilis<=Total number of pregnant women registered for ANC', 
                                    'Number of Pregnant women tested found sero positive for Syphilis<=Number of Pregnant women tested for Syphilis',
                                    'Number of Syphilis positive pregnant women treated for Syphilis<=Number of Pregnant women tested found sero positive for Syphilis',
                                    'Number of babies treated for congenital Syphilis<=Number of babies diagnosed with congenital Syphilis',
                                    'C-sections, performed at night (8 PM- 8 AM)<=Total C -Section deliveries performed',
                                    'Total C -Section deliveries performed<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Number of Pre term newborns ( < 37 weeks of pregnancy)<=Live Birth - Male+Live Birth - Female',
                                        'Live Birth - Male+Live Birth - Female+Still Birth>=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Post Abortion/ MTP Complications Identified<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                        'Post Abortion/ MTP Complications Treated<=Post Abortion/ MTP Complications Identified',
                                        'Number of women provided with post abortion/ MTP contraception <=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                        'Number of newborns weighed at birth<=Live Birth - Male+Live Birth - Female',
                                        'Number of newborns having weight less than 2.5 kg<=Number of newborns weighed at birth',
                                        'Number of Newborns breast fed within 1 hour of birth <=Live Birth - Male+Live Birth - Female',
                                        'Women receiving 1st post partum checkup within 48 hours of home delivery<=Number of Home Deliveries attended by Skill Birth Attendant (SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)',
                                        'Number of mothers provided full course of 180 IFA tablets after delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of mothers provided 360 Calcium tablets after delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'RTI/STI for which treatment initiated - Male<=New RTI/STI cases identified - Male',
                                            'RTI/STI for which treatment initiated -Female<=New RTI/STI cases identified - Female',
                                            'Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of Post Partum (within 48 hours of delivery) IUCD insertions<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted (Including C-Sections)',
                                                'Number of complications following IUCD Insertion<=Number of Interval IUCD Insertions (excluding PPIUCD and PAIUCD)+ Number of post partum (with in 48 hours of delivery) IUCD insertion +Number of post abortion (with 12 days of spontaneous or surgical abortions)  IUCD incertion',
                                                'Complications following male sterilization<=Number of Non Scalpel Vasectomy (NSV) / Conventional Vasectomy conducted',
                                                'Complications following female sterilization<=Number of Laparoscopic sterilizations (excluding post abortion) conducted + Number of Interval Mini-lap (other than post-partum and post abortion) sterilizations conducted + Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted + Number of Post Abortion sterilizations (within 7 days of spontaneous or surgical abortion) conducted',
                                                'Child immunisation - OPV 0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                                'Child immunisation - Hepatitis-B0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                                'Children aged between 9 and 11 months fully immunized- Male+Children aged between 9 and 11 months fully immunized - Female<=Child immunisation (9-11months) - Measles & Rubella (MR) 1st dose  & Child immunisation (9-11months) - Measles 1st dose',
                                                'Kala Azar Positive Cases<=Kala Azar (RDT) - Tests Conducted',
                                                    'Out of registered, Girls received clinical services<=Girls registered in AFHC',
                                                    'Out of registered, Boys received clinical services<=Boys registered in AFHC',
                                                    'Out of registered, Girls received counselling<=Girls registered in AFHC',
                                                    'Out of registered, Boys received counselling<=Boys registered in AFHC',
                                                        'Allopathic- Outpatient attendance+Ayush - Outpatient attendance >= 14.1.1+14.1.2+14.1.3+14.1.4+14.1.5+14.1.6+14.1.7+14.1.8',
                                                        'Number of Left Against Medical Advice (LAMA) cases<=Male Admissions +Female Admissions',
                                                        'Out of Operation major, Gynecology- Hysterectomy surgeries<=Operation major (General and spinal anaesthesia)',
                                                        'out of the above, Number screened positive<=Number of Pregnant Women screened for HIV',
                                                        'number positive for HIV (Number confirmed positive at ICTCs)<=out of the above, Number screened positive',
                                                        'Widal tests - Number Positive<=Widal tests - Number Tested',
                                                        'Number of cases of AEFI - Abscess<=Number of Children Immunized (9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4)',
                                                            'Number of cases of AEFI - Death<=Number of Children Immunized (9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4)',
                                                            'Number of cases of AEFI - Others<=Number of Children Immunized (9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4)',
                                                            'Out of the new cases of PW with hypertension detected, cases managed at institution <=New cases of PW with hypertension detected',
                                                            'Immunisation sessions held <=Immunisation sessions planned',
                                                            'Number of Immunisation sessions where ASHAs were present<=Immunisation sessions held', 
                                                                'Malaria (Microscopy Tests ) - Plasmodium Falciparum test positive<=Total Blood Smears Examined for Malaria', 
                                                                'Malaria (Microscopy Tests ) - Plasmodium Falciparum test positive<=Total Blood Smears Examined for Malaria',
                                                                'Malaria (RDT) - Plasmodium Vivax test positive<=RDT conducted for Malaria',
                                                                'Malaria (RDT) - Plamodium Falciparum test positive<=RDT conducted for Malaria',
                                                                'Inpatient - Malaria <=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                'Inpatient - Dengue<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                    'Inpatient - Typhoid<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                    'Inpatient - Asthma, Chronic Obstructive Pulmonary Disease (COPD), Respiratory infections<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                    'Inpatient - Tuberculosis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                    'Inpatient - Pyrexia of unknown origin (PUO)<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                    'Inpatient - Diarrhea with dehydration<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                        'Inpatient - Hepatitis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                        'Inpatient Deaths - Male<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                        'Inpatient Deaths - Female<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                        'Number of children discharged with target weight gain from the NRCs<=Number of children admitted in NRC',
                                                                        'Out of the total number of Hb tests done, Number having Hb < 7 mg<=Number of Hb tests conducted',
                                                                            'Male HIV - Number Positive<=Male HIV - Number Tested',
                                                                            'Female Non ANC HIV - Number Positive<=Female Non ANC HIV - Number Tested',
                                                                            'Number of Male STI/RTI attendees found sero Positive for syphilis<=Number of Male STI/RTI attendees tested for syphilis',
                                                                            'Number of Female (Non ANC) STI/RTI attendees found sero Positive for syphilis<=Number of Female (Non ANC)STI/RTI attendees tested for syphilis',
                                                                            'Child immunisation - BCG<=Live Birth - Male+Live Birth - Female']


        # For Sub District Hospital
        elif FType == 'Community Health Centre':
            df_SummReport = df.iloc[:, 318:]     ## Taking columns after 305th
            val_Description = [
                                'Out of the ANC registered, number registered with in 1st trimester(Within 12 weeks)<=Total number of pregnant women registered for ANC',
                                'Male HIV Number Positive <= Male HIV - Number Tested',
                                'Number of PW given 180 Iron Folic Acid (IFA) tablets <=Total number of pregnant women registered for ANC ',
                                'Number of PW given 360 Calcium tablets <=Total number of pregnant women registered for ANC ',
                                'Number of PW received 4 or more ANC check ups<=Total number of pregnant women registered for ANC ',
                                'Out of the new cases of PW with hypertension detected, cases managed at institution<=New cases of PW with hypertension detected ',
                                'Number of Eclampsia cases managed during delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                'No. of PW having severe anaemia (Hb<7) treated could be greater than No. of PW having severe anaemia (Hb<7)  tested cases',
                                'Number of PW tested positive for GDM<=Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)',
                                'Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)<=Total number of pregnant women registered for ANC ',
                                    'Number of PW given insulin out of total tested positive for GDM<=Number of PW tested positive for GDM',
                                    'Number of Pregnant women tested found sero positive for Syphilis<=Number of Pregnant women tested for Syphilis',
                                    'Number of Pregnant women tested for Syphilis<=Total number of pregnant women registered for ANC ',
                                    'Number of Syphilis positive pregnant women treated for Syphilis<=Number of Pregnant women tested found sero positive for Syphilis',
                                    'Number of babies treated for congenital Syphilis<=Number of babies diagnosed with congenital Syphilis',
                                    'Out of total institutional deliveries number of women discharged within 48 hours of delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'Total C -Section deliveries performed<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'C-sections, performed at night (8 PM- 8 AM)<=Total C -Section deliveries performed',
                                    'Live Birth - Male + Live Birth - Female + Still Birth>=Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'Number of Pre term newborns ( < 37 weeks of pregnancy)<=Live Birth - Male+Live Birth - Female',
                                    'Post Abortion/ MTP Complications Identified<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                    'Post Abortion/ MTP Complications Treated<=Post Abortion/ MTP Complications Identified',
                                        'Number of women provided with post abortion/ MTP contraception<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                        'Number of newborns weighed at birth<=Live Birth - Male+Live Birth - Female',
                                        'Number of newborns having weight less than 2.5 kg<=Number of newborns weighed at birth',
                                        'Number of Newborns breast fed within 1 hour of birth<=Live Birth - Male+Live Birth - Female',
                                        'Number of Complicated pregnancies treated with Blood Transfusion<=Number of cases of pregnant women with Obstetric Complications attended (Antepartum haemorrhage (APH), Post-Partum Hemorrhage (PPH), Sepsis, Eclampsia and others) ',
                                        'Number of mothers provided full course of 180 IFA tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Number of mothers provided 360 Calcium tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'RTI/STI for which treatment initiated - Male<=New RTI/STI cases identified - Male',
                                        'RTI/STI for which treatment initiated -Female<=New RTI/STI cases identified - Female',
                                        'Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Number of Post Partum (within 48 hours of delivery) IUCD insertions<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of complications following IUCD Insertion<=Number of Interval IUCD Insertions (excluding PPIUCD and PAIUCD)+ Number of post partum (with in 48 hours of delivery) IUCD insertion +Number of post abortion (with 12 days of spontaneous or surgical abortions)  IUCD incertion',
                                            'Complications following male sterilization<=Number of Non Scalpel Vasectomy (NSV) / Conventional Vasectomy conducted',
                                            'Complications following female sterilization<=Number of Laparoscopic sterilizations (excluding post abortion) conducted + Number of Interval Mini-lap (other than post-partum and post abortion) sterilizations conducted + Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted + Number of Post Abortion sterilizations (within 7 days of spontaneous or surgical abortion) conducted',
                                            'Child immunisation - Vitamin K1(Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                            'Child immunisation - BCG<=Live Birth - Male+Live Birth - Female',
                                            'Child immunisation - OPV-0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                            'Child immunisation - Hepatitis-B0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                            'Children aged between 9 and 11 months fully immunized- Male+Children aged between 9 and 11 months fully immunize<=Child immunisation (9-11months) - Measles & Rubella (MR) 1st dose  & Child immunisation (9-11months) - Measles 1st dose',
                                            'Kala Azar Positive Cases<=Kala Azar (RDT) - Tests Conducted',
                                            'Out of registered, Girls received clinical services<=Girls registered in AFHC',
                                            'Out of registered, Boys received clinical services<=Boys registered in AFHC',
                                            'Out of registered, Girls received counselling<=Girls registered in AFHC',
                                            'Out of registered, Boys received counselling<=Boys registered in AFHC',
                                                'Allopathic- Outpatient attendance+Ayush - Outpatient attendance >= Number of outpatients (Diabetes + Hypertension +  Stroke (Paralysis) + Acute Heart Diseases + Mental illness + Epilepsy + Ophthalmic Related + Dental + Oncology',
                                                'Number of Left Against Medical Advice (LAMA) cases<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Malaria<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Dengue<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Typhoid<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Asthma, Chronic Obstructive Pulmonary Disease (COPD), Respiratory infections<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Tuberculosis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Pyrexia of unknown origin (PUO)<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Diarrhea with dehydration<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Hepatitis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Emergency - Trauma ( accident, injury, poisoning etc)<= Patients registered at Emergency Department',
                                                'Emergency - Burn<= Patients registered at Emergency Department',
                                                'Emergency - Obstetrics complications<= Patients registered at Emergency Department',
                                                    'Emergency - Snake Bite<=Patients registered at Emergency Department',
                                                    'Emergency - Acute Caridiac Emergencies<= Patients registered at Emergency Department',
                                                    'Number of deaths occurring at Emergency Department<= Patients registered at Emergency Department',
                                                    'Number of children discharged with target weight gain from the NRCs<=Number of children admitted in NRC',
                                                    'Out of the total number of Hb tests done, Number having Hb < 7 mg<=Number of Hb tests conducted',
                                                    'Female Non ANC HIV - Number Positive<=Female Non ANC HIV - Number Tested',
                                                    'out of the above, Number screened positive<=Number of Pregnant Women screened for HIV',
                                                    'out of the above, Number screened positive, number confirmed with HIV infection at Integrated Counselling and Testing Centre (ICTC) <=out of the above, Number screened positive',
                                                    'Widal tests - Number Positive<=Widal tests - Number Tested',
                                                    'Number of cases of AEFI - Abscess<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                    'Number of cases of AEFI - Death<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                    'Number of cases of AEFI - Others<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                    'Immunisation sessions held <=Immunisation sessions planned ',
                                                        'Number of Immunisation sessions where ASHAs were present<=Immunisation sessions held ',
                                                        'Malaria (Microscopy Tests ) - Plasmodium Vivax test positive<=Total Blood Smears Examined for Malaria ',
                                                        'Malaria (Microscopy Tests ) - Plasmodium Falciparum test positive<=Total Blood Smears Examined for Malaria ',
                                                        'Malaria (RDT) - Plasmodium Vivax test positive<=RDT conducted for Malaria',
                                                        'Malaria (RDT) - Plamodium Falciparum test positive<=RDT conducted for Malaria',
                                                        'Inpatient Deaths - Male <=Inpatient (Male)- Children<18yrs+Inpatient (Male)',
                                                        'Inpatient Deaths - Female<=Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                        'Number of deaths occurring at SNCU<=Special Newborn Care Unit (SNCU Admissions) - Inborn Male + Special Newborn Care Unit (SNCU Admissions) - Inborn Female + Outborn – Male + Outborn - Female + Number of newborns admitted in SNCU - referred by ASHA',
                                                        'Out of Operation major, Gynecology- Hysterectomy surgeries<=Operation major (General and spinal anaesthesia)',
                                                        'Number of Male STI/RTI attendees found sero Positive for syphilis<=Number of Male STI/RTI attendees tested for syphilis',
                                                        'Number of Female (Non ANC) STI/RTI attendees found sero Positive for syphilis<=Number of Female (Non ANC)STI/RTI attendees tested for syphilis'
                                                        ]

        # For Primary Health Centre
        elif FType == 'Sub District Hospital':
            df_SummReport = df.iloc[:, 321:]     ## Taking columns after 321th
            val_Description = [
                                'Out of the ANC registered, number registered with in 1st trimester(Within 12 weeks)<=Total number of pregnant women registered for ANC',
                                'Male HIV Number Positive <= Male HIV - Number Tested',
                                'Number of PW given 180 Iron Folic Acid (IFA) tablets <=Total number of pregnant women registered for ANC ',
                                'Number of PW given 360 Calcium tablets <=Total number of pregnant women registered for ANC ',
                                'Number of PW received 4 or more ANC check ups<=Total number of pregnant women registered for ANC ',
                                'Out of the new cases of PW with hypertension detected, cases managed at institution<=New cases of PW with hypertension detected ',
                                'Number of Eclampsia cases managed during delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'No. of PW having severe anaemia (Hb<7) treated could be greater than No. of PW having severe anaemia (Hb<7)  tested cases',
                                    'Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)<=Total number of pregnant women registered for ANC ',
                                    'Number of PW tested positive for GDM<=Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)',
                                    'Number of PW given insulin out of total tested positive for GDM<=Number of PW tested positive for GDM',
                                    'Number of Pregnant women tested for Syphilis<=Total number of pregnant women registered for ANC ',
                                    'Number of Pregnant women tested found sero positive for Syphilis<=Number of Pregnant women tested for Syphilis',
                                    'Number of Syphilis positive pregnant women treated for Syphilis<=Number of Pregnant women tested found sero positive for Syphilis',
                                    'Number of babies treated for congenital Syphilis<=Number of babies diagnosed with congenital Syphilis',
                                        'Out of total institutional deliveries number of women discharged within 48 hours of delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Total C -Section deliveries performed<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'C-sections, performed at night (8 PM- 8 AM)<=Total C -Section deliveries performed',
                                        'Live Birth - Male + Live Birth - Female + Still Birth>=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Number of Pre term newborns ( < 37 weeks of pregnancy)<=Live Birth - Male+Live Birth - Female',
                                        'Post Abortion/ MTP Complications Identified<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                        'Post Abortion/ MTP Complications Treated<=Post Abortion/ MTP Complications Identified',
                                        'Number of women provided with post abortion/ MTP contraception<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                        'Number of newborns weighed at birth<=Live Birth - Male+Live Birth - Female',
                                        'Number of newborns having weight less than 2.5 kg<=Number of newborns weighed at birth',
                                        'Number of Newborns breast fed within 1 hour of birth<=Live Birth - Male+Live Birth - Female',
                                            'Number of Complicated pregnancies treated with Blood Transfusion<=Number of cases of pregnant women with Obstetric Complications attended (Antepartum haemorrhage (APH), Post-Partum Hemorrhage (PPH), Sepsis, Eclampsia and others) ',
                                            'Number of mothers provided full course of 180 IFA tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of mothers provided 360 Calcium tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'RTI/STI for which treatment initiated - Male<=New RTI/STI cases identified - Male',
                                            'RTI/STI for which treatment initiated -Female<=New RTI/STI cases identified - Female',
                                            'Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of Post Partum (within 48 hours of delivery) IUCD insertions<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of complications following IUCD Insertion<=Number of Interval IUCD Insertions (excluding PPIUCD and PAIUCD)+ Number of post partum (with in 48 hours of delivery) IUCD insertion +Number of post abortion (with 12 days of spontaneous or surgical abortions)  IUCD incertion',
                                            'Complications following male sterilization<=Number of Non Scalpel Vasectomy (NSV) / Conventional Vasectomy conducted',
                                            'Complications following female sterilization<=Number of Laparoscopic sterilizations (excluding post abortion) conducted + Number of Interval Mini-lap (other than post-partum and post abortion) sterilizations conducted + Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted + Number of Post Abortion sterilizations (within 7 days of spontaneous or surgical abortion) conducted',
                                            'Child immunisation - Vitamin K1(Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                                'Child immunisation - BCG<=Live Birth - Male+Live Birth - Female',
                                                'Child immunisation - OPV-0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                                'Child immunisation - Hepatitis-B0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                                'Children aged between 9 and 11 months fully immunized- Male+Children aged between 9 and 11 months fully immunize<=Child immunisation (9-11months) - Measles & Rubella (MR) 1st dose  & Child immunisation (9-11months) - Measles 1st dose',
                                                'Kala Azar Positive Cases<=Kala Azar (RDT) - Tests Conducted',
                                                'Tests Positive for JE<=Tests Conducted for JE',
                                                'Out of registered, Girls received clinical services<=Girls registered in AFHC',
                                                'Out of registered, Boys received clinical services<=Boys registered in AFHC',
                                                'Out of registered, Girls received counselling<=Girls registered in AFHC',
                                                'Out of registered, Boys received counselling<=Boys registered in AFHC',
                                                'Allopathic- Outpatient attendance+Ayush - Outpatient attendance >= Number of outpatients (Diabetes + Hypertension +  Stroke (Paralysis) + Acute Heart Diseases + Mental illness + Epilepsy + Ophthalmic Related + Dental + Oncology)',
                                                    'Number of Left Against Medical Advice (LAMA) cases<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Malaria<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Dengue<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Typhoid<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Asthma, Chronic Obstructive Pulmonary Disease (COPD), Respiratory infections<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Tuberculosis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Pyrexia of unknown origin (PUO)<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Diarrhea with dehydration<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Hepatitis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Emergency - Trauma ( accident, injury, poisoning etc)<= Patients registered at Emergency Department',
                                                    'Emergency - Burn<= Patients registered at Emergency Department',
                                                    'Emergency - Obstetrics complications<= Patients registered at Emergency Department',
                                                        'Emergency - Snake Bite<=Patients registered at Emergency Department',
                                                        'Emergency - Acute Caridiac Emergencies<= Patients registered at Emergency Department',
                                                        'Emergency - CVA ( Cerebovascular Disease)<= Patients registered at Emergency Department',
                                                        'Number of deaths occurring at Emergency Department<= Patients registered at Emergency Department',
                                                        'Number of children discharged with target weight gain from the NRCs<=Number of children admitted in NRC',
                                                        'Out of the total number of Hb tests done, Number having Hb < 7 mg<=Number of Hb tests conducted',
                                                        'Female Non ANC HIV - Number Positive<=Female Non ANC HIV - Number Tested',
                                                        'out of the above, Number screened positive<=Number of Pregnant Women screened for HIV',
                                                        'out of the above, Number screened positive, number confirmed with HIV infection at Integrated Counselling and Testing Centre (ICTC) <=out of the above, Number screened positive',
                                                        'Widal tests - Number Positive<=Widal tests - Number Tested',
                                                        'Number of cases of AEFI - Abscess<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                        'Number of cases of AEFI - Death<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                        'Number of cases of AEFI - Others<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                        'Immunisation sessions held <=Immunisation sessions planned ',
                                                            'Number of Immunisation sessions where ASHAs were present<=Immunisation sessions held ',
                                                            'Malaria (Microscopy Tests ) - Plasmodium Vivax test positive<=Total Blood Smears Examined for Malaria ',
                                                            'Malaria (Microscopy Tests ) - Plasmodium Falciparum test positive<=Total Blood Smears Examined for Malaria ',
                                                            'Malaria (RDT) - Plasmodium Vivax test positive<=RDT conducted for Malaria',
                                                            'Malaria (RDT) - Plamodium Falciparum test positive<=RDT conducted for Malaria',
                                                            'Inpatient Deaths - Male <=Inpatient (Male)- Children<18yrs+Inpatient (Male)',
                                                            'Inpatient Deaths - Female<=Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                            'Number of deaths occurring at SNCU<=Special Newborn Care Unit (SNCU Admissions) - Inborn Male + Special Newborn Care Unit (SNCU Admissions) - Inborn Female + Outborn – Male + Outborn - Female + Number of newborns admitted in SNCU - referred by ASHA',
                                                            'Out of Operation major, Gynecology- Hysterectomy surgeries<=Operation major (General and spinal anaesthesia)',
                                                            'Number of Male STI/RTI attendees found sero Positive for syphilis<=Number of Male STI/RTI attendees tested for syphilis',
                                                            'Number of Female (Non ANC) STI/RTI attendees found sero Positive for syphilis<=Number of Female (Non ANC)STI/RTI attendees tested for syphilis']
            
        # For District Hospital
        elif FType == 'District Hospital':
            df_SummReport = df.iloc[:, 326:]     ## Taking columns after 326th
            val_Description = [
                                'Out of the ANC registered, number registered with in 1st trimester(Within 12 weeks)<=Total number of pregnant women registered for ANC',
                                'Male HIV Number Positive <= Male HIV - Number Tested',
                                'Number of PW given 180 Iron Folic Acid (IFA) tablets <=Total number of pregnant women registered for ANC ',
                                'Number of PW given 360 Calcium tablets <=Total number of pregnant women registered for ANC ',
                                'Number of PW received 4 or more ANC check ups<=Total number of pregnant women registered for ANC ',
                                'Out of the new cases of PW with hypertension detected, cases managed at institution<=New cases of PW with hypertension detected ',
                                'Number of Eclampsia cases managed during delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                'No. of PW having severe anaemia (Hb<7) treated could be greater than No. of PW having severe anaemia (Hb<7)  tested cases',
                                'Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)<=Total number of pregnant women registered for ANC ',
                                    'Number of PW tested positive for GDM<=Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)',
                                    'Number of PW given insulin out of total tested positive for GDM<=Number of PW tested positive for GDM',
                                    'Number of Pregnant women tested for Syphilis<=Total number of pregnant women registered for ANC ',
                                    'Number of Pregnant women tested found sero positive for Syphilis<=Number of Pregnant women tested for Syphilis',
                                    'Number of Syphilis positive pregnant women treated for Syphilis<=Number of Pregnant women tested found sero positive for Syphilis',
                                    'Number of babies treated for congenital Syphilis<=Number of babies diagnosed with congenital Syphilis',
                                    'Out of total institutional deliveries number of women discharged within 48 hours of delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'Total C -Section deliveries performed<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'C-sections, performed at night (8 PM- 8 AM)<=Total C -Section deliveries performed',
                                    'Live Birth - Male + Live Birth - Female + Still Birth>=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Number of Pre term newborns ( < 37 weeks of pregnancy)<=Live Birth - Male+Live Birth - Female',
                                        'Post Abortion/ MTP Complications Identified<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                        'Post Abortion/ MTP Complications Treated<=Post Abortion/ MTP Complications Identified',
                                        'Number of women provided with post abortion/ MTP contraception<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                        'Number of newborns weighed at birth<=Live Birth - Male+Live Birth - Female',
                                        'Number of newborns having weight less than 2.5 kg<=Number of newborns weighed at birth',
                                        'Number of Newborns breast fed within 1 hour of birth<=Live Birth - Male+Live Birth - Female',
                                        'Number of Complicated pregnancies treated with Blood Transfusion<=Number of cases of pregnant women with Obstetric Complications attended (Antepartum haemorrhage (APH), Post-Partum Hemorrhage (PPH), Sepsis, Eclampsia and others) ',
                                        'Number of mothers provided full course of 180 IFA tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Number of mothers provided 360 Calcium tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'RTI/STI for which treatment initiated - Male<=New RTI/STI cases identified - Male',
                                            'RTI/STI for which treatment initiated -Female<=New RTI/STI cases identified - Female',
                                            'Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of Post Partum (within 48 hours of delivery) IUCD insertions<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of complications following IUCD Insertion<=Number of Interval IUCD Insertions (excluding PPIUCD and PAIUCD)+ Number of post partum (with in 48 hours of delivery) IUCD insertion +Number of post abortion (with 12 days of spontaneous or surgical abortions)  IUCD incertion',
                                            'Complications following male sterilization<=Number of Non Scalpel Vasectomy (NSV) / Conventional Vasectomy conducted',
                                            'Complications following female sterilization<=Number of Laparoscopic sterilizations (excluding post abortion) conducted + Number of Interval Mini-lap (other than post-partum and post abortion) sterilizations conducted + Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted + Number of Post Abortion sterilizations (within 7 days of spontaneous or surgical abortion) conducted',
                                            'Child immunisation - Vitamin K1(Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                            'Child immunisation - BCG<=Live Birth - Male+Live Birth - Female',
                                            'Child immunisation - OPV-0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                            'Child immunisation - Hepatitis-B0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                            'Children aged between 9 and 11 months fully immunized- Male+Children aged between 9 and 11 months fully immunize<=Child immunisation (9-11months) - Measles & Rubella (MR) 1st dose  & Child immunisation (9-11months) - Measles 1st dose',
                                                'Kala Azar Positive Cases<=Kala Azar (RDT) - Tests Conducted',
                                                'Tests Positive for JE<=Tests Conducted for JE',
                                                'Out of registered, Girls received clinical services<=Girls registered in AFHC',
                                                'Out of registered, Boys received clinical services<=Boys registered in AFHC',
                                                'Out of registered, Girls received counselling<=Girls registered in AFHC',
                                                'Out of registered, Boys received counselling<=Boys registered in AFHC',
                                                'Allopathic- Outpatient attendance+Ayush - Outpatient attendance >= Number of outpatients (Diabetes + Hypertension +  Stroke (Paralysis) + Acute Heart Diseases + Mental illness + Epilepsy + Ophthalmic Related + Dental + Oncology',
                                                'Number of Left Against Medical Advice (LAMA) cases<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Malaria<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Dengue<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Typhoid<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Asthma, Chronic Obstructive Pulmonary Disease (COPD), Respiratory infections<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Tuberculosis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Pyrexia of unknown origin (PUO)<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Diarrhea with dehydration<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Hepatitis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Emergency - Trauma ( accident, injury, poisoning etc)<= Patients registered at Emergency Department',
                                                    'Emergency - Burn<= Patients registered at Emergency Department',
                                                    'Emergency - Obstetrics complications<= Patients registered at Emergency Department',
                                                    'Emergency - Snake Bite<=Patients registered at Emergency Department',
                                                    'Emergency - Acute Caridiac Emergencies<= Patients registered at Emergency Department',
                                                    'Emergency - CVA ( Cerebovascular Disease)<= Patients registered at Emergency Department',
                                                        'Number of deaths occurring at Emergency Department<= Patients registered at Emergency Department',
                                                        'Number of children discharged with target weight gain from the NRCs<=Number of children admitted in NRC',
                                                        'Out of the total number of Hb tests done, Number having Hb < 7 mg<=Number of Hb tests conducted',
                                                        'Female Non ANC HIV - Number Positive<=Female Non ANC HIV - Number Tested',
                                                        'out of the above, Number screened positive<=Number of Pregnant Women screened for HIV',
                                                        'out of the above, Number screened positive, number confirmed with HIV infection at Integrated Counselling and Testing Centre (ICTC) <=out of the above, Number screened positive',
                                                        'Widal tests - Number Positive<=Widal tests - Number Tested',
                                                        'Number of cases of AEFI - Abscess<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                        'Number of cases of AEFI - Death<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                        'Number of cases of AEFI - Others<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                        'Immunisation sessions held <=Immunisation sessions planned ',
                                                            'Number of Immunisation sessions where ASHAs were present<=Immunisation sessions held ',
                                                            'Malaria (Microscopy Tests ) - Plasmodium Vivax test positive<=Total Blood Smears Examined for Malaria ',
                                                            'Malaria (Microscopy Tests ) - Plasmodium Falciparum test positive<=Total Blood Smears Examined for Malaria ',
                                                            'Malaria (RDT) - Plasmodium Vivax test positive<=RDT conducted for Malaria',
                                                            'Malaria (RDT) - Plamodium Falciparum test positive<=RDT conducted for Malaria',
                                                            'Total number of blood units issued during the month>=Number of blood units issued (Excluding C-Section)',
                                                            'Inpatient Deaths - Male <=Inpatient (Male)- Children<18yrs+Inpatient (Male)',
                                                            'Inpatient Deaths - Female<=Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                            'Number of deaths occurring at SNCU<=Special Newborn Care Unit (SNCU Admissions) - Inborn Male + Special Newborn Care Unit (SNCU Admissions) - Inborn Female + Outborn – Male + Outborn - Female + Number of newborns admitted in SNCU - referred by ASHA',
                                                            'Out of Operation major, Gynecology- Hysterectomy surgeries<=Operation major (General and spinal anaesthesia)',
                                                            'Major Surgeries excluding Obstetrics, Gynaecology and Opthalmology etc.<=Operation major (General and spinal anaesthesia)',
                                                            'Number of Male STI/RTI attendees found sero Positive for syphilis<=Number of Male STI/RTI attendees tested for syphilis',
                                                            'Number of Female (Non ANC) STI/RTI attendees found sero Positive for syphilis<=Number of Female (Non ANC)STI/RTI attendees tested for syphilis']

        

        '''
        ## First Summary Report
        ## ---------------------
        '''

        # count_Consistent = []
        count_Inconsistent = []
        # count_Blank = []
        count_ProbableRErr = []

        Columns = list(df_SummReport.columns.values.tolist())

        for col_name in Columns:

            c2 = df_SummReport[col_name].str.count("Inconsistent").sum()
            count_Inconsistent.append(c2)

            c4 = df_SummReport[col_name].str.count("Probable Reporting Error").sum()
            count_ProbableRErr.append(c4)

        print(len(val_Description), len(count_Inconsistent), len(count_ProbableRErr))   

        # To show facilities in a column
        colInterest = df['col_14']
        colInterest = colInterest.tolist()
        
        ''' For Inconsistent '''
        inconsistent_list = []

        lg = len(df_SummReport.columns)

        for i in range(0, lg):
            temp = []
            colComparison = df_SummReport.iloc[:,i]
            colComparison = colComparison.tolist()
            for j in range(0, len(colComparison)):
                primString = colComparison[j]
                
                if primString == 'Inconsistent':
                    temp.append(colInterest[j])
                else:
                    continue

            inconsistent_list.append(temp)           

        ''' For PRE '''
        PRE_list = []
        # TO show facility in the sheet5
        lg = len(df_SummReport.columns)

        for i in range(0, lg):
            temp = []
            colComparison = df_SummReport.iloc[:,i]
            colComparison = colComparison.tolist()
            for j in range(0, len(colComparison)):
                primString = colComparison[j]

                pattern = re.compile("^P")
                if pattern.match(str(primString)):
                    temp.append(colInterest[j])
                else:
                    continue

            PRE_list.append(temp)      


        final_result_summ1 = pd.DataFrame({"Conditions": df_SummReport.columns, 
                                            "Description": val_Description,
                                                "Facilities(Name) Showing Inconsistent": inconsistent_list,
                                                    "Inconsistent": count_Inconsistent,
                                                        "Facilities (Name) Showing Probable Reporting Error": PRE_list,
                                                            "Probable Reporting Error": count_ProbableRErr,
                                                            })
        final_result_summ1 = final_result_summ1.sort_values(by=['Inconsistent'], ascending=False)
        final_result_summ1 = final_result_summ1.reset_index(drop=True)
                                                 
        FList1 = final_result_summ1["Facilities(Name) Showing Inconsistent"].tolist()
        FList2 = final_result_summ1["Facilities (Name) Showing Probable Reporting Error"].tolist()

        dataframeForSheet4 = final_result_summ1[['Conditions', 'Description', 'Inconsistent', 'Facilities(Name) Showing Inconsistent']]
        dataframeForSheet5 = final_result_summ1[['Conditions', 'Description', 'Probable Reporting Error', 'Facilities (Name) Showing Probable Reporting Error']]

        # Counting total numnber of facility names and calculating percentage and 
        # Coloring percentage >=25% to RED
        count_rows = df.shape[0]
            
            # Percentage for Validation Summary Sheet to show color codes
        final_result_summ1['PerIncSheet1'] = final_result_summ1['Inconsistent']/count_rows*100
        final_result_summ1['Inconsistent'].value_counts()
        final_result_summ1['PerPRErrSheet1'] = final_result_summ1['Probable Reporting Error']/count_rows*100

        len0 = len(final_result_summ1['PerIncSheet1'])
        len1 = len(final_result_summ1['PerPRErrSheet1'])

        # Deleting unnecessary columns
        # del final_result_summ1['PerIncSheet1']
        # del final_result_summ1['PerPRErrSheet1']
            
        def select_col_SumSheet(X):
            # COLORS
            c = ['background-color:  #EF5350',                  #   >=25% RED
                        'background-color: #FFAF00',            #   10 - 25% LIGHTER RED
                            'background-color: #C0C000',        #   5 - 10% MORE LIGHTER RED
                                'background-color: #00AF5F',    #   < 5% LIGHTEST RED
                                    '']

            mask_30 = (X['Inconsistent'] == 0)
            cnt30 = mask_30.values.sum()
            mask_29 = (X['Probable Reporting Error'] == 0)
            cnt29 = mask_29.values.sum()

            #DataFrame with same index and columns names as original filled empty strings
            df1 =  pd.DataFrame(c[4], X.index, columns=X.columns)
            df1.loc[mask_29, 'Probable Reporting Error'] = c[3]
            df1.loc[mask_30, 'Inconsistent'] = c[3]

            return df1

        # Remoiving inconsistent facility names from validation summary sheet
        final_result_summ1.drop(['Facilities(Name) Showing Inconsistent'], axis = 1, inplace=True)
        # Remoiving PRE facility names from validation summary sheet
        final_result_summ1.drop(['Facilities (Name) Showing Probable Reporting Error'], axis = 1, inplace=True)

        final_result_summ1 = final_result_summ1.style.apply(select_col_SumSheet, axis=None)


        '''
        ## Second Summary Report
        ## --------------------
        '''
        summ2_countInconsistent = []
        summ2_countProbableRErr = []
        All_Blank = []

        # Iterating over indices of each row and calculating number of Blanks for each Facility Name 
        for index in range(len(df_SummReport)):
            '''   For no. of Inconsistent   '''
            inconsistent = df_SummReport.iloc[index, :].str.count("Inconsistent").sum()
            summ2_countInconsistent.append(inconsistent)
            
            '''   For no. of Probable Reporting Errors   '''
            probableRErr = df_SummReport.iloc[index, :].str.count('Probable Reporting Error').sum()
            summ2_countProbableRErr.append(probableRErr)

            blank = df_SummReport.iloc[index, :].str.count("Blank").sum()
            if blank == len(df_SummReport.columns):
                All_Blank.append('Yes')
            else:
                All_Blank.append('No')

        

        #########################################################   
        #  Facility Specific Inconsistent (Sheet 6)      
        colInterest = df_['col_14']
        
        ''' For Inconsistent '''
        inc_list = []

        lg = len(df_SummReport.columns)
        len_df = df_SummReport.shape[0]

        for i in range(0, len_df):
            temp = []

            colComparison = df_SummReport.iloc[i,:]
            for j in range(0, lg):
                primString = colComparison[j]
                
                if primString == 'Inconsistent':
                    temp.append(df_SummReport.columns[j])

            inc_list.append(temp)

        ''' For PRE '''
        pre_list = []

        lg = len(df_SummReport.columns)
        len_df = df_SummReport.shape[0]

        for i in range(0, len_df):
            temp = []

            colComparison = df_SummReport.iloc[i,:]
            for j in range(0, lg):
                primString = colComparison[j]

                pattern = re.compile("^P")
                if pattern.match(str(primString)):
                    temp.append(df_SummReport.columns[j])

            pre_list.append(temp)

        

        ###################################################
        
        final_result_summ2 = pd.DataFrame({ "State": df['col_3'].tolist(),
                                                "District": df['col_5'].tolist(),
                                                    "Rural/Urban": df['col_18'].tolist(),
                                                        "Ownership": df['col_19'].tolist(),
                                                            "Facility Name": df['col_14'].tolist(),
                                                                "Inconsistent": summ2_countInconsistent,
                                                                    "Probable Reporting Error": summ2_countProbableRErr,
                                                                        "All Blank": All_Blank,
                                                                            "Checks (Inconsistent)" : inc_list,
                                                                                "Checks (PRE)": pre_list
                                                                                    })

        # Sorting in alphabetical  order
        # final_result_summ2 = final_result_summ2.sort_values(by=['Facility Name'], ascending=True)
        final_result_summ2 = final_result_summ2.sort_values(by=['Inconsistent'], ascending=False)
        final_result_summ2 = final_result_summ2.reset_index(drop=True)

        FList3 = final_result_summ2["Checks (Inconsistent)"].tolist()
        FList4 = final_result_summ2["Checks (PRE)"].tolist()


        '''
        ## Third Summary Report (Top 10 Validation Checks not performing good)
        ## --------------------
        '''
        # dft_ARCheckWiseInc = dataframeForSheet4.sort_values(by=['Inconsistent'], ascending=False)
        # dft_ARCheckWiseInc = dft_ARCheckWiseInc[dft_ARCheckWiseInc['Inconsistent'] != 0]
        # dft_ARCheckWiseInc = dft_ARCheckWiseInc.reset_index(drop=True)

        '''
        ## Fourth Summary Report (Top 10 Validation Checks not performing good)
        ## --------------------
        '''
        # dft_ARCheckWisePRE = dataframeForSheet5.sort_values(by=['Probable Reporting Error'], ascending=False)
        # dft_ARCheckWisePRE = dft_ARCheckWisePRE[dft_ARCheckWisePRE['Probable Reporting Error'] != 0]
        # dft_ARCheckWisePRE = dft_ARCheckWisePRE.reset_index(drop=True)

        '''
        ## 5th and 6th Summary Report (Top 10 Validation Checks not performing good)
        ## --------------------
        '''
        dataframeForSheet6 = final_result_summ2[['Facility Name', 'Inconsistent', 'Checks (Inconsistent)']]
        dataframeForSheet7 = final_result_summ2[['Facility Name', 'Probable Reporting Error', 'Checks (PRE)']]



        '''  To find percentage Facility Type Wise   '''
                                                    
        # For Health Sub Centre
        if FType == 'Health Sub Centre':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/31 * 100
            final_result_summ2['PercentagePRErr'] = final_result_summ2['Probable Reporting Error']/ 31 * 100
        
        # For Primary Health Centre
        elif FType == 'Primary Health Centre':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/78 * 100
            final_result_summ2['PercentagePRErr'] = final_result_summ2['Probable Reporting Error']/ 78 * 100

        # For Community Health Centre
        elif FType == 'Community Health Centre':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/83 * 100
            final_result_summ2['PercentagePRErr'] = final_result_summ2['Probable Reporting Error']/ 83 * 100

        # For Sub District Hospital
        elif FType == 'Sub District Hospital':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/85 * 100
            final_result_summ2['PercentagePRErr'] = final_result_summ2['Probable Reporting Error']/ 85 * 100

        # For District Hospital
        elif FType == 'District Hospital':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/85 * 100
            final_result_summ2['PercentagePRErr'] = final_result_summ2['Probable Reporting Error']/ 85 * 100
        
        ## Deleting unnecessary columns
        # del final_result_summ2['PercentageInc']
        # del final_result_summ2['PercentagePRErr']
        del final_result_summ2['Checks (Inconsistent)']
        del final_result_summ2['Checks (PRE)']


        def select_col(X):
            global c
            # COLORS
            c = ['background-color:  #EF5350',                  #>=50% RED
                    'background-color: #FFAF00',                #25-50% ORANGE
                        'background-color: #C0C000',            #10-25% YELLOW
                            'background-color: #00FF00',        #5-10% L GREEN
                                'background-color: #00AF5F',    #<5% GREEN
                                    '']
   
                
            mask_AllBlank = (X['All Blank'] == 'Yes')   
            mask_16 = (X['Inconsistent'] == 0)
            mask_15 = (X['Probable Reporting Error'] == 0)

            #DataFrame with same index and columns names as original filled empty strings
            df1 =  pd.DataFrame(c[5], X.index, columns=X.columns)
            df1.loc[mask_AllBlank, 'All Blank'] = c[0]
            df1.loc[mask_15, 'Probable Reporting Error'] = c[4]
            df1.loc[mask_16, 'Inconsistent'] = c[4]
            return df1

        final_result_summ2 = final_result_summ2.style.apply(select_col, axis=None)

        return final_result_summ1, final_result_summ2, dataframeForSheet4, dataframeForSheet5, dataframeForSheet6, dataframeForSheet7


    # EXPORT FILE
    # ===========
    def export(self):
        table_result1, table_result2, table_result3, table_result4, table_result5, table_result6  = self.summaryReport(df)

        # Rename orignal headers
        df.rename(res_dict , axis=1, inplace=True)

        # Remove column names month and year
        df.drop(['Month', 'Year'], axis = 1, inplace=True)

        new_list = [[""]]
        table_result_content = pd.DataFrame(new_list)


        # Save file dialog
        filename = QFileDialog.getSaveFileName(TabWidget, "Save to Excel", "Summary_Sheet",
                                                "Excel Spreadsheet (*.xlsx);;"
                                                "All Files (*)")[0]

        self.popup.start_progress()
        
        # Taking transpose of data 
        table_result3 = table_result3.T
        table_result4 = table_result4.T
        table_result5 = table_result5.T
        table_result6 = table_result6.T

        try:
        # exporting to excel
            with pd.ExcelWriter(filename) as writer: 
                table_result_content.to_excel(writer, sheet_name='Description', engine='openpyxl')
                table_result2.to_excel(writer, sheet_name='Facility Level Summary', engine='openpyxl')
                table_result1.to_excel(writer, sheet_name='Validation checkwise summary', engine='openpyxl')
                table_result3.to_excel(writer, sheet_name='Facility with Inconsistent', engine='openpyxl')
                table_result4.to_excel(writer, sheet_name='Facility with PRE', engine='openpyxl')
                table_result5.to_excel(writer, sheet_name='Checks giving Inconsistent', engine='openpyxl')
                table_result6.to_excel(writer, sheet_name='Checks giving PRE', engine='openpyxl')
                df.to_excel(writer, sheet_name='Validated_Data', engine='openpyxl')

        except:
            msg = QMessageBox()
            msg.setWindowTitle("Saving File Error Message / फ़ाइल सहेजें त्रुटि संदेश")
            msg.setIcon(QMessageBox.Critical)
            msg.setText(
                "The file is already opened , CLOSE IT ... / फ़ाइल पहले ही खोली जा चुकी है, इसे बंद करें...")
            msg.setText(
                "\n WINDOWS PERMISSION DENIED !  The file is already opened , CLOSE IT FIRST ... / विंडोज़ अनुमति अस्वीकृत! फ़ाइल पहले ही खोली जा चुकी है, इसे पहले बंद करें...")
            msg.exec()
        

        # PALETTES
        workbook = load_workbook(filename)
        sheet_0 = workbook['Description']
        sheet = workbook['Facility Level Summary']
        sheet_1 = workbook['Validation checkwise summary']
        sheet_2 = workbook['Facility with Inconsistent']
        sheet_3 = workbook['Facility with PRE']
        sheet_4 = workbook['Checks giving Inconsistent']
        sheet_5 = workbook['Checks giving PRE']
        sheet_6 = workbook['Validated_Data'] 

        # Activating sheets 
        workbook.active = sheet
        workbook.active = sheet_1

        sheet.sheet_view.showGridLines = False
        sheet_1.sheet_view.showGridLines = False

        # Function to generate sequences according to the excel sheet
        def excel_cols_link3():
            n1 = 1
            while True:
                yield from (''.join(group) for group in itertools.product(string.ascii_uppercase, repeat=n1))
                n1 += 1

        # Function to generate sequences according to the excel sheet
        def excel_cols_link4():
            n2 = 1
            while True:
                yield from (''.join(group) for group in itertools.product(string.ascii_uppercase, repeat=n2))
                n2 += 1
       
        cnt1, cnt2, cnt3, cnt4, cnt5 = 0, 0, 0, 0, 0
        for i in range(len(FList3)):
            if sheet.cell(row=i+2, column=10).value >= 50:
                cnt1 += 1
            elif sheet.cell(row=i+2, column=10).value < 50 and sheet.cell(row=i+2, column=10).value >= 25:
                cnt2 += 1
            elif sheet.cell(row=i+2, column=10).value < 25 and sheet.cell(row=i+2, column=10).value >= 10:
                cnt3 += 1
            elif sheet.cell(row=i+2, column=10).value < 10 and sheet.cell(row=i+2, column=10).value >= 5:
                cnt4 += 1
            elif sheet.cell(row=i+2, column=10).value < 5:
                cnt5 += 1

            for j in range(len(FList3[i])):
                '''
                # Create hyperlink to relevant cell
                '''
                link1 = "#'Checks giving Inconsistent'!B2"

                l1 = list(itertools.islice(excel_cols_link3(), len(FList3)+1))
                
                #update link
                link1 = link1.replace("B", l1[i+1])

                sheet.cell(row=i+2, column=7).hyperlink = link1
                sheet.cell(row=i+2, column=7).style = "Hyperlink"

                if sheet.cell(row=i+2, column=10).value >= 50:
                    sheet.cell(row=i+2, column=7).fill = PatternFill(fgColor='EF5350', fill_type = "solid")
                elif sheet.cell(row=i+2, column=10).value < 50 and sheet.cell(row=i+2, column=10).value >= 25:
                    sheet.cell(row=i+2, column=7).fill = PatternFill(fgColor='FFAF00', fill_type = "solid")
                elif sheet.cell(row=i+2, column=10).value < 25 and sheet.cell(row=i+2, column=10).value >= 10:
                    sheet.cell(row=i+2, column=7).fill = PatternFill(fgColor='C0C000', fill_type = "solid")
                elif sheet.cell(row=i+2, column=10).value < 10 and sheet.cell(row=i+2, column=10).value >= 5:
                    sheet.cell(row=i+2, column=7).fill = PatternFill(fgColor='00FF00', fill_type = "solid")
                elif sheet.cell(row=i+2, column=10).value < 5:
                    sheet.cell(row=i+2, column=7).fill = PatternFill(fgColor='00AF5F', fill_type = "solid")

        cnt6, cnt7, cnt8, cnt9, cnt10 = 0, 0, 0, 0, 0
        for i in range(len(FList4)):
            if sheet.cell(row=i+2, column=11).value >= 50:
                cnt6 += 1
            elif sheet.cell(row=i+2, column=11).value < 50 and sheet.cell(row=i+2, column=11).value >= 25:
                cnt7 += 1
            elif sheet.cell(row=i+2, column=11).value < 25 and sheet.cell(row=i+2, column=11).value >= 10:
                cnt8 += 1
            elif sheet.cell(row=i+2, column=11).value < 10 and sheet.cell(row=i+2, column=11).value >= 5:
                cnt9 += 1
            elif sheet.cell(row=i+2, column=11).value < 5:
                cnt10 += 1

            for j in range(len(FList4[i])):

                '''
                # Create hyperlink to relevant cell
                '''
                link2 = "#'Checks giving PRE'!B2"

                l2 = list(itertools.islice(excel_cols_link4(), len(FList4)+1))
                
                #update link
                link2 = link2.replace("B", l2[i+1])
                sheet_2.cell

                sheet.cell(row=i+2, column=8).hyperlink = link2
                sheet.cell(row=i+2, column=8).style = "Hyperlink"

                if sheet.cell(row=i+2, column=11).value >= 50:
                    sheet.cell(row=i+2, column=8).fill = PatternFill(fgColor='EF5350', fill_type = "solid")
                elif sheet.cell(row=i+2, column=11).value < 50 and sheet.cell(row=i+2, column=11).value >= 25:
                    sheet.cell(row=i+2, column=8).fill = PatternFill(fgColor='FFAF00', fill_type = "solid")
                elif sheet.cell(row=i+2, column=11).value < 25 and sheet.cell(row=i+2, column=11).value >= 10:
                    sheet.cell(row=i+2, column=8).fill = PatternFill(fgColor='C0C000', fill_type = "solid")
                elif sheet.cell(row=i+2, column=11).value < 10 and sheet.cell(row=i+2, column=11).value >= 5:
                    sheet.cell(row=i+2, column=8).fill = PatternFill(fgColor='00FF00', fill_type = "solid")
                elif sheet.cell(row=i+2, column=11).value < 5:
                    sheet.cell(row=i+2, column=8).fill = PatternFill(fgColor='00AF5F', fill_type = "solid")


        # Coloring and palettes of Facility Guidance Sheet
        sheet['O4'] = 'Color Brackets'
        sheet['O5'].fill = PatternFill(fgColor="EF5350", fill_type = "solid")
        sheet['O6'].fill = PatternFill(fgColor="FFAF00", fill_type = "solid")
        sheet['O7'].fill = PatternFill(fgColor="C0C000", fill_type = "solid")
        sheet['O8'].fill = PatternFill(fgColor="00FF00", fill_type = "solid")
        sheet['O9'].fill = PatternFill(fgColor="00AF5F", fill_type = "solid")

        sheet["P4"] = "Range"
        sheet["P5"] = ">= 50%"
        sheet["P6"] = "25 - 50%"
        sheet["P7"] = "10 - 25%"
        sheet["P8"] = "5 - 10%"
        sheet["P9"] = "< 5%"
        sheet["P10"] = "Total Facilities"

        sheet["Q4"] = "Inconsistent"
        sheet["Q5"] = cnt1
        sheet["Q6"] = cnt2
        sheet["Q7"] = cnt3
        sheet["Q8"] = cnt4
        sheet["Q9"] = cnt5
        sheet["Q10"] = cnt1 + cnt2 + cnt3 + cnt4 + cnt5

        sheet["R4"] = "Probable Reporting Error"
        sheet["R5"] = cnt6
        sheet["R6"] = cnt7
        sheet["R7"] = cnt8
        sheet["R8"] = cnt9
        sheet["R9"] = cnt10
        sheet["R10"] = cnt6 + cnt7 + cnt8 + cnt9 + cnt10

        ''' 
        GRAPH PLOTS
        '''
        
        Ranges = ['< 5%', '5 - 10%' , '10 - 25%', '25 - 50%', '>= 50%']
        Numbers_Inc = [cnt5, cnt4, cnt3, cnt2, cnt1]
        Numbers_PRE = [cnt10, cnt9, cnt8, cnt7, cnt6]
        
        # plotting a bar graph
        X_axis = np.arange(len(Ranges))

        plt.bar(X_axis - 0.2, Numbers_Inc, 0.4, label = 'Number of Inconsistents')
        plt.bar(X_axis + 0.2, Numbers_PRE, 0.4, label = 'Number of Probable Reporting Errors')
        # figure(figsize=(8, 8), dpi=50)
        plt.xticks(X_axis, Ranges)
        plt.xlabel("Condition")
        plt.ylabel("Number")
        plt.legend()
        plt.title('Error Summary Facility Wise')
        plt.savefig("myplot2.png", dpi = 80)

        img = openpyxl.drawing.image.Image('myplot2.png')
        img.anchor='N13'

        sheet.add_image(img)
        workbook.save(filename=filename)
        workbook.save(filename=filename)


        # Attention Required Sheet (Inconsistent)
        # =======================================

        workbook.active = sheet_2
        sheet_2.sheet_view.showGridLines = False

        def excel_cols_link1():
            n1 = 1
            while True:
                yield from (''.join(group) for group in itertools.product(string.ascii_uppercase, repeat=n1))
                n1 += 1

        def excel_cols_link2():
            n2 = 1
            while True:
                yield from (''.join(group) for group in itertools.product(string.ascii_uppercase, repeat=n2))
                n2 += 1
        
        cnt21, cnt22, cnt23, cnt24 = 0, 0, 0, 0
        for i in range(len(FList1)):
            if sheet_1.cell(row=i+2, column=6).value >= 25:
                cnt21 += 1
            elif sheet_1.cell(row=i+2, column=6).value < 25 and sheet_1.cell(row=i+2, column=6).value >= 10:
                cnt22 += 1
            elif sheet_1.cell(row=i+2, column=6).value < 10 and sheet_1.cell(row=i+2, column=6).value >= 5:
                cnt23 += 1
            elif sheet_1.cell(row=i+2, column=6).value < 5:
                cnt24 += 1

            for j in range(len(FList1[i])):
                sheet_2.cell(row=j+5,column=i+2).value = FList1[i][j]

                '''
                # Create hyperlink to relevant cell for sheet_1
                '''
                link1 = "#'Facility with Inconsistent'!B2"

                l1 = list(itertools.islice(excel_cols_link1(), len(FList1)+1))
                
                #update link
                link1 = link1.replace("B", l1[i+1])

                sheet_1.cell(row=i+2, column=4).hyperlink = link1
                sheet_1.cell(row=i+2, column=4).style = "Hyperlink"

                if sheet_1.cell(row=i+2, column=6).value >= 25:
                    sheet_1.cell(row=i+2, column=4).fill = PatternFill(fgColor='EF5350', fill_type = "solid")
                elif sheet_1.cell(row=i+2, column=6).value < 25 and sheet_1.cell(row=i+2, column=6).value >= 10:
                    sheet_1.cell(row=i+2, column=4).fill = PatternFill(fgColor='FFAF00', fill_type = "solid")
                elif sheet_1.cell(row=i+2, column=6).value < 10 and sheet_1.cell(row=i+2, column=6).value >= 5:
                    sheet_1.cell(row=i+2, column=4).fill = PatternFill(fgColor='C0C000', fill_type = "solid")
                elif sheet_1.cell(row=i+2, column=6).value < 5:
                    sheet_1.cell(row=i+2, column=4).fill = PatternFill(fgColor='00AF5F', fill_type = "solid")
                
                
            # Colors
            for k in range(1, len(FList1[i])+100):
                sheet_2.cell(row=4, column=i+1).alignment = Alignment(horizontal='center')
                sheet.cell(row=k, column=7).alignment = Alignment(horizontal='center')
                sheet.cell(row=k, column=8).alignment = Alignment(horizontal='center')
                sheet.cell(row=k, column=9).alignment = Alignment(horizontal='center')
                sheet_1.cell(row=k, column=4).alignment = Alignment(horizontal='center')
                sheet_1.cell(row=k, column=5).alignment = Alignment(horizontal='center')
                sheet_2.cell(row=2, column=i+2).fill = PatternFill(fgColor="fff5be", fill_type = "solid")
                sheet_2.cell(row=k+2, column=i+2).fill = PatternFill(fgColor="EF9A9A", fill_type = "solid")
                # set the width of the column
        
        workbook.save(filename=filename)


        # Attention Required Sheet (PRE)
        # =======================================
        workbook.active = sheet_3

        sheet_3.sheet_view.showGridLines = False

        cnt25, cnt26, cnt27, cnt28 = 0, 0, 0, 0
        for i in range(len(FList2)):
            if sheet_1.cell(row=i+2, column=7).value >= 25:
                cnt25 += 1
            elif sheet_1.cell(row=i+2, column=7).value < 25 and sheet_1.cell(row=i+2, column=7).value >= 10:
                cnt26 += 1
            elif sheet_1.cell(row=i+2, column=7).value < 10 and sheet_1.cell(row=i+2, column=7).value >= 5:
                cnt27 += 1
            elif sheet_1.cell(row=i+2, column=7).value < 5:
                cnt28 += 1

            for j in range(len(FList2[i])):
                sheet_3.cell(row=j+5,column=i+2).value = FList2[i][j]
                sheet_3.cell(row=4, column=i+1).alignment = Alignment(horizontal='center')
                sheet_3.cell(row=2, column=i+2).fill = PatternFill(fgColor="fff5be", fill_type = "solid")
                # Top 10 Facility Checks not working well
                sheet_3.cell(row=j+3, column=i+2).fill = PatternFill(fgColor="EF9A9A", fill_type = "solid")

                '''
                # Create hyperlink to relevant cell for sheet_1
                '''
                link1 = "#'Facility with PRE'!B2"

                l1 = list(itertools.islice(excel_cols_link2(), len(FList2)+1))
                
                #update link
                link1 = link1.replace("B", l1[i+1])

                sheet_1.cell(row=i+2, column=5).hyperlink = link1
                sheet_1.cell(row=i+2, column=5).style = "Hyperlink"

                if sheet_1.cell(row=i+2, column=7).value >= 25:
                    sheet_1.cell(row=i+2, column=5).fill = PatternFill(fgColor='EF5350', fill_type = "solid")
                elif sheet_1.cell(row=i+2, column=7).value < 25 and sheet_1.cell(row=i+2, column=7).value >= 10:
                    sheet_1.cell(row=i+2, column=5).fill = PatternFill(fgColor='FFAF00', fill_type = "solid")
                elif sheet_1.cell(row=i+2, column=7).value < 10 and sheet_1.cell(row=i+2, column=7).value >= 5:
                    sheet_1.cell(row=i+2, column=5).fill = PatternFill(fgColor='C0C000', fill_type = "solid")
                elif sheet_1.cell(row=i+2, column=7).value < 5:
                    sheet_1.cell(row=i+2, column=5).fill = PatternFill(fgColor='00AF5F', fill_type = "solid")
                
            # Colors
            for k in range(1, len(FList2[i])+100):
                sheet_3.cell(row=4, column=i+1).alignment = Alignment(horizontal='center')
                sheet_3.cell(row=2, column=i+2).fill = PatternFill(fgColor="fff5be", fill_type = "solid")


        # Coloring of Validation Summary Sheet
        sheet_1['J10'] = "Color Brackets"
        sheet_1['J11'].fill = PatternFill(fgColor="EF5350", fill_type = "solid")
        sheet_1['J12'].fill = PatternFill(fgColor="FFAF00", fill_type = "solid")
        sheet_1['J13'].fill = PatternFill(fgColor="C0C000", fill_type = "solid")
        sheet_1['J14'].fill = PatternFill(fgColor="00AF5F", fill_type = "solid")

        sheet_1['K10'] = "Range"
        sheet_1['K11'] = ">= 25%"
        sheet_1['K12'] = "10 - 25%"
        sheet_1['K13'] = "5 - 10%"
        sheet_1['K14'] = "< 5%"
        sheet_1['K15'] = "Total Indicators"

        sheet_1['L10'] = "Inconsistent"
        sheet_1['L11'] = cnt21
        sheet_1['L12'] = cnt22
        sheet_1['L13'] = cnt23
        sheet_1['L14'] = cnt24
        sheet_1['L15'] = cnt21 + cnt22 + cnt23 + cnt24

        sheet_1['M10'] = "Probable Reporting Error"
        sheet_1['M11'] = cnt25
        sheet_1['M12'] = cnt26
        sheet_1['M13'] = cnt27
        sheet_1['M14'] = cnt28
        sheet_1['M15'] = cnt25 + cnt26 + cnt27 + cnt28
        
        workbook.save(filename=filename)


        # Checks Sheet (Inconsistent)
        # =======================================
        workbook.active = sheet_4

        sheet_4.sheet_view.showGridLines = False
        for i in range(len(FList3)):

            for j in range(len(FList3[i])):
                sheet_4.cell(row=j+4,column=i+2).value = FList3[i][j]

            # Colors
            for k in range(1, len(FList3[i])+5):
                sheet_4.cell(row=2, column=i+2).fill = PatternFill(fgColor="fff5be", fill_type = "solid")
                sheet_4.cell(row=k+2, column=i+2).fill = PatternFill(fgColor="EF9A9A", fill_type = "solid")

        workbook.save(filename=filename)

        # Checks Sheet (PRE)
        # =======================================
        workbook.active = sheet_5

        sheet_5.sheet_view.showGridLines = False
        for i in range(len(FList4)):
            for j in range(len(FList4[i])):
                sheet_5.cell(row=j+4,column=i+2).value = FList4[i][j]

            # Colors
            for k in range(1, len(FList4[i])+5):
                sheet_5.cell(row=4, column=i+1).alignment = Alignment(horizontal='center')
                sheet_5.cell(row=2, column=i+2).fill = PatternFill(fgColor="fff5be", fill_type = "solid")
                # Top 10 Facility Checks not working well
                sheet_5.cell(row=k+2, column=i+2).fill = PatternFill(fgColor="EF9A9A", fill_type = "solid")

        workbook.save(filename=filename)

        # Coloring validated data sheet
        workbook.active = sheet_6

        for i in range(len(df.axes[0])):
            for j in range(len(df.axes[1])):
                if sheet_6.cell(row=i+2, column=j+2) == 'Inconsistent':
                    sheet_6.cell(row=i+2, column=j+2).fill = PatternFill(fgColor="EF9A9A", fill_type = "solid")

        workbook.save(filename = filename)



        # Handling contents sheet
        from openpyxl.styles.borders import Border, Side
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))


        workbook.active = sheet_0

        for i in range(4, 13):
            sheet_0.cell(row=i+1, column=2).border = thin_border
            sheet_0.cell(row=i+1, column=3).border = thin_border

        sheet_0.sheet_view.showGridLines = False
        sheet_0.cell(row=3, column=3).alignment = Alignment(horizontal='right')
        sheet_0['B4'] = "   Data Validation Check Tool"
        sheet_0['B5'] = "Facility Type:" + self.lineEdit_5.text()
        sheet_0['B6'] = "Sheet Name"
        sheet_0['B7'] = "Facility Level Summary"
        sheet_0['B8'] = "Validation checkwise summary"
        sheet_0['B9'] = "Facility with Inconsistent"
        sheet_0['B10'] = "Facility with PRE"
        sheet_0['B11'] = "Checks giving Inconsistent"
        sheet_0['B12'] = "Checks giving PRE"
        sheet_0['B13'] = "Validated Data"
        sheet_0['B16'] = "Important Terminologies"
        sheet_0['B17'] = "Consistent:"
        sheet_0['B18'] = "Inconsistent:"
        sheet_0['B19'] = "Probable reporting error:"
        sheet_0['B20'] = "Blank:"

        sheet_0['C4'] = ""
        sheet_0['C5'] = "Duration:" + self.lineEdit_6.text()
        sheet_0['C6'] = "Sheet Description"
        sheet_0['C7'] = "This sheet gives the counts of errors corresponding to each facility name. The colour coding is done as per the buckets, considering the percentage of the number of inconsistent/PRE out of the total validation checks in that facility type for each facility name. This is also shown graphically below the bucketing table. Clicking on the figures of the original table, the user will be directed to the “Checks giving inconsistent” or “Checks giving PRE” tabs, showing the validation checks for which, the errors creeped in."
        sheet_0['C8'] = "This sheet gives us the count of the facilities giving inconsistent and probable reporting error within the datasets. The colour coding is done as per the buckets, which were created considering the percentage of the number of facility names out of all the facilities who have reported inconsistent/PRE. This is also shown graphically below the bucketing table. Clicking on the figures of the original table, the user will be directed to the “Facility with inconsistent” or “Facility with PRE” tabs, showing the facilities for which, the errors creeped in."
        sheet_0['C9'] = "This sheet provides the list of the facilities giving Inconsistencies within the dataset with the count and the condition and the description of the condition."
        sheet_0['C10'] = "This sheet provides the list of the facilities giving probable reporting error within the dataset with the count and the condition and the description of the condition."
        sheet_0['C11'] = "This sheet provides the list of the checks giving Inconsistencies within the dataset with the count and the description."
        sheet_0['C12'] = "This sheet provides the list of the checks giving Probable Reporting Error within the dataset with the count and the description"
        sheet_0['C13'] = "This sheet is the complete raw data with the checks embedded with the sheet which gives a detailed information, row wise identification of the inconsistent and PRE with the problematic indicators because of which the user is getting errors in its dataset"
        sheet_0['C16'] = ""
        sheet_0['C17'] = "The validation check holds true and needs no scrutiny."
        sheet_0['C18'] = "The validation check fails and the inconsistent data item is flagged. "
        sheet_0['C19'] = "The validation check may fail and it is subject to confirmation with the concerned authority. Check and verify."
        sheet_0['C20'] = "The validation where all the items are blank."

        sheet_0['B4'].font = Font(size = 14, bold = True)
        sheet_0['B16'].font = Font(size = 14, bold = True)
        sheet_0['B4'].fill = PatternFill(fgColor="FFFFCC", fill_type = "solid")
        sheet_0['B16'].fill = PatternFill(fgColor="C0C000", fill_type = "solid")
        # sheet_0['C4'].fill = PatternFill(fgColor="FFFFCC", fill_type = "solid")
        sheet_0['B5'].fill = PatternFill(fgColor="00FF00", fill_type = "solid")
        sheet_0['C5'].fill = PatternFill(fgColor="00FF00", fill_type = "solid")

        # delete useless columns of Facility Level Summary
        sheet.delete_cols(10, 2)
        # delete useless columns of sheet_1
        sheet_1.delete_cols(6, 2)
        workbook.save(filename=filename)

        self.popup.close()

        # Create the messagebox object
        self.msg = QMessageBox()
        # Set the information icon
        self.msg.setWindowIcon(QtGui.QIcon('checked.png'))
        self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
        # Set the main message
        self.msg.setText("Excel file downloaded in the selected location \n\n एक्सेल फ़ाइल चयनित स्थान पर डाउनलोड की गई।")
        # Set the title of the window
        self.msg.setWindowTitle(" ")
        # Display the message box
        self.msg.show()

        os.remove("FileName.csv")
        os.remove("myplot2.png")

    # Reset
    def reset(self):
        QtCore.QCoreApplication.quit()
        status = QtCore.QProcess.startDetached(sys.executable, sys.argv)
      
        print('Please WAIT the data and settings are resetting ..... ')
        print('After resetting TOOL will open again .....')


    # # RESET FUNCTION
    # def reset(self):
    #     self.lineEdit.clear()
    #     self.lineEdit_2.clear()
    #     self.lineEdit_3.clear()
    #     self.pushButton.setEnabled(True)
    #     self.pushButton_2.setEnabled(True)
    #     self.pushButton_3.setEnabled(True)
    #     self.pushButton_4.setEnabled(True)
    #     self.pushButton_5.setEnabled(True)
    #     self.pushButton_6.setEnabled(True)
    #     self.pushButton_7.setEnabled(True)
    #     self.pushButton_3.setText('-- All Selected --')
    #     self.pushButton_4.setText('-- All Selected --')
    #     self.pushButton_5.setText('-- All Selected --')
    #     self.pushButton_6.setText('-- All Selected --')


    #     # Create the messagebox object
    #     self.msg = QMessageBox()
    #     # Set the information icon
    #     self.msg.setWindowIcon(QtGui.QIcon('checked.png'))
    #     self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
    #     # Set the main message
    #     self.msg.setText("Reset Successful , Now you can upload data again.")
    #     # Set the title of the window
    #     self.msg.setWindowTitle("Reset Successful Message")
    #     # Display the message box
    #     self.msg.show()

    # Display methodology pdf in browser
    def methodology(self):
        os.system('start Steps.pdf')

    # Display methodology pdf in browser
    def UserManual(self):
        os.system('start Manual.pdf')

    # Data Extraction from HMIS website pdf in browser
    def data_extraction(self):
        os.system('data_extraction.pdf')


################################### Main Function ############################
if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    TabWidget = QtWidgets.QTabWidget()
    ui = Ui_TabWidget()
    ui.setupUi(TabWidget)
    TabWidget.show()
    sys.exit(app.exec_())