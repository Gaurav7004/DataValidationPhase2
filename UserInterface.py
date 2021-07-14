import re, jinja2
from typing_extensions import final
import numpy as np
import pandas as pd
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QFileDialog,QMessageBox
import matplotlib.pyplot as plt
from matplotlib.pyplot import figure
from openpyxl.styles import PatternFill, Alignment, Border, Font
from HealthSubCentreVal import *
from PrimaryHealthCentre import *
from SubDistrictHospitalVal import *
from DistrictHospitalVal import *
from CommunityHealthCentreVal import *

hiddenimports=['pandas', 'jinja2']

# ---------------------------------------------------- UI -----------------------------------------------
class Ui_Dialog(object):
    def setupUi(self, Dialog):
        Dialog.setObjectName("Dialog")
        Dialog.resize(1536, 860)
        Dialog.setStyleSheet("")
        self.horizontalLayout = QtWidgets.QHBoxLayout(Dialog)
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.widget = QtWidgets.QWidget(Dialog)
        self.widget.setStyleSheet("QWidget{\n"
"background-color: rgb(255, 255, 255);}")
        self.widget.setObjectName("widget")
        self.frame = QtWidgets.QFrame(self.widget)
        self.frame.setGeometry(QtCore.QRect(10, 10, 1491, 91))
        self.frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame.setObjectName("frame")
        self.verticalLayoutWidget_2 = QtWidgets.QWidget(self.frame)
        self.verticalLayoutWidget_2.setGeometry(QtCore.QRect(0, 0, 1491, 91))
        self.verticalLayoutWidget_2.setObjectName("verticalLayoutWidget_2")
        self.gridLayout = QtWidgets.QGridLayout(self.verticalLayoutWidget_2)
        self.gridLayout.setContentsMargins(0, 0, 0, 0)
        self.gridLayout.setObjectName("gridLayout")
        self.label = QtWidgets.QLabel(self.verticalLayoutWidget_2)
        font = QtGui.QFont()
        font.setPointSize(28)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setStyleSheet("QLabel{background-color: #00A5DB; color : white;border: 3px solid black;}\n"
"\n"
"")
        self.label.setObjectName("label")
        self.gridLayout.addWidget(self.label, 0, 0, 1, 1)
        self.frame_3 = QtWidgets.QFrame(self.widget)
        self.frame_3.setGeometry(QtCore.QRect(10, 110, 1491, 721))
        self.frame_3.setMinimumSize(QtCore.QSize(1331, 561))
        self.frame_3.setStyleSheet("")
        self.frame_3.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_3.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_3.setObjectName("frame_3")
        self.pushButton = QtWidgets.QPushButton(self.frame_3)
        self.pushButton.clicked.connect(self.upload)
        self.pushButton.setGeometry(QtCore.QRect(20, 10, 271, 61))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton.setFont(font)
        self.pushButton.setStyleSheet("background-color: #73C067;\n"
"color: white;\n"
"border: 3px solid black;")
        self.pushButton.setObjectName("pushButton")
        self.lineEdit = QtWidgets.QLineEdit(self.frame_3)
        self.lineEdit.setGeometry(QtCore.QRect(310, 10, 741, 61))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit.setFont(font)
        self.lineEdit.setStyleSheet("background-color: white;\n"
"color: #656565;\n"
"border: 3px solid black;")
        self.lineEdit.setObjectName("lineEdit")
        self.pushButton_2 = QtWidgets.QPushButton(self.frame_3)
        self.pushButton_2.clicked.connect(self.VerifyFType)
        self.pushButton_2.setGeometry(QtCore.QRect(700, 170, 351, 61))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_2.setFont(font)
        self.pushButton_2.setStyleSheet("background-color: #00A5DB;color: white;border: 3px solid black;\n"
"")
        self.pushButton_2.setObjectName("pushButton_2")
        self.frame_4 = QtWidgets.QFrame(self.frame_3)
        self.frame_4.setGeometry(QtCore.QRect(20, 270, 661, 441))
        self.frame_4.setStyleSheet("QLabel{background-color: rgb(185, 255, 238);}\n"
"\n"
"")
        self.frame_4.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_4.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_4.setObjectName("frame_4")
        self.pushButton_3 = QtWidgets.QPushButton(self.frame_4)
        self.pushButton_3.installEventFilter(Dialog)
        self.pushButton_3.clicked.connect(self.onSelectDistrict)
        self.pushButton_3.setGeometry(QtCore.QRect(330, 140, 311, 51))
        self.pushButton_3.setStyleSheet("border: 3px solid black;\n"
"background-color: rgb(222, 222, 222);\n"
"")
        self.pushButton_3.setObjectName("pushButton_3")
        self.pushButton_4 = QtWidgets.QPushButton(self.frame_4)
        self.pushButton_4 = QtWidgets.QPushButton(self.frame_4)
        self.pushButton_4.clicked.connect(self.onSelectRuralUrban)
        self.pushButton_4.setGeometry(QtCore.QRect(330, 220, 311, 51))
        self.pushButton_4.setStyleSheet("border: 3px solid black;\n"
"background-color: rgb(222, 222, 222);")
        self.pushButton_4.setObjectName("pushButton_4")
        self.pushButton_5 = QtWidgets.QPushButton(self.frame_4)
        self.pushButton_5.installEventFilter(Dialog)
        self.pushButton_5.clicked.connect(self.onSelectOwnership)
        self.pushButton_5.setGeometry(QtCore.QRect(330, 300, 311, 51))
        self.pushButton_5.setStyleSheet("border: 3px solid black;\n"
"background-color: rgb(222, 222, 222);\n"
"")
        self.pushButton_5.setObjectName("pushButton_5")
        self.pushButton_6 = QtWidgets.QPushButton(self.frame_4)
        self.pushButton_6.installEventFilter(Dialog)
        self.pushButton_6.clicked.connect(self.onSelectFacilityName)
        self.pushButton_6.setGeometry(QtCore.QRect(330, 380, 311, 51))
        self.pushButton_6.setStyleSheet("border: 3px solid black;\n"
"background-color: rgb(222, 222, 222);")
        self.pushButton_6.setObjectName("pushButton_6")
        self.label_4 = QtWidgets.QLabel(self.frame_4)
        self.label_4.setGeometry(QtCore.QRect(10, 140, 251, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_4.setFont(font)
        self.label_4.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_4.setObjectName("label_4")
        self.label_5 = QtWidgets.QLabel(self.frame_4)
        self.label_5.setGeometry(QtCore.QRect(10, 220, 251, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_5.setFont(font)
        self.label_5.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_5.setObjectName("label_5")
        self.label_6 = QtWidgets.QLabel(self.frame_4)
        self.label_6.setGeometry(QtCore.QRect(10, 300, 251, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_6.setFont(font)
        self.label_6.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_6.setObjectName("label_6")
        self.label_7 = QtWidgets.QLabel(self.frame_4)
        self.label_7.setGeometry(QtCore.QRect(10, 380, 251, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_7.setFont(font)
        self.label_7.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_7.setObjectName("label_7")
        self.label_8 = QtWidgets.QLabel(self.frame_4)
        self.label_8.setGeometry(QtCore.QRect(230, 30, 271, 51))
        font = QtGui.QFont()
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_8.setFont(font)
        self.label_8.setStyleSheet("background-color: white;\n"
"color: grey;")
        self.label_8.setObjectName("label_8")
        self.pushButton_7 = QtWidgets.QPushButton(self.frame_3)
        self.pushButton_7.clicked.connect(self.export)
        self.pushButton_7.setGeometry(QtCore.QRect(1020, 390, 231, 71))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_7.setFont(font)
        self.pushButton_7.setStyleSheet("background-color: #73C067;color: white;border: 3px solid black;")
        self.pushButton_7.setObjectName("pushButton_7")
        self.pushButton_8 = QtWidgets.QPushButton(self.frame_3)
        self.pushButton_8.clicked.connect(self.reset)
        self.pushButton_8.setGeometry(QtCore.QRect(1020, 550, 231, 71))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_8.setFont(font)
        self.pushButton_8.setStyleSheet("background-color: #B00020;color: white;border: 3px solid black;")
        self.pushButton_8.setObjectName("pushButton_8")
        self.lineEdit_2 = QtWidgets.QLineEdit(self.frame_3)
        self.lineEdit_2.setGeometry(QtCore.QRect(20, 170, 441, 61))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_2.setFont(font)
        self.lineEdit_2.setStyleSheet("background-color: white;\n"
"color: #656565;\n"
"border: 3px solid black;")
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.label_9 = QtWidgets.QLabel(self.frame_3)
        self.label_9.setGeometry(QtCore.QRect(1200, 670, 281, 31))
        font = QtGui.QFont()
        font.setFamily("MS Sans Serif")
        font.setPointSize(10)
        font.setItalic(True)
        self.label_9.setFont(font)
        self.label_9.setStyleSheet("color:#00395D;")
        self.label_9.setObjectName("label_9")
        self.label_2 = QtWidgets.QLabel(self.frame_3)
        self.label_2.setGeometry(QtCore.QRect(20, 90, 611, 31))
        font = QtGui.QFont()
        font.setFamily("MS Sans Serif")
        font.setPointSize(10)
        font.setBold(False)
        font.setItalic(True)
        font.setWeight(50)
        self.label_2.setFont(font)
        self.label_2.setStyleSheet("QLabel{color : red;}")
        self.label_2.setObjectName("label_2")
        self.pushButton_9 = QtWidgets.QPushButton(self.frame_3)
        self.pushButton_9.clicked.connect(self.methodology)
        self.pushButton_9.setGeometry(QtCore.QRect(1260, 10, 221, 61))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_9.setFont(font)
        self.pushButton_9.setStyleSheet("background-color: #F47B1F;\n"
"color: white;\n"
"border: 3px solid black;")
        self.pushButton_9.setObjectName("pushButton_9")
        self.pushButton_10 = QtWidgets.QPushButton(self.frame_3)
        self.pushButton_10.clicked.connect(self.UserManual)
        self.pushButton_10.setGeometry(QtCore.QRect(1260, 100, 221, 61))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_10.setFont(font)
        self.pushButton_10.setStyleSheet("background-color: #F47B1F;\n"
"color: white;\n"
"border: 3px solid black;")
        self.pushButton_10.setObjectName("pushButton_10")
        self.label_3 = QtWidgets.QLabel(self.frame_3)
        self.label_3.setGeometry(QtCore.QRect(700, 240, 661, 21))
        font = QtGui.QFont()
        font.setFamily("MS Sans Serif")
        font.setPointSize(10)
        font.setBold(False)
        font.setItalic(True)
        font.setWeight(50)
        self.label_3.setFont(font)
        self.label_3.setStyleSheet("QLabel{color : red;}")
        self.label_3.setObjectName("label_3")
        self.label_10 = QtWidgets.QLabel(self.frame_3)
        self.label_10.setGeometry(QtCore.QRect(1020, 470, 551, 21))
        font = QtGui.QFont()
        font.setFamily("MS Sans Serif")
        font.setPointSize(10)
        font.setBold(False)
        font.setItalic(True)
        font.setWeight(50)
        self.label_10.setFont(font)
        self.label_10.setStyleSheet("QLabel{color : red;}")
        self.label_10.setObjectName("label_10")
        self.lineEdit_3 = QtWidgets.QLineEdit(self.frame_3)
        self.lineEdit_3.setGeometry(QtCore.QRect(480, 170, 201, 61))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_3.setFont(font)
        self.lineEdit_3.setStyleSheet("background-color: white;\n"
"color: #656565;\n"
"border: 3px solid black;")
        self.lineEdit_3.setText("")
        self.lineEdit_3.setObjectName("lineEdit_3")
        self.horizontalLayout.addWidget(self.widget)

        self.retranslateUi(Dialog)
        QtCore.QMetaObject.connectSlotsByName(Dialog)

    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", "Dialog"))
        self.label.setText(_translate("Dialog", "                            Data Validation Check Tool"))
        self.pushButton.setText(_translate("Dialog", "Upload"))
        self.lineEdit.setPlaceholderText(_translate("Dialog", "  Your uploaded file name will display here ..."))
        self.pushButton_2.setText(_translate("Dialog", "Validate"))
        self.pushButton_3.setText(_translate("Dialog", "-- All Selected --"))
        self.pushButton_4.setText(_translate("Dialog", "-- All Selected --"))
        self.pushButton_5.setText(_translate("Dialog", "-- All Selected --"))
        self.pushButton_6.setText(_translate("Dialog", "-- All Selected --"))
        self.label_4.setText(_translate("Dialog", "             District"))
        self.label_5.setText(_translate("Dialog", "          Rural/Urban"))
        self.label_6.setText(_translate("Dialog", "           Ownership"))
        self.label_7.setText(_translate("Dialog", "         Facility Name"))
        self.label_8.setText(_translate("Dialog", "  Select Filters"))
        self.pushButton_7.setText(_translate("Dialog", "Export"))
        self.pushButton_8.setText(_translate("Dialog", "Reset"))
        self.lineEdit_2.setPlaceholderText(_translate("Dialog", "  Facility Type selected will display here ..."))
        self.label_9.setText(_translate("Dialog", "@Powered by NDQF"))
        self.label_2.setText(_translate("Dialog", "* Upload data in .xls / .xlsx format for one month facility wise."))
        self.pushButton_9.setText(_translate("Dialog", "Methodology"))
        self.pushButton_10.setText(_translate("Dialog", "User Manual"))
        self.label_3.setText(_translate("Dialog", "* Press Validate button to perform validation check on your data"))
        self.label_10.setText(_translate("Dialog", "* Press Export button to export your validated data."))
        self.lineEdit_3.setPlaceholderText(_translate("Dialog", " Month, Year"))

    ''' Upload Function '''
    # =====================

    def upload(self):
        global df_, res_dict

        # Validation for uploaded valid excel file
        try:
            # Upload file by opening filedialog
            fileName,_ = QFileDialog.getOpenFileName( Dialog, "Open Excel", (QtCore.QDir.homePath()), "Excel (*.xls *.xlsx)")

            # Read uploaded excel file
            df_ = pd.read_excel(fileName)

            # Converted again to csv file
            df_.to_csv("FileName.csv")

            # Read converted csv file
            df_ = pd.read_csv("FileName.csv")
        except:
            msg = QMessageBox()
            msg.setWindowTitle("Uploaded File Error Message")
            msg.setText(
                "The file which you have uploaded is not in the valid format of excel, Please upload valid excel file")
            msg.exec()

            try:
                # Upload file by opening filedialog
                fileName, _ = QFileDialog.getOpenFileName(
                    Dialog, "Open Excel", (QtCore.QDir.homePath()), "Excel (*.xls *.xlsx)")

                # Read uploaded excel file
                df_ = pd.read_excel(fileName)

                # Converted again to csv file
                df_.to_csv("FileName.csv")

                # Read converted csv file
                df_ = pd.read_csv("FileName.csv")
            except:
                msg = QMessageBox()
                msg.setWindowTitle("Uploaded File Error Message")
                msg.setText("Please upload valid excel file, Try again!")
                msg.exec()

        
        self.lineEdit.setText(fileName)

        # Dropping last two rows
        df_.drop(df_.index[[-1, -2]], inplace=True)

        # Extracting string from 1st cell of dataframe
        str_to_extr_MonthYear = str(df_.iloc[0])

        # grab the first row for the header
        new_header = df_.iloc[1]

        # #take the data less the header row
        df_ = df_[1:]

        # set the header row as the df header
        df_.columns = new_header

        # Extracting Month , Year from string
        results = re.findall(
            r"[abceglnoprtuvyADFJMNOS|]{3}[\s-]\d{2,4}", str_to_extr_MonthYear)
        print(results)

        # Splitting Month and Year
        MYList = results[0].split('-')
        print(MYList)

        # Partial list of headers
        lst1 = df_.columns[:21].values

        # Picking row items after 18th row to merge with lst1
        lst2 = df_.iloc[1, 21:].to_numpy()

        # Merging both lists
        lst3 = np.concatenate((lst1, lst2))

        # Assign lst3 as new column header
        df_.columns = lst3

        # Taking DataFrame from second row
        df_ = df_[2:]

        # Insering Month and Year to the orignal dataframe
        df_.insert(1, 'Month', MYList[0])
        df_.insert(2, 'Year', MYList[1])

        # Removing A column named as # coming from orignal data
        df_ = df_.loc[:, df_.columns != '#']

        # Reindexing dataframe
        df_ = df_.reset_index(drop=True)

        df_ = df_.iloc[:, 1:]

        # Temporary column to verify modified checks
        temp_columns = ['col_' + str(index)
                        for index in range(1, len(df_.columns)+1)]

        # Merging and converting temp_columns to orignal header to dictionary
        res_dict = {temp_columns[i]: df_.columns[i]
                    for i in range(len(temp_columns))}

        # Picking the temporary column names and renaming column headers with it
        df_.columns = [i for i in res_dict.keys()]

        # Orignal Header
        df_OrgHeaders = [i for i in res_dict.values()]

        # convert the set to the list and fill inside comboBox to select facility type
        list_set = df_['col_12'].tolist()
        unique_list = set(list_set)

        self.lineEdit_2.setText(["{0}".format(col) for col in unique_list][0])
        self.lineEdit_3.setText(results[0])

        # Disabling upload Button
        self.pushButton.setDisabled(True)


        # Create the messagebox object
        self.msg = QMessageBox()
        # Set the information icon
        self.msg.setWindowIcon(QtGui.QIcon('checked.png'))
        self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
        # Set the main message
        self.msg.setText("The file has been uploaded successfully.")
        # Set the title of the window
        self.msg.setWindowTitle("Upload Successful Message")
        # Display the message box
        self.msg.show()
        return df_


    # Upload file button functionality
    # ================================
    def loadFile(self, df_):
        return df_


    # Filtering Facility Type
    # =======================
    def VerifyFType(self):
        global df, FType

        #df = self.loadFile(df_)
        FType = self.lineEdit_2.text()

        if (FType == 'Primary Health Centre'):
            print('Facility Type - ',FType)

            # Signaling PHC_Validate function i.e function where validation checks are present
            df = PHC_Validate(self, df_)

        elif (FType == 'Health Sub Centre' ):
            print('Facility Type - ', FType)
            # Signaling HSC_Validate function i.e function where validation checks are present
            df = HSC_Validate(self, df_)

        elif (FType == 'District Hospital'):
            print('Facility Type - ',FType)
            df = DH_Validate(self, df_)

        elif (FType == 'Sub District Hospital'):
            print('Facility Type - ',FType)
            df = SDH_Validate(self, df_)

        elif (FType == 'Community Health Centre'):
            print('Facility Type - ',FType)
            df = CHC_Validate(self, df_)

        else:
            raise Exception('Facility Type Name is not matching')

        self.pushButton_2.setEnabled(False)


    '''
    # Filter to decide which filter button user clicked
    # =================================================
    '''
    def eventFilter(self, target, event):
        if target == self.pushButton_3 and event.type() == QtCore.QEvent.MouseButtonPress:
            print("Checked District")
            self.pushButton_3.clicked.connect(self.onSelectDistrict)
            return True

        elif target == self.pushButton_4 and event.type() == QtCore.QEvent.MouseButtonPress:
            print("Checked Facility Name")
            self.pushButton_4.clicked.connect(self.onSelectFacilityName)
            return True

        elif target == self.pushButton_5 and event.type() == QtCore.QEvent.MouseButtonPress:
            print("Checked Rural/Urban")
            self.pushButton_5.clicked.connect(self.onSelectRuralUrban)
            return True

        elif target == self.pushButton_6 and event.type() == QtCore.QEvent.MouseButtonPress:
            print("Checked Ownership")
            self.pushButton_6.clicked.connect(self.onSelectOwnership)
            return True

        return False


    ################################################################################
    # Select District

    # Filter District Functionality
    def onSelectDistrict(self, index):
        self.keywords = dict([(i, []) for i in range(df.shape[0])])
        print(self.keywords)
        self.menu = QtWidgets.QMenu(Dialog)
        self.menu.setStyleSheet('QMenu { menu-scrollable: true; width: 200 }')
        font = self.menu.font()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.menu.setFont(font)

        index = 5
        self.col = index

        data_unique = []

        self.checkBoxs = []

        # Selectall added into Dropdown
        checkBox = QtWidgets.QCheckBox("Select all/ Deselect all", self.menu)
        checkBox.setStyleSheet("color: red; spacing: 5px; font-size:15px;")
        # All the checkboxes are enabled to check
        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(checkBox)
        self.menu.addAction(checkableAction)
        checkBox.setChecked(True)
        checkBox.stateChanged.connect(self.slotSelectDistrict)

        # list storing state data
        df['col_5'].fillna('Blank', inplace = True)
        list_set = df['col_5'].to_list()

        item = list_set
        item = sorted(list_set, key=str.upper)

        # looping to fill checkboxes, initially all checkboxes will be checked
        for i in range(len(item)):
            if item[i] not in data_unique:
                data_unique.append(item[i])
                checkBox = QtWidgets.QCheckBox(item[i], self.menu)
                checkBox.setChecked(True)
                checkableAction = QtWidgets.QWidgetAction(self.menu)
                checkableAction.setDefaultWidget(checkBox)
                self.menu.addAction(checkableAction)
                self.checkBoxs.append(checkBox)

        # Ok, cancel button
        btn = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel,
                                            QtCore.Qt.Vertical, self.menu)

        # ok selected
        btn.accepted.connect(self.menuCloseDistrict)
        # rejected , nothing selected
        btn.rejected.connect(self.menu.close)

        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(btn)
        self.menu.addAction(checkableAction)
        self.pushButton_3.setMenu(self.menu)

    # method to check -> uncheck and vice versa
    def slotSelectDistrict(self, state):
        for checkbox in self.checkBoxs:
            checkbox.setChecked(QtCore.Qt.Checked == state)

    # after ok selected
    def menuCloseDistrict(self):
        self.keywords[self.col] = []
        for element in self.checkBoxs:
            if element.isChecked():
                self.keywords[self.col].append(element.text())
        print(self.keywords[self.col])
        self.filterdataDistrict()
        self.pushButton_3.setText('Selected')
        self.pushButton_3.setEnabled(False)
        self.menu.close()

    # Filter data columnwise
    def filterdataDistrict(self):
        global df
        #keywords = dict([(i, []) for i in range(self.filterall.columnCount())])
        columnsShow = dict([(i, True) for i in range(df['col_5'].shape[0])])
        print(columnsShow)

        j = 0
        for j in range(df['col_5'].shape[0]):
            item = df['col_5'].to_list()

            # if self.keywords[self.col]:
            if item[j] not in self.keywords[self.col]:
                columnsShow[j] = False

        # for key, value in columnsShow.items():
        final_lst = [i for i in columnsShow.values()]
        df = df[final_lst]
        return df


    ################################################################################
    # Filter Facility Name

    # Filter FacilityName Functionality

    def onSelectFacilityName(self, index):
        self.keywords = dict([(i, []) for i in range(df.shape[0])])
        print(self.keywords)
        self.menu = QtWidgets.QMenu(Dialog)
        self.menu.setStyleSheet('QMenu { menu-scrollable: 1; width: 400 }')
        font = self.menu.font()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.menu.setFont(font)

        index = 14
        self.col = index

        data_unique = []

        self.checkBoxs = []

        # Selectall added into Dropdown
        checkBox = QtWidgets.QCheckBox("Select all/ Deselect all", self.menu)
        checkBox.setStyleSheet("color: red; spacing: 5px; font-size:15px;")
        # All the checkboxes are enabled to check
        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(checkBox)
        self.menu.addAction(checkableAction)
        checkBox.setChecked(True)
        checkBox.stateChanged.connect(self.slotSelectFacilityName)

        # list storing Facility Name data
        df['col_14'].fillna('Blank', inplace = True)
        list_set = df['col_14'].to_list()

        item = sorted(list_set, key=str.upper)

        # looping to fill checkboxes, initially all checkboxes will be checked
        for i in range(len(item)):
            if item[i] not in data_unique:
                data_unique.append(item[i])
                checkBox = QtWidgets.QCheckBox(item[i], self.menu)
                checkBox.setChecked(True)
                checkableAction = QtWidgets.QWidgetAction(self.menu)
                checkableAction.setDefaultWidget(checkBox)
                self.menu.addAction(checkableAction)
                self.checkBoxs.append(checkBox)

        # Ok, cancel button
        btn = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel,
                                            QtCore.Qt.Vertical, self.menu)

        # ok selected
        btn.accepted.connect(self.menuCloseFacilityName)
        # rejected , nothing selected
        btn.rejected.connect(self.menu.close)

        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(btn)
        self.menu.addAction(checkableAction)

        ############# Always set Pushbutton ####################
        self.pushButton_6.setMenu(self.menu)


    # method to check -> uncheck and vice versa
    def slotSelectFacilityName(self, state):
        for checkbox in self.checkBoxs:
            checkbox.setChecked(QtCore.Qt.Checked == state)

    # after ok selected
    def menuCloseFacilityName(self):
        self.keywords[self.col] = []
        for element in self.checkBoxs:
            if element.isChecked():
                self.keywords[self.col].append(element.text())
        print(self.keywords[self.col])
        self.filterdataFacilityName()
        self.pushButton_6.setText('Selected')
        self.pushButton_6.setEnabled(False)
        self.menu.close()

    # Filter data columnwise
    def filterdataFacilityName(self):
        global df
        #keywords = dict([(i, []) for i in range(self.filterall.columnCount())])
        columnsShow = dict([(i, True) for i in range(df['col_14'].shape[0])])
        print(columnsShow)

        j = 0
        for j in range(df['col_14'].shape[0]):
            item = df['col_14'].to_list()

            # if self.keywords[self.col]:
            if item[j] not in self.keywords[self.col]:
                columnsShow[j] = False

        # for key, value in columnsShow.items():
        final_lst = [i for i in columnsShow.values()]
        print(final_lst, 'this is final list of Select District')

        # matching list of facility type with col of dataframe returned by onSelectDistrict fun
        df = df[final_lst]
        return df


    ################################################################################
    # Filter Rural/Urban

    def onSelectRuralUrban(self, index):
        self.keywords = dict([(i, []) for i in range(df.shape[0])])
        print(self.keywords)
        self.menu = QtWidgets.QMenu(Dialog)
        self.menu.setStyleSheet('QMenu { menu-scrollable: true; width: 300 }')
        font = self.menu.font()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.menu.setFont(font)

        index = 18
        self.col = index

        data_unique = []

        self.checkBoxs = []

        # Selectall added into Dropdown
        checkBox = QtWidgets.QCheckBox("Select all/ Deselect all", self.menu)
        checkBox.setStyleSheet("color: red; spacing: 5px; font-size:15px;")
        # All the checkboxes are enabled to check
        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(checkBox)
        self.menu.addAction(checkableAction)
        checkBox.setChecked(True)
        checkBox.stateChanged.connect(self.slotSelectRuralUrban)

        # list storing Facility Name data
        df['col_18'].fillna('Blank', inplace = True)
        list_set = df['col_18'].to_list()

        item = list_set
        item = sorted(list_set, key=str.upper)

        # looping to fill checkboxes, initially all checkboxes will be checked
        for i in range(len(item)):
            if item[i] not in data_unique:
                data_unique.append(item[i])
                checkBox = QtWidgets.QCheckBox(item[i], self.menu)
                checkBox.setChecked(True)
                checkableAction = QtWidgets.QWidgetAction(self.menu)
                checkableAction.setDefaultWidget(checkBox)
                self.menu.addAction(checkableAction)
                self.checkBoxs.append(checkBox)

        # Ok, cancel button
        btn = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel,
                                            QtCore.Qt.Vertical, self.menu)

        # ok selected
        btn.accepted.connect(self.menuCloseRuralUrban)
        # rejected , nothing selected
        btn.rejected.connect(self.menu.close)

        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(btn)
        self.menu.addAction(checkableAction)

        ############# Always set Pushbutton ####################
        self.pushButton_4.setMenu(self.menu)

    # method to check -> uncheck and vice versa
    def slotSelectRuralUrban(self, state):
        for checkbox in self.checkBoxs:
            checkbox.setChecked(QtCore.Qt.Checked == state)

    # after ok selected
    def menuCloseRuralUrban(self):
        self.keywords[self.col] = []
        for element in self.checkBoxs:
            if element.isChecked():
                self.keywords[self.col].append(element.text())
        print(self.keywords[self.col])
        self.filterdataRuralUrban()
        self.pushButton_4.setText('Selected')
        self.pushButton_4.setEnabled(False)
        self.menu.close()

    # Filter data columnwise
    def filterdataRuralUrban(self):
        global df
        #keywords = dict([(i, []) for i in range(self.filterall.columnCount())])
        columnsShow = dict([(i, True) for i in range(df['col_18'].shape[0])])
        print(columnsShow)

        j = 0
        for j in range(df['col_18'].shape[0]):
            item = df['col_18'].to_list()

            # if self.keywords[self.col]:
            if item[j] not in self.keywords[self.col]:
                columnsShow[j] = False

        # for key, value in columnsShow.items():
        final_lst = [i for i in columnsShow.values()]
        print(final_lst, 'this is final list of Select Month')

        # matching list of facility type with col of dataframe returned by onSelectDistrict fun
        df = df[final_lst]
        return df


    ################################################################################
    # Select Ownership

    # Select Ownership Filter
    def onSelectOwnership(self, index):
        self.keywords = dict([(i, []) for i in range(df.shape[0])])
        print(self.keywords)
        self.menu = QtWidgets.QMenu(Dialog)
        self.menu.setStyleSheet('QMenu { menu-scrollable: true; width: 300 }')
        font = self.menu.font()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.menu.setFont(font)

        index = 19
        self.col = index

        data_unique = []

        self.checkBoxs = []

        # Selectall added into Dropdown
        checkBox = QtWidgets.QCheckBox("Select all/ Deselect all", self.menu)
        checkBox.setStyleSheet("color: red; spacing: 5px; font-size:15px;")
        # All the checkboxes are enabled to check
        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(checkBox)
        self.menu.addAction(checkableAction)
        checkBox.setChecked(True)
        checkBox.stateChanged.connect(self.slotSelectOwnership)

        # list storing Facility Name data
        df['col_19'].fillna('blank', inplace = True)
        list_set = df['col_19'].to_list()

        item = list_set
        item = sorted(list_set, key=str.upper)

        # looping to fill checkboxes, initially all checkboxes will be checked
        for i in range(len(item)):
            if item[i] not in data_unique:
                data_unique.append(item[i])
                checkBox = QtWidgets.QCheckBox(item[i], self.menu)
                checkBox.setChecked(True)
                checkableAction = QtWidgets.QWidgetAction(self.menu)
                checkableAction.setDefaultWidget(checkBox)
                self.menu.addAction(checkableAction)
                self.checkBoxs.append(checkBox)

        # Ok, cancel button
        btn = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel,
                                            QtCore.Qt.Vertical, self.menu)

        # ok selected
        btn.accepted.connect(self.menuCloseOwnership)
        # rejected , nothing selected
        btn.rejected.connect(self.menu.close)

        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(btn)
        self.menu.addAction(checkableAction)

        ############# Always set Pushbutton ####################
        self.pushButton_5.setMenu(self.menu)

    # method to check -> uncheck and vice versa
    def slotSelectOwnership(self, state):
        for checkbox in self.checkBoxs:
            checkbox.setChecked(QtCore.Qt.Checked == state)

    # after ok selected
    def menuCloseOwnership(self):
        self.keywords[self.col] = []
        for element in self.checkBoxs:
            if element.isChecked():
                self.keywords[self.col].append(element.text())
        print(self.keywords[self.col])
        self.filterdataOwnership()
        self.pushButton_5.setText('Selected')
        self.pushButton_5.setEnabled(False)
        self.menu.close()

    # Filter data columnwise
    def filterdataOwnership(self):
        global df
        #keywords = dict([(i, []) for i in range(self.filterall.columnCount())])
        columnsShow = dict([(i, True) for i in range(df['col_19'].shape[0])])
        print(columnsShow)

        j = 0
        for j in range(df['col_19'].shape[0]):
            item = df['col_19'].to_list()

            # if self.keywords[self.col]:
            if item[j] not in self.keywords[self.col]:
                columnsShow[j] = False

        # for key, value in columnsShow.items():
        final_lst = [i for i in columnsShow.values()]

        # matching list of facility type with col of dataframe returned by onSelectDistrict fun
        df = df[final_lst]
        return df


    # To count summary of the Modified Checks
    # =======================================
    def summaryReport(self):
        global final_result_summ1, final_result_summ2, col_sum, dft_ARFacilityWise, dft_ARCheckWiseInc, FList1, dft_ARCheckWisePRE, FList2, dft_FacilityWisePRE, dft_FacilityWiseInc, FList3, FList4
        FType = self.lineEdit_2.text()

        # For Health Sub Centre
        if FType == 'Health Sub Centre':
            df = load_HSC(self)
            df_SummReport = df.iloc[:, 200:]     ## Taking columns after 200th
            val_Description = [
                                'Number of mothers provided full course of 180 IFA tablets after delivery <= Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) + Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+ Number of Institutional Deliveries conducted',
                                'Out of the total ANC registered, number registered within 1st trimester (within 12 weeks)<=Total number of pregnant women registered for ANC', 
                                'Out of the new cases of PW with hypertension detected, cases managed at institution <=New cases of PW with hypertension detected',
                                'Number of PW received 4 or more ANC check ups<=Total number of pregnant women registered for ANC',
                                'Number of PW tested using POC test for Syphilis<=Total number of pregnant women registered for ANC',
                                    'Out of the above, Number of PW found sero positive for Syphilis<=Number of PW tested using POC test for Syphilis',
                                    'Number of PW given Tablet Misoprostol during home delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)',
                                    'Number of newborns received 7 Home Based Newborn Care (HBNC) visits in case of Home delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)',
                                    'Number of newborns received 6 HBNC  visits after Institutional Delivery<=Number of Institutional Deliveries conducted',
                                        'Number of mothers provided 360 Calcium tablets after delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted',
                                        'Child immunisation - Vitamin K1 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                        'Child immunisation - OPV 0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                        'Child immunisation - Hepatitis-B0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                        'Out of total institutional deliveries number of women discharged within 48 hours of delivery<=Number of Institutional Deliveries conducted',
                                        'Number of Pre term newborns ( < 37 weeks of pregnancy)<=Live Birth - Male+Live Birth - Female',
                                        'Number of newborns weighed at birth<=Live Birth - Male+Live Birth - Female',
                                            'Number of newborns having weight less than 2.5 kg<=Number of newborns weighed at birth',
                                            'Number of Newborns breast fed within 1 hour of birth<=Live Birth - Male+Live Birth - Female',
                                            'Women receiving 1st post partum checkup within 48 hours of home delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)',
                                            'Number of Post Partum (within 48 hours of delivery) IUCD insertions<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Children aged between 9 and 11 months fully immunized- Male+Children aged between 9 and 11 months fully immunized - Female<=Child immunisation (9-11months) - Measles & Rubella (MR) 1st dose  & Child immunisation (9-11months) - Measles 1st dose',
                                            'Number of cases of AEFI - Abscess<=Number of children immunized (6.1.1+6.1.2+6.1.3+6.1.4+6.1.5+6.1.6+6.1.7+6.1.8+6.1.13+6.1.14+6.1.15+6.1.16+6.1.17+6.1.18+6.1.19+6.1.20+6.1.21+6.2.1+6.2.2+6.2.3+6.3.1+6.3.2+6.3.3+6.4.1+6.4.2+6.4.3+6.4.5+6.4.6+6.5.1+6.5.2+6.5.3+6.5.4)',
                                                'Number of cases of AEFI - Death<=Number of children immunized (6.1.1+6.1.2+6.1.3+6.1.4+6.1.5+6.1.6+6.1.7+6.1.8+6.1.13+6.1.14+6.1.15+6.1.16+6.1.17+6.1.18+6.1.19+6.1.20+6.1.21+6.2.1+6.2.2+6.2.3+6.3.1+6.3.2+6.3.3+6.4.1+6.4.2+6.4.3+6.4.5+6.4.6+6.5.1+6.5.2+6.5.3+6.5.4)',
                                                'Number of cases of AEFI - Others<=Number of children immunized (6.1.1+6.1.2+6.1.3+6.1.4+6.1.5+6.1.6+6.1.7+6.1.8+6.1.13+6.1.14+6.1.15+6.1.16+6.1.17+6.1.18+6.1.19+6.1.20+6.1.21+6.2.1+6.2.2+6.2.3+6.3.1+6.3.2+6.3.3+6.4.1+6.4.2+6.4.3+6.4.5+6.4.6+6.5.1+6.5.2+6.5.3+6.5.4)',
                                                'Number of Immunisation sessions where ASHAs were present<=Immunisation sessions held',
                                                'Out of the total number of Hb tests done, Number having Hb < 7 mg <=Number of Hb tests conducted',
                                                'out of the above, Number screened positive<=Number of Pregnant Women screened for HIV',
                                                    'Live Birth - Male+Live Birth - Female+Still Birth>=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted',
                                                    'Malaria (RDT) - Plamodium Falciparum test positive<=RDT conducted for Malaria',
                                                    'Allopathic- Outpatient attendance+Ayush - Outpatient attendance >= 9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8',
                                                    'Malaria (RDT) - Plasmodium Vivax test positive<=RDT conducted for Malaria' ]

        # For Primary Health Centre
        elif FType == 'Primary Health Centre':
            df = load_PHC(self)
            df_SummReport = df.iloc[:, 305:]     ## Taking columns after 305th
            val_Description = [
                                'Child immunisation - Vitamin K (Birth Dose) <= Live Birth - Male+Live Birth - Female',
                                'Out of the total ANC registered, number registered within 1st trimester (within 12 weeks)<=Total number of pregnant women registered for ANC',
                                'Number of PW given 180 Iron Folic Acid (IFA) tablets <=Total number of pregnant women registered for ANC', 
                                'Number of PW given 360 Calcium tablets <=Total number of pregnant women registered for ANC', 
                                'Number of PW received 4 or more ANC check ups<=Total number of pregnant women registered for ANC', 
                                'Number of PW given Tablet Misoprostol during home delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)',
                                'Number of newborns received 7 Home Based Newborn Care (HBNC) visits in case of Home delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)',
                                'Out of total institutional deliveries number of women discharged within 48 hours of delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                'Number of Eclampsia cases managed during delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'No. of PW having severe anaemia (Hb<7) treated could be greater than No. of PW having severe anaemia (Hb<7)  tested cases',
                                    'Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)<=Total number of pregnant women registered for ANC',
                                    'Number of PW tested positive for GDM<=Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)',
                                    'Number of PW given insulin out of total tested positive for GDM<=Number of PW tested positive for GDM',
                                    'Number of Pregnant women tested for Syphilis<=Total number of pregnant women registered for ANC', 
                                    'Number of Pregnant women tested found sero positive for Syphilis<=Number of Pregnant women tested for Syphilis',
                                    'Number of Syphilis positive pregnant women treated for Syphilis<=Number of Pregnant women tested found sero positive for Syphilis',
                                    'Number of babies treated for congenital Syphilis<=Number of babies diagnosed with congenital Syphilis',
                                    'C-sections, performed at night (8 PM- 8 AM)<=Total C -Section deliveries performed',
                                    'Total C -Section deliveries performed<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Number of Pre term newborns ( < 37 weeks of pregnancy)<=Live Birth - Male+Live Birth - Female',
                                        'Live Birth - Male+Live Birth - Female+Still Birth>=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Post Abortion/ MTP Complications Identified<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                        'Post Abortion/ MTP Complications Treated<=Post Abortion/ MTP Complications Identified',
                                        'Number of women provided with post abortion/ MTP contraception <=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                        'Number of newborns weighed at birth<=Live Birth - Male+Live Birth - Female',
                                        'Number of newborns having weight less than 2.5 kg<=Number of newborns weighed at birth',
                                        'Number of Newborns breast fed within 1 hour of birth <=Live Birth - Male+Live Birth - Female',
                                        'Women receiving 1st post partum checkup within 48 hours of home delivery<=Number of Home Deliveries attended by Skill Birth Attendant (SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)',
                                        'Number of mothers provided full course of 180 IFA tablets after delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of mothers provided 360 Calcium tablets after delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'RTI/STI for which treatment initiated - Male<=New RTI/STI cases identified - Male',
                                            'RTI/STI for which treatment initiated -Female<=New RTI/STI cases identified - Female',
                                            'Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of Post Partum (within 48 hours of delivery) IUCD insertions<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted (Including C-Sections)',
                                                'Number of complications following IUCD Insertion<=Number of Interval IUCD Insertions (excluding PPIUCD and PAIUCD)+ Number of post partum (with in 48 hours of delivery) IUCD insertion +Number of post abortion (with 12 days of spontaneous or surgical abortions)  IUCD incertion',
                                                'Complications following male sterilization<=Number of Non Scalpel Vasectomy (NSV) / Conventional Vasectomy conducted',
                                                'Complications following female sterilization<=Number of Laparoscopic sterilizations (excluding post abortion) conducted + Number of Interval Mini-lap (other than post-partum and post abortion) sterilizations conducted + Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted + Number of Post Abortion sterilizations (within 7 days of spontaneous or surgical abortion) conducted',
                                                'Child immunisation - OPV 0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                                'Child immunisation - Hepatitis-B0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                                'Children aged between 9 and 11 months fully immunized- Male+Children aged between 9 and 11 months fully immunized - Female<=Child immunisation (9-11months) - Measles & Rubella (MR) 1st dose  & Child immunisation (9-11months) - Measles 1st dose',
                                                'Kala Azar Positive Cases<=Kala Azar (RDT) - Tests Conducted',
                                                    'Out of registered, Girls received clinical services<=Girls registered in AFHC',
                                                    'Out of registered, Boys received clinical services<=Boys registered in AFHC',
                                                    'Out of registered, Girls received counselling<=Girls registered in AFHC',
                                                    'Out of registered, Boys received counselling<=Boys registered in AFHC',
                                                        'Allopathic- Outpatient attendance+Ayush - Outpatient attendance >= 14.1.1+14.1.2+14.1.3+14.1.4+14.1.5+14.1.6+14.1.7+14.1.8',
                                                        'Number of Left Against Medical Advice (LAMA) cases<=Male Admissions +Female Admissions',
                                                        'Out of Operation major, Gynecology- Hysterectomy surgeries<=Operation major (General and spinal anaesthesia)',
                                                        'out of the above, Number screened positive<=Number of Pregnant Women screened for HIV',
                                                        'number positive for HIV (Number confirmed positive at ICTCs)<=out of the above, Number screened positive',
                                                        'Widal tests - Number Positive<=Widal tests - Number Tested',
                                                        'Number of cases of AEFI - Abscess<=Number of Children Immunized (9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4)',
                                                            'Number of cases of AEFI - Death<=Number of Children Immunized (9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4)',
                                                            'Number of cases of AEFI - Others<=Number of Children Immunized (9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4)',
                                                            'Out of the new cases of PW with hypertension detected, cases managed at institution <=New cases of PW with hypertension detected',
                                                            'Immunisation sessions held <=Immunisation sessions planned',
                                                            'Number of Immunisation sessions where ASHAs were present<=Immunisation sessions held', 
                                                                'Malaria (Microscopy Tests ) - Plasmodium Falciparum test positive<=Total Blood Smears Examined for Malaria', 
                                                                'Malaria (Microscopy Tests ) - Plasmodium Falciparum test positive<=Total Blood Smears Examined for Malaria',
                                                                'Malaria (RDT) - Plasmodium Vivax test positive<=RDT conducted for Malaria',
                                                                'Malaria (RDT) - Plamodium Falciparum test positive<=RDT conducted for Malaria',
                                                                'Inpatient - Malaria <=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                'Inpatient - Dengue<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                    'Inpatient - Typhoid<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                    'Inpatient - Asthma, Chronic Obstructive Pulmonary Disease (COPD), Respiratory infections<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                    'Inpatient - Tuberculosis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                    'Inpatient - Pyrexia of unknown origin (PUO)<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                    'Inpatient - Diarrhea with dehydration<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                        'Inpatient - Hepatitis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                        'Inpatient Deaths - Male<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                        'Inpatient Deaths - Female<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                        'Number of children discharged with target weight gain from the NRCs<=Number of children admitted in NRC',
                                                                        'Out of the total number of Hb tests done, Number having Hb < 7 mg<=Number of Hb tests conducted',
                                                                            'Male HIV - Number Positive<=Male HIV - Number Tested',
                                                                            'Female Non ANC HIV - Number Positive<=Female Non ANC HIV - Number Tested',
                                                                            'Number of Male STI/RTI attendees found sero Positive for syphilis<=Number of Male STI/RTI attendees tested for syphilis',
                                                                            'Number of Female (Non ANC) STI/RTI attendees found sero Positive for syphilis<=Number of Female (Non ANC)STI/RTI attendees tested for syphilis',
                                                                            'Child immunisation - BCG<=Live Birth - Male+Live Birth - Female']


        # For Sub District Hospital
        elif FType == 'Community Health Centre':
            df = load_CHC(self)
            df_SummReport = df.iloc[:, 318:]     ## Taking columns after 305th
            val_Description = [
                                'Out of the ANC registered, number registered with in 1st trimester(Within 12 weeks)<=Total number of pregnant women registered for ANC',
                                'Male HIV Number Positive <= Male HIV - Number Tested',
                                'Number of PW given 180 Iron Folic Acid (IFA) tablets <=Total number of pregnant women registered for ANC ',
                                'Number of PW given 360 Calcium tablets <=Total number of pregnant women registered for ANC ',
                                'Number of PW received 4 or more ANC check ups<=Total number of pregnant women registered for ANC ',
                                'Out of the new cases of PW with hypertension detected, cases managed at institution<=New cases of PW with hypertension detected ',
                                'Number of Eclampsia cases managed during delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                'No. of PW having severe anaemia (Hb<7) treated could be greater than No. of PW having severe anaemia (Hb<7)  tested cases',
                                'Number of PW tested positive for GDM<=Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)',
                                'Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)<=Total number of pregnant women registered for ANC ',
                                    'Number of PW given insulin out of total tested positive for GDM<=Number of PW tested positive for GDM',
                                    'Number of Pregnant women tested found sero positive for Syphilis<=Number of Pregnant women tested for Syphilis',
                                    'Number of Pregnant women tested for Syphilis<=Total number of pregnant women registered for ANC ',
                                    'Number of Syphilis positive pregnant women treated for Syphilis<=Number of Pregnant women tested found sero positive for Syphilis',
                                    'Number of babies treated for congenital Syphilis<=Number of babies diagnosed with congenital Syphilis',
                                    'Out of total institutional deliveries number of women discharged within 48 hours of delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'Total C -Section deliveries performed<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'C-sections, performed at night (8 PM- 8 AM)<=Total C -Section deliveries performed',
                                    'Live Birth - Male + Live Birth - Female + Still Birth>=Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'Number of Pre term newborns ( < 37 weeks of pregnancy)<=Live Birth - Male+Live Birth - Female',
                                    'Post Abortion/ MTP Complications Identified<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                    'Post Abortion/ MTP Complications Treated<=Post Abortion/ MTP Complications Identified',
                                        'Number of women provided with post abortion/ MTP contraception<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                        'Number of newborns weighed at birth<=Live Birth - Male+Live Birth - Female',
                                        'Number of newborns having weight less than 2.5 kg<=Number of newborns weighed at birth',
                                        'Number of Newborns breast fed within 1 hour of birth<=Live Birth - Male+Live Birth - Female',
                                        'Number of Complicated pregnancies treated with Blood Transfusion<=Number of cases of pregnant women with Obstetric Complications attended (Antepartum haemorrhage (APH), Post-Partum Hemorrhage (PPH), Sepsis, Eclampsia and others) ',
                                        'Number of mothers provided full course of 180 IFA tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Number of mothers provided 360 Calcium tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'RTI/STI for which treatment initiated - Male<=New RTI/STI cases identified - Male',
                                        'RTI/STI for which treatment initiated -Female<=New RTI/STI cases identified - Female',
                                        'Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Number of Post Partum (within 48 hours of delivery) IUCD insertions<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of complications following IUCD Insertion<=Number of Interval IUCD Insertions (excluding PPIUCD and PAIUCD)+ Number of post partum (with in 48 hours of delivery) IUCD insertion +Number of post abortion (with 12 days of spontaneous or surgical abortions)  IUCD incertion',
                                            'Complications following male sterilization<=Number of Non Scalpel Vasectomy (NSV) / Conventional Vasectomy conducted',
                                            'Complications following female sterilization<=Number of Laparoscopic sterilizations (excluding post abortion) conducted + Number of Interval Mini-lap (other than post-partum and post abortion) sterilizations conducted + Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted + Number of Post Abortion sterilizations (within 7 days of spontaneous or surgical abortion) conducted',
                                            'Child immunisation - Vitamin K1(Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                            'Child immunisation - BCG<=Live Birth - Male+Live Birth - Female',
                                            'Child immunisation - OPV-0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                            'Child immunisation - Hepatitis-B0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                            'Children aged between 9 and 11 months fully immunized- Male+Children aged between 9 and 11 months fully immunize<=Child immunisation (9-11months) - Measles & Rubella (MR) 1st dose  & Child immunisation (9-11months) - Measles 1st dose',
                                            'Kala Azar Positive Cases<=Kala Azar (RDT) - Tests Conducted',
                                            'Out of registered, Girls received clinical services<=Girls registered in AFHC',
                                            'Out of registered, Boys received clinical services<=Boys registered in AFHC',
                                            'Out of registered, Girls received counselling<=Girls registered in AFHC',
                                            'Out of registered, Boys received counselling<=Boys registered in AFHC',
                                                'Allopathic- Outpatient attendance+Ayush - Outpatient attendance >= Number of outpatients (Diabetes + Hypertension +  Stroke (Paralysis) + Acute Heart Diseases + Mental illness + Epilepsy + Ophthalmic Related + Dental + Oncology',
                                                'Number of Left Against Medical Advice (LAMA) cases<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Malaria<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Dengue<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Typhoid<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Asthma, Chronic Obstructive Pulmonary Disease (COPD), Respiratory infections<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Tuberculosis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Pyrexia of unknown origin (PUO)<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Diarrhea with dehydration<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Hepatitis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Emergency - Trauma ( accident, injury, poisoning etc)<= Patients registered at Emergency Department',
                                                'Emergency - Burn<= Patients registered at Emergency Department',
                                                'Emergency - Obstetrics complications<= Patients registered at Emergency Department',
                                                    'Emergency - Snake Bite<=Patients registered at Emergency Department',
                                                    'Emergency - Acute Caridiac Emergencies<= Patients registered at Emergency Department',
                                                    'Number of deaths occurring at Emergency Department<= Patients registered at Emergency Department',
                                                    'Number of children discharged with target weight gain from the NRCs<=Number of children admitted in NRC',
                                                    'Out of the total number of Hb tests done, Number having Hb < 7 mg<=Number of Hb tests conducted',
                                                    'Female Non ANC HIV - Number Positive<=Female Non ANC HIV - Number Tested',
                                                    'out of the above, Number screened positive<=Number of Pregnant Women screened for HIV',
                                                    'out of the above, Number screened positive, number confirmed with HIV infection at Integrated Counselling and Testing Centre (ICTC) <=out of the above, Number screened positive',
                                                    'Widal tests - Number Positive<=Widal tests - Number Tested',
                                                    'Number of cases of AEFI - Abscess<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                    'Number of cases of AEFI - Death<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                    'Number of cases of AEFI - Others<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                    'Immunisation sessions held <=Immunisation sessions planned ',
                                                        'Number of Immunisation sessions where ASHAs were present<=Immunisation sessions held ',
                                                        'Malaria (Microscopy Tests ) - Plasmodium Vivax test positive<=Total Blood Smears Examined for Malaria ',
                                                        'Malaria (Microscopy Tests ) - Plasmodium Falciparum test positive<=Total Blood Smears Examined for Malaria ',
                                                        'Malaria (RDT) - Plasmodium Vivax test positive<=RDT conducted for Malaria',
                                                        'Malaria (RDT) - Plamodium Falciparum test positive<=RDT conducted for Malaria',
                                                        'Inpatient Deaths - Male <=Inpatient (Male)- Children<18yrs+Inpatient (Male)',
                                                        'Inpatient Deaths - Female<=Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                        'Number of deaths occurring at SNCU<=Special Newborn Care Unit (SNCU Admissions) - Inborn Male + Special Newborn Care Unit (SNCU Admissions) - Inborn Female + Outborn – Male + Outborn - Female + Number of newborns admitted in SNCU - referred by ASHA',
                                                        'Out of Operation major, Gynecology- Hysterectomy surgeries<=Operation major (General and spinal anaesthesia)',
                                                        'Number of Male STI/RTI attendees found sero Positive for syphilis<=Number of Male STI/RTI attendees tested for syphilis',
                                                        'Number of Female (Non ANC) STI/RTI attendees found sero Positive for syphilis<=Number of Female (Non ANC)STI/RTI attendees tested for syphilis'
                                                        ]

        # For Primary Health Centre
        elif FType == 'Sub District Hospital':
            df = load_SDH(self)
            df_SummReport = df.iloc[:, 321:]     ## Taking columns after 321th
            val_Description = [
                                'Out of the ANC registered, number registered with in 1st trimester(Within 12 weeks)<=Total number of pregnant women registered for ANC',
                                'Male HIV Number Positive <= Male HIV - Number Tested',
                                'Number of PW given 180 Iron Folic Acid (IFA) tablets <=Total number of pregnant women registered for ANC ',
                                'Number of PW given 360 Calcium tablets <=Total number of pregnant women registered for ANC ',
                                'Number of PW received 4 or more ANC check ups<=Total number of pregnant women registered for ANC ',
                                'Out of the new cases of PW with hypertension detected, cases managed at institution<=New cases of PW with hypertension detected ',
                                'Number of Eclampsia cases managed during delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'No. of PW having severe anaemia (Hb<7) treated could be greater than No. of PW having severe anaemia (Hb<7)  tested cases',
                                    'Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)<=Total number of pregnant women registered for ANC ',
                                    'Number of PW tested positive for GDM<=Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)',
                                    'Number of PW given insulin out of total tested positive for GDM<=Number of PW tested positive for GDM',
                                    'Number of Pregnant women tested for Syphilis<=Total number of pregnant women registered for ANC ',
                                    'Number of Pregnant women tested found sero positive for Syphilis<=Number of Pregnant women tested for Syphilis',
                                    'Number of Syphilis positive pregnant women treated for Syphilis<=Number of Pregnant women tested found sero positive for Syphilis',
                                    'Number of babies treated for congenital Syphilis<=Number of babies diagnosed with congenital Syphilis',
                                        'Out of total institutional deliveries number of women discharged within 48 hours of delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Total C -Section deliveries performed<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'C-sections, performed at night (8 PM- 8 AM)<=Total C -Section deliveries performed',
                                        'Live Birth - Male + Live Birth - Female + Still Birth>=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Number of Pre term newborns ( < 37 weeks of pregnancy)<=Live Birth - Male+Live Birth - Female',
                                        'Post Abortion/ MTP Complications Identified<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                        'Post Abortion/ MTP Complications Treated<=Post Abortion/ MTP Complications Identified',
                                        'Number of women provided with post abortion/ MTP contraception<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                        'Number of newborns weighed at birth<=Live Birth - Male+Live Birth - Female',
                                        'Number of newborns having weight less than 2.5 kg<=Number of newborns weighed at birth',
                                        'Number of Newborns breast fed within 1 hour of birth<=Live Birth - Male+Live Birth - Female',
                                            'Number of Complicated pregnancies treated with Blood Transfusion<=Number of cases of pregnant women with Obstetric Complications attended (Antepartum haemorrhage (APH), Post-Partum Hemorrhage (PPH), Sepsis, Eclampsia and others) ',
                                            'Number of mothers provided full course of 180 IFA tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of mothers provided 360 Calcium tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'RTI/STI for which treatment initiated - Male<=New RTI/STI cases identified - Male',
                                            'RTI/STI for which treatment initiated -Female<=New RTI/STI cases identified - Female',
                                            'Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of Post Partum (within 48 hours of delivery) IUCD insertions<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of complications following IUCD Insertion<=Number of Interval IUCD Insertions (excluding PPIUCD and PAIUCD)+ Number of post partum (with in 48 hours of delivery) IUCD insertion +Number of post abortion (with 12 days of spontaneous or surgical abortions)  IUCD incertion',
                                            'Complications following male sterilization<=Number of Non Scalpel Vasectomy (NSV) / Conventional Vasectomy conducted',
                                            'Complications following female sterilization<=Number of Laparoscopic sterilizations (excluding post abortion) conducted + Number of Interval Mini-lap (other than post-partum and post abortion) sterilizations conducted + Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted + Number of Post Abortion sterilizations (within 7 days of spontaneous or surgical abortion) conducted',
                                            'Child immunisation - Vitamin K1(Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                                'Child immunisation - BCG<=Live Birth - Male+Live Birth - Female',
                                                'Child immunisation - OPV-0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                                'Child immunisation - Hepatitis-B0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                                'Children aged between 9 and 11 months fully immunized- Male+Children aged between 9 and 11 months fully immunize<=Child immunisation (9-11months) - Measles & Rubella (MR) 1st dose  & Child immunisation (9-11months) - Measles 1st dose',
                                                'Kala Azar Positive Cases<=Kala Azar (RDT) - Tests Conducted',
                                                'Tests Positive for JE<=Tests Conducted for JE',
                                                'Out of registered, Girls received clinical services<=Girls registered in AFHC',
                                                'Out of registered, Boys received clinical services<=Boys registered in AFHC',
                                                'Out of registered, Girls received counselling<=Girls registered in AFHC',
                                                'Out of registered, Boys received counselling<=Boys registered in AFHC',
                                                'Allopathic- Outpatient attendance+Ayush - Outpatient attendance >= Number of outpatients (Diabetes + Hypertension +  Stroke (Paralysis) + Acute Heart Diseases + Mental illness + Epilepsy + Ophthalmic Related + Dental + Oncology)',
                                                    'Number of Left Against Medical Advice (LAMA) cases<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Malaria<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Dengue<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Typhoid<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Asthma, Chronic Obstructive Pulmonary Disease (COPD), Respiratory infections<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Tuberculosis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Pyrexia of unknown origin (PUO)<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Diarrhea with dehydration<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Hepatitis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Emergency - Trauma ( accident, injury, poisoning etc)<= Patients registered at Emergency Department',
                                                    'Emergency - Burn<= Patients registered at Emergency Department',
                                                    'Emergency - Obstetrics complications<= Patients registered at Emergency Department',
                                                        'Emergency - Snake Bite<=Patients registered at Emergency Department',
                                                        'Emergency - Acute Caridiac Emergencies<= Patients registered at Emergency Department',
                                                        'Emergency - CVA ( Cerebovascular Disease)<= Patients registered at Emergency Department',
                                                        'Number of deaths occurring at Emergency Department<= Patients registered at Emergency Department',
                                                        'Number of children discharged with target weight gain from the NRCs<=Number of children admitted in NRC',
                                                        'Out of the total number of Hb tests done, Number having Hb < 7 mg<=Number of Hb tests conducted',
                                                        'Female Non ANC HIV - Number Positive<=Female Non ANC HIV - Number Tested',
                                                        'out of the above, Number screened positive<=Number of Pregnant Women screened for HIV',
                                                        'out of the above, Number screened positive, number confirmed with HIV infection at Integrated Counselling and Testing Centre (ICTC) <=out of the above, Number screened positive',
                                                        'Widal tests - Number Positive<=Widal tests - Number Tested',
                                                        'Number of cases of AEFI - Abscess<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                        'Number of cases of AEFI - Death<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                        'Number of cases of AEFI - Others<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                        'Immunisation sessions held <=Immunisation sessions planned ',
                                                            'Number of Immunisation sessions where ASHAs were present<=Immunisation sessions held ',
                                                            'Malaria (Microscopy Tests ) - Plasmodium Vivax test positive<=Total Blood Smears Examined for Malaria ',
                                                            'Malaria (Microscopy Tests ) - Plasmodium Falciparum test positive<=Total Blood Smears Examined for Malaria ',
                                                            'Malaria (RDT) - Plasmodium Vivax test positive<=RDT conducted for Malaria',
                                                            'Malaria (RDT) - Plamodium Falciparum test positive<=RDT conducted for Malaria',
                                                            'Inpatient Deaths - Male <=Inpatient (Male)- Children<18yrs+Inpatient (Male)',
                                                            'Inpatient Deaths - Female<=Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                            'Number of deaths occurring at SNCU<=Special Newborn Care Unit (SNCU Admissions) - Inborn Male + Special Newborn Care Unit (SNCU Admissions) - Inborn Female + Outborn – Male + Outborn - Female + Number of newborns admitted in SNCU - referred by ASHA',
                                                            'Out of Operation major, Gynecology- Hysterectomy surgeries<=Operation major (General and spinal anaesthesia)',
                                                            'Number of Male STI/RTI attendees found sero Positive for syphilis<=Number of Male STI/RTI attendees tested for syphilis',
                                                            'Number of Female (Non ANC) STI/RTI attendees found sero Positive for syphilis<=Number of Female (Non ANC)STI/RTI attendees tested for syphilis']
            
        # For District Hospital
        elif FType == 'District Hospital':
            df = load_DH(self)
            df_SummReport = df.iloc[:, 326:]     ## Taking columns after 326th
            val_Description = [
                                'Out of the ANC registered, number registered with in 1st trimester(Within 12 weeks)<=Total number of pregnant women registered for ANC',
                                'Male HIV Number Positive <= Male HIV - Number Tested',
                                'Number of PW given 180 Iron Folic Acid (IFA) tablets <=Total number of pregnant women registered for ANC ',
                                'Number of PW given 360 Calcium tablets <=Total number of pregnant women registered for ANC ',
                                'Number of PW received 4 or more ANC check ups<=Total number of pregnant women registered for ANC ',
                                'Out of the new cases of PW with hypertension detected, cases managed at institution<=New cases of PW with hypertension detected ',
                                'Number of Eclampsia cases managed during delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                'No. of PW having severe anaemia (Hb<7) treated could be greater than No. of PW having severe anaemia (Hb<7)  tested cases',
                                'Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)<=Total number of pregnant women registered for ANC ',
                                    'Number of PW tested positive for GDM<=Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)',
                                    'Number of PW given insulin out of total tested positive for GDM<=Number of PW tested positive for GDM',
                                    'Number of Pregnant women tested for Syphilis<=Total number of pregnant women registered for ANC ',
                                    'Number of Pregnant women tested found sero positive for Syphilis<=Number of Pregnant women tested for Syphilis',
                                    'Number of Syphilis positive pregnant women treated for Syphilis<=Number of Pregnant women tested found sero positive for Syphilis',
                                    'Number of babies treated for congenital Syphilis<=Number of babies diagnosed with congenital Syphilis',
                                    'Out of total institutional deliveries number of women discharged within 48 hours of delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'Total C -Section deliveries performed<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'C-sections, performed at night (8 PM- 8 AM)<=Total C -Section deliveries performed',
                                    'Live Birth - Male + Live Birth - Female + Still Birth>=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Number of Pre term newborns ( < 37 weeks of pregnancy)<=Live Birth - Male+Live Birth - Female',
                                        'Post Abortion/ MTP Complications Identified<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                        'Post Abortion/ MTP Complications Treated<=Post Abortion/ MTP Complications Identified',
                                        'Number of women provided with post abortion/ MTP contraception<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                        'Number of newborns weighed at birth<=Live Birth - Male+Live Birth - Female',
                                        'Number of newborns having weight less than 2.5 kg<=Number of newborns weighed at birth',
                                        'Number of Newborns breast fed within 1 hour of birth<=Live Birth - Male+Live Birth - Female',
                                        'Number of Complicated pregnancies treated with Blood Transfusion<=Number of cases of pregnant women with Obstetric Complications attended (Antepartum haemorrhage (APH), Post-Partum Hemorrhage (PPH), Sepsis, Eclampsia and others) ',
                                        'Number of mothers provided full course of 180 IFA tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Number of mothers provided 360 Calcium tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'RTI/STI for which treatment initiated - Male<=New RTI/STI cases identified - Male',
                                            'RTI/STI for which treatment initiated -Female<=New RTI/STI cases identified - Female',
                                            'Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of Post Partum (within 48 hours of delivery) IUCD insertions<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of complications following IUCD Insertion<=Number of Interval IUCD Insertions (excluding PPIUCD and PAIUCD)+ Number of post partum (with in 48 hours of delivery) IUCD insertion +Number of post abortion (with 12 days of spontaneous or surgical abortions)  IUCD incertion',
                                            'Complications following male sterilization<=Number of Non Scalpel Vasectomy (NSV) / Conventional Vasectomy conducted',
                                            'Complications following female sterilization<=Number of Laparoscopic sterilizations (excluding post abortion) conducted + Number of Interval Mini-lap (other than post-partum and post abortion) sterilizations conducted + Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted + Number of Post Abortion sterilizations (within 7 days of spontaneous or surgical abortion) conducted',
                                            'Child immunisation - Vitamin K1(Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                            'Child immunisation - BCG<=Live Birth - Male+Live Birth - Female',
                                            'Child immunisation - OPV-0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                            'Child immunisation - Hepatitis-B0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                            'Children aged between 9 and 11 months fully immunized- Male+Children aged between 9 and 11 months fully immunize<=Child immunisation (9-11months) - Measles & Rubella (MR) 1st dose  & Child immunisation (9-11months) - Measles 1st dose',
                                                'Kala Azar Positive Cases<=Kala Azar (RDT) - Tests Conducted',
                                                'Tests Positive for JE<=Tests Conducted for JE',
                                                'Out of registered, Girls received clinical services<=Girls registered in AFHC',
                                                'Out of registered, Boys received clinical services<=Boys registered in AFHC',
                                                'Out of registered, Girls received counselling<=Girls registered in AFHC',
                                                'Out of registered, Boys received counselling<=Boys registered in AFHC',
                                                'Allopathic- Outpatient attendance+Ayush - Outpatient attendance >= Number of outpatients (Diabetes + Hypertension +  Stroke (Paralysis) + Acute Heart Diseases + Mental illness + Epilepsy + Ophthalmic Related + Dental + Oncology',
                                                'Number of Left Against Medical Advice (LAMA) cases<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Malaria<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Dengue<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Typhoid<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Asthma, Chronic Obstructive Pulmonary Disease (COPD), Respiratory infections<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Tuberculosis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Pyrexia of unknown origin (PUO)<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Diarrhea with dehydration<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Hepatitis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Emergency - Trauma ( accident, injury, poisoning etc)<= Patients registered at Emergency Department',
                                                    'Emergency - Burn<= Patients registered at Emergency Department',
                                                    'Emergency - Obstetrics complications<= Patients registered at Emergency Department',
                                                    'Emergency - Snake Bite<=Patients registered at Emergency Department',
                                                    'Emergency - Acute Caridiac Emergencies<= Patients registered at Emergency Department',
                                                    'Emergency - CVA ( Cerebovascular Disease)<= Patients registered at Emergency Department',
                                                        'Number of deaths occurring at Emergency Department<= Patients registered at Emergency Department',
                                                        'Number of children discharged with target weight gain from the NRCs<=Number of children admitted in NRC',
                                                        'Out of the total number of Hb tests done, Number having Hb < 7 mg<=Number of Hb tests conducted',
                                                        'Female Non ANC HIV - Number Positive<=Female Non ANC HIV - Number Tested',
                                                        'out of the above, Number screened positive<=Number of Pregnant Women screened for HIV',
                                                        'out of the above, Number screened positive, number confirmed with HIV infection at Integrated Counselling and Testing Centre (ICTC) <=out of the above, Number screened positive',
                                                        'Widal tests - Number Positive<=Widal tests - Number Tested',
                                                        'Number of cases of AEFI - Abscess<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                        'Number of cases of AEFI - Death<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                        'Number of cases of AEFI - Others<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                        'Immunisation sessions held <=Immunisation sessions planned ',
                                                            'Number of Immunisation sessions where ASHAs were present<=Immunisation sessions held ',
                                                            'Malaria (Microscopy Tests ) - Plasmodium Vivax test positive<=Total Blood Smears Examined for Malaria ',
                                                            'Malaria (Microscopy Tests ) - Plasmodium Falciparum test positive<=Total Blood Smears Examined for Malaria ',
                                                            'Malaria (RDT) - Plasmodium Vivax test positive<=RDT conducted for Malaria',
                                                            'Malaria (RDT) - Plamodium Falciparum test positive<=RDT conducted for Malaria',
                                                            'Total number of blood units issued during the month>=Number of blood units issued (Excluding C-Section)',
                                                            'Inpatient Deaths - Male <=Inpatient (Male)- Children<18yrs+Inpatient (Male)',
                                                            'Inpatient Deaths - Female<=Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                            'Number of deaths occurring at SNCU<=Special Newborn Care Unit (SNCU Admissions) - Inborn Male + Special Newborn Care Unit (SNCU Admissions) - Inborn Female + Outborn – Male + Outborn - Female + Number of newborns admitted in SNCU - referred by ASHA',
                                                            'Out of Operation major, Gynecology- Hysterectomy surgeries<=Operation major (General and spinal anaesthesia)',
                                                            'Major Surgeries excluding Obstetrics, Gynaecology and Opthalmology etc.<=Operation major (General and spinal anaesthesia)',
                                                            'Number of Male STI/RTI attendees found sero Positive for syphilis<=Number of Male STI/RTI attendees tested for syphilis',
                                                            'Number of Female (Non ANC) STI/RTI attendees found sero Positive for syphilis<=Number of Female (Non ANC)STI/RTI attendees tested for syphilis']



        '''
        ## First Summary Report
        ## ---------------------
        '''

        # count_Consistent = []
        count_Inconsistent = []
        # count_Blank = []
        count_ProbableRErr = []
        Columns = list(df_SummReport.columns.values.tolist())

        for col_name in Columns:

            # c1 = df_SummReport[col_name].str.count("consistent").sum()
            # count_Consistent.append(c1)

            c2 = df_SummReport[col_name].str.count("Inconsistent").sum()
            count_Inconsistent.append(c2)

            # c3 = df_SummReport[col_name].str.count("Blank").sum()
            # count_Blank.append(c3)

            c4 = df_SummReport[col_name].str.count("Probable Reporting Error").sum()
            count_ProbableRErr.append(c4)

        print(len(val_Description), len(count_Inconsistent), len(count_ProbableRErr))   

        # To show facilities in a column
        colInterest = df_['col_14']
        
        ''' For Inconsistent '''
        inconsistent_list = []

        lg = len(df_SummReport.columns)

        for i in range(0, lg):
            temp = []
            colComparison = df_SummReport.iloc[:,i]

            for j in range(0, len(colComparison)):
                primString = colComparison[j]
                
                if primString is 'Inconsistent':
                    temp.append(colInterest[j])

            inconsistent_list.append(temp)           

        ''' For PRE '''
        PRE_list = []
        # TO show facility in the sheet5
        lg = len(df_SummReport.columns)

        for i in range(0, lg):
            temp = []
            colComparison = df_SummReport.iloc[:,i]

            for j in range(0, len(colComparison)):
                primString = colComparison[j]

                pattern = re.compile("^P")
                if pattern.match(str(primString)):
                    temp.append(colInterest[j])

            PRE_list.append(temp)      


        final_result_summ1 = pd.DataFrame({"Conditions": df_SummReport.columns, 
                                            "Description": val_Description,
                                                "Facilities(Name) Showing Inconsistent": inconsistent_list,
                                                    "Inconsistent": count_Inconsistent,
                                                        "Facilities (Name) Showing Probable Reporting Error": PRE_list,
                                                            "Probable Reporting Error": count_ProbableRErr,
                                                            })

                                                            
        FList1 = final_result_summ1["Facilities(Name) Showing Inconsistent"].tolist()
        FList2 = final_result_summ1["Facilities (Name) Showing Probable Reporting Error"].tolist()

        # Sorting ascending to descending
        FList1.sort(key=lambda k: len(k), reverse=True)
        FList2.sort(key=lambda k: len(k), reverse=True)


        dataframeForSheet4 = final_result_summ1[['Conditions', 'Description', 'Inconsistent', 'Facilities(Name) Showing Inconsistent']]
        dataframeForSheet5 = final_result_summ1[['Conditions', 'Description', 'Probable Reporting Error', 'Facilities (Name) Showing Probable Reporting Error']]

        # Counting total numnber of facility names and calculating percentage and 
        # Coloring percentage >=25% to RED
        count_rows = df.shape[0]
            
            # Percentage for Validation Summary Sheet to show color codes
        final_result_summ1['PerIncSheet1'] = final_result_summ1['Inconsistent']/count_rows*100
        final_result_summ1['PerPRErrSheet1'] = final_result_summ1['Probable Reporting Error']/count_rows*100

        len0 = len(final_result_summ1['PerIncSheet1'])
        len1 = len(final_result_summ1['PerPRErrSheet1'])

        col_percentageInc_rank1 = []
        col_percentageErr_rank1 = []

        # For All Inconsistents in all five ranges respectively
        # Blank list for Inconsistents
        range11_Inc, range21_Inc, range31_Inc, range41_Inc = [], [], [], []
        # Blank list for Probable Reporting Error
        range11_PRE, range21_PRE, range31_PRE, range41_PRE = [], [], [], []

        for k in range(0,len0):
            primcol_Inc1 = final_result_summ1['PerIncSheet1'][k]
            if primcol_Inc1 < 5 :
                col_percentageInc_rank1.append('1')
                #col_percentageInc_Bucket.append('< 5')
                range11_Inc.append(primcol_Inc1)
            elif primcol_Inc1 >= 5 and primcol_Inc1 < 10:
                col_percentageInc_rank1.append('2')
                #col_percentageInc_Bucket.append('5 & < 10')
                range21_Inc.append(primcol_Inc1)
            elif primcol_Inc1 >= 10 and primcol_Inc1 < 25:
                col_percentageInc_rank1.append('3')
                #col_percentageInc_Bucket.append('10 & < 25')
                range31_Inc.append(primcol_Inc1)
            elif primcol_Inc1 >= 25:
                col_percentageInc_rank1.append('4')
                #col_percentageInc_Bucket.append('>=25')
                range41_Inc.append(primcol_Inc1)
            else:
                col_percentageInc_rank1.append('NaN')
                #col_percentageInc_Bucket.append('NaN')

        for i in range(0,len1):
            primcol_Prc1 = final_result_summ1['PerPRErrSheet1'][i]
            if primcol_Prc1 < 5 :
                col_percentageErr_rank1.append('1')
                #col_percentageErr_Bucket.append('< 5')
                range11_PRE.append(primcol_Prc1)
            elif primcol_Prc1 >= 5 and primcol_Prc1 < 10:
                col_percentageErr_rank1.append('2')
                #col_percentageErr_Bucket.append('5 & < 10')
                range21_PRE.append(primcol_Prc1)
            elif primcol_Prc1 >= 10 and primcol_Prc1 < 25:
                col_percentageErr_rank1.append('3')
                #col_percentageErr_Bucket.append('10 & < 25')
                range31_PRE.append(primcol_Prc1)
            elif primcol_Prc1 >= 25:
                col_percentageErr_rank1.append('4')
                #col_percentageErr_Bucket.append('>= 20')
                range41_PRE.append(primcol_Prc1)
            else:
                col_percentageErr_rank1.append('NaN')
                #col_percentageErr_Bucket.append('NaN')

        # Deleting unnecessary columns
        del final_result_summ1['PerIncSheet1']
        del final_result_summ1['PerPRErrSheet1']
            
        def select_col_SumSheet(X):
            global cnt21, cnt22, cnt23, cnt24, cnt25, cnt26, cnt27, cnt28
                # COLORS
            c = ['background-color:  #EF5350',                  #   >=25% RED
                        'background-color: #FFAF00',            #   10 - 25% LIGHTER RED
                            'background-color: #C0C000',        #   5 - 10% MORE LIGHTER RED
                                'background-color: #00AF5F',    #   < 5% LIGHTEST RED
                                    '']


            mask_21 = (X['Inconsistent'] >= 4)
            cnt21 = mask_21.values.sum()
            mask_22 = (X['Inconsistent'] == 3)
            cnt22= mask_22.values.sum()
            mask_23 = (X['Inconsistent'] == 2)
            cnt23= mask_23.values.sum()
            mask_24 = (X['Inconsistent'] <= 1)
            cnt24 = mask_24.values.sum()

            mask_25 = (X['Probable Reporting Error'] >= 4)
            cnt25 = mask_25.values.sum()
            mask_26 = (X['Probable Reporting Error'] == 3)
            cnt26 = mask_26.values.sum()
            mask_27 = (X['Probable Reporting Error'] == 2)
            cnt27 = mask_27.values.sum()
            mask_28 = (X['Probable Reporting Error'] <= 1)
            cnt28 = mask_28.values.sum()

            #DataFrame with same index and columns names as original filled empty strings
            df1 =  pd.DataFrame(c[4], X.index, columns=X.columns)

            df1.loc[mask_21, 'Inconsistent'] = c[0]
            df1.loc[mask_22, 'Inconsistent'] = c[1]
            df1.loc[mask_23, 'Inconsistent'] = c[2]
            df1.loc[mask_24, 'Inconsistent'] = c[3]
            df1.loc[mask_25, 'Probable Reporting Error'] = c[0]
            df1.loc[mask_26, 'Probable Reporting Error'] = c[1]
            df1.loc[mask_27, 'Probable Reporting Error'] = c[2]
            df1.loc[mask_28, 'Probable Reporting Error'] = c[3]
            return df1

        # Remoiving inconsistent facility names from validation summary sheet
        final_result_summ1.drop(['Facilities(Name) Showing Inconsistent'], axis = 1, inplace=True)
        # Remoiving PRE facility names from validation summary sheet
        final_result_summ1.drop(['Facilities (Name) Showing Probable Reporting Error'], axis = 1, inplace=True)

        final_result_summ1 = final_result_summ1.style.apply(select_col_SumSheet, axis=None)


        '''
        ## Second Summary Report
        ## --------------------
        '''
        summ2_countInconsistent = []
        summ2_countProbableRErr = []
        All_Blank = []

        # Iterating over indices of each row and calculating number of Blanks for each Facility Name 
        for index in range(len(df_SummReport)):
            '''   For no. of Inconsistent   '''
            inconsistent = df_SummReport.iloc[index, :].str.count("Inconsistent").sum()
            summ2_countInconsistent.append(inconsistent)
            
            '''   For no. of Probable Reporting Errors   '''
            probableRErr = df_SummReport.iloc[index, :].str.count('Probable Reporting Error').sum()
            summ2_countProbableRErr.append(probableRErr)

            blank = df_SummReport.iloc[index, :].str.count("Blank").sum()
            if blank == 0:
                All_Blank.append('Yes')
            else:
                All_Blank.append('No')

        

        #########################################################   
        #  Facility Specific Inconsistent (Sheet 6)      
        colInterest = df_['col_14']
        
        ''' For Inconsistent '''
        inc_list = []

        lg = len(df_SummReport.columns)
        len_df = df_SummReport.shape[0]

        for i in range(0, len_df):
            temp = []

            colComparison = df_SummReport.iloc[i,:]
            for j in range(0, lg):
                primString = colComparison[j]
                
                if primString is 'Inconsistent':
                    temp.append(df_SummReport.columns[j])

            inc_list.append(temp)

        ''' For PRE '''
        pre_list = []

        lg = len(df_SummReport.columns)
        len_df = df_SummReport.shape[0]

        for i in range(0, len_df):
            temp = []

            colComparison = df_SummReport.iloc[i,:]
            for j in range(0, lg):
                primString = colComparison[j]

                pattern = re.compile("^P")
                if pattern.match(str(primString)):
                    temp.append(df_SummReport.columns[j])

            pre_list.append(temp)

        

        ###################################################
        
        final_result_summ2 = pd.DataFrame({ "Facility Name": df['col_14'].tolist(),
                                                "District": df['col_5'].tolist(),
                                                    "Rural/Urban": df['col_18'].tolist(),
                                                        "Ownership": df['col_19'].tolist(),
                                                            "Inconsistent": summ2_countInconsistent,
                                                                "Probable Reporting Error": summ2_countProbableRErr,
                                                                    "All Blank": All_Blank,
                                                                        "Checks (Inconsistent)" : inc_list,
                                                                            "Checks (PRE)": pre_list
                                                                                })

        # Sorting in alphabetical  order
        final_result_summ2 = final_result_summ2.sort_values(by=['Facility Name'], ascending=True)
        final_result_summ2 = final_result_summ2.reset_index(drop=True)

        FList3 = final_result_summ2["Checks (Inconsistent)"].tolist()
        FList4 = final_result_summ2["Checks (PRE)"].tolist()

        # Sorting ascending to descending
        FList3.sort(key=lambda k: len(k), reverse=True)
        FList4.sort(key=lambda k: len(k), reverse=True)

        
        '''
        ## Third Summary Report (Top 10 Validation Checks not performing good)
        ## --------------------
        '''
        dft_ARCheckWiseInc = dataframeForSheet4.sort_values(by=['Inconsistent'], ascending=False)
        dft_ARCheckWiseInc = dft_ARCheckWiseInc.reset_index(drop=True)

        '''
        ## Fourth Summary Report (Top 10 Validation Checks not performing good)
        ## --------------------
        '''
        dft_ARCheckWisePRE = dataframeForSheet5.sort_values(by=['Probable Reporting Error'], ascending=False)
        dft_ARCheckWisePRE = dft_ARCheckWisePRE.reset_index(drop=True)

        '''
        ## 5th and 6th Summary Report (Top 10 Validation Checks not performing good)
        ## --------------------
        '''
        dataframeForSheet6 = final_result_summ2[['Facility Name', 'Inconsistent', 'Checks (Inconsistent)']]
        dataframeForSheet7 = final_result_summ2[['Facility Name', 'Probable Reporting Error', 'Checks (PRE)']]
        dft_FacilityWiseInc = dataframeForSheet6.sort_values(by=['Inconsistent'], ascending=False)
        dft_FacilityWiseInc = dft_FacilityWiseInc.reset_index(drop=True)
        dft_FacilityWisePRE = dataframeForSheet7.sort_values(by=['Probable Reporting Error'], ascending=False)
        dft_FacilityWisePRE = dft_FacilityWisePRE.reset_index(drop=True)



        '''  To find percentage Facility Type Wise   '''
                                                    
        # For Health Sub Centre
        if FType == 'Health Sub Centre':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/31 * 100
            final_result_summ2['PercentagePRErr'] = final_result_summ2['Probable Reporting Error']/ 31 * 100
        
        # For Primary Health Centre
        elif FType == 'Primary Health Centre':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/78 * 100
            final_result_summ2['PercentagePRErr'] = final_result_summ2['Probable Reporting Error']/ 78 * 100

        # For Community Health Centre
        elif FType == 'Community Health Centre':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/83 * 100
            final_result_summ2['PercentagePRErr'] = final_result_summ2['Probable Reporting Error']/ 83 * 100

        # For Sub District Hospital
        elif FType == 'Sub District Hospital':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/85 * 100
            final_result_summ2['PercentagePRErr'] = final_result_summ2['Probable Reporting Error']/ 85 * 100

        # For District Hospital
        elif FType == 'District Hospital':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/85 * 100
            final_result_summ2['PercentagePRErr'] = final_result_summ2['Probable Reporting Error']/ 85 * 100
            
        #  # Top 10 percentile of Inconsistent values
        # df_2 = final_result_summ2[final_result_summ2['Inconsistent'].ge(final_result_summ2['Inconsistent'].quantile(0.9))]
        # print(df_2)

        len0 = len(final_result_summ2['PercentageInc'])
        len1 = len(final_result_summ2['PercentagePRErr'])

        col_percentageInc_rank = []
        col_percentageErr_rank = []

        # For All Inconsistents in all five ranges respectively
        # Blank list for Inconsistents
        range1_Inc, range2_Inc, range3_Inc, range4_Inc, range5_Inc = [], [], [], [], []
        # Blank list for Probable Reporting Error
        range1_PRE, range2_PRE, range3_PRE, range4_PRE, range5_PRE = [], [], [], [], []

        for k in range(0,len0):
            primcol_Inc = final_result_summ2['PercentageInc'][k]
            if primcol_Inc < 5 :
                col_percentageInc_rank.append('1')
                #col_percentageInc_Bucket.append('< 5')
                range1_Inc.append(primcol_Inc)
            elif primcol_Inc >= 5 and primcol_Inc < 10:
                col_percentageInc_rank.append('2')
                #col_percentageInc_Bucket.append('5 & < 10')
                range2_Inc.append(primcol_Inc)
            elif primcol_Inc >= 10 and primcol_Inc < 25:
                col_percentageInc_rank.append('3')
                #col_percentageInc_Bucket.append('10 & < 25')
                range3_Inc.append(primcol_Inc)
            elif primcol_Inc >= 25 and primcol_Inc < 50:
                col_percentageInc_rank.append('4')
                #col_percentageInc_Bucket.append('25 & < 5')
                range4_Inc.append(primcol_Inc)
            elif primcol_Inc >= 50:
                col_percentageInc_rank.append('5')
                #col_percentageInc_Bucket.append('>= 5')
                range5_Inc.append(primcol_Inc)
            else:
                col_percentageInc_rank.append('NaN')
                #col_percentageInc_Bucket.append('NaN')

        for i in range(0,len1):
            primcol_Prc = final_result_summ2['PercentagePRErr'][i]
            if primcol_Prc < 5 :
                col_percentageErr_rank.append('1')
                #col_percentageErr_Bucket.append('< 5')
                range1_PRE.append(primcol_Prc)
            elif primcol_Prc >= 5 and primcol_Prc < 10:
                col_percentageErr_rank.append('2')
                #col_percentageErr_Bucket.append('5 & < 10')
                range2_PRE.append(primcol_Prc)
            elif primcol_Prc >= 10 and primcol_Prc < 25:
                col_percentageErr_rank.append('3')
                #col_percentageErr_Bucket.append('10 & < 25')
                range3_PRE.append(primcol_Prc)
            elif primcol_Prc >= 25 and primcol_Prc < 50:
                col_percentageErr_rank.append('4')
                #col_percentageErr_Bucket.append('25 & < 50')
                range4_PRE.append(primcol_Prc)
            elif primcol_Prc >= 50:
                col_percentageErr_rank.append('5')
                #col_percentageErr_Bucket.append('>= 50')
                range5_PRE.append(primcol_Prc)
            else:
                col_percentageErr_rank.append('NaN')


        # Deleting unnecessary columns
        del final_result_summ2['PercentageInc']
        del final_result_summ2['PercentagePRErr']
        del final_result_summ2['Checks (Inconsistent)']
        del final_result_summ2['Checks (PRE)']


        def select_col(X):
            global cnt1, cnt2, cnt3, cnt4, cnt5, cnt6, cnt7, cnt8, cnt9, cnt10
            # COLORS
            c = ['background-color:  #EF5350',                  #>=50% RED
                    'background-color: #FFAF00',                #25-50% ORANGE
                        'background-color: #C0C000',            #10-25% YELLOW
                            'background-color: #00FF00',        #5-10% L GREEN
                                'background-color: #00AF5F',    #<5% GREEN
                                    '']
   
                
            mask_AllBlank = (X['All Blank'] == 'Yes')   

            mask_10 = (X['Inconsistent'] >= 5)
            cnt1 = mask_10.values.sum()
            mask_11 = (X['Inconsistent'] == 4)
            cnt2 = mask_11.values.sum()
            mask_12 = (X['Inconsistent'] == 3)
            cnt3 = mask_12.values.sum()
            mask_13 = (X['Inconsistent'] == 2)
            cnt4 = mask_13.values.sum()
            mask_14 = (X['Inconsistent'] <= 1)
            cnt5 = mask_14.values.sum()

            mask_5 = (X['Probable Reporting Error'] >= 5)
            cnt6 = mask_5.values.sum()
            mask_6 = (X['Probable Reporting Error'] == 4)
            cnt7 = mask_6.values.sum()
            mask_7 = (X['Probable Reporting Error'] == 3)
            cnt8 = mask_7.values.sum()
            mask_8 = (X['Probable Reporting Error'] == 2)
            cnt9 = mask_8.values.sum()
            mask_9 = (X['Probable Reporting Error'] <= 1)
            cnt10 = mask_9.values.sum()

            #DataFrame with same index and columns names as original filled empty strings
            df1 =  pd.DataFrame(c[5], X.index, columns=X.columns)

            #modify values of df1 column by boolean mask
            df1.loc[mask_5, 'Probable Reporting Error'] = c[0]
            df1.loc[mask_6, 'Probable Reporting Error'] = c[1]
            df1.loc[mask_7, 'Probable Reporting Error'] = c[2]
            df1.loc[mask_8, 'Probable Reporting Error'] = c[3]
            df1.loc[mask_9, 'Probable Reporting Error'] = c[4]

            df1.loc[mask_AllBlank, 'All Blank'] = c[0]
            df1.loc[mask_10, 'Inconsistent'] = c[0]
            df1.loc[mask_11, 'Inconsistent'] = c[1]
            df1.loc[mask_12, 'Inconsistent'] = c[2]
            df1.loc[mask_13, 'Inconsistent'] = c[3]
            df1.loc[mask_14, 'Inconsistent'] = c[4]

            return df1

        final_result_summ2 = final_result_summ2.style.apply(select_col, axis=None)

        return final_result_summ1, final_result_summ2, dft_ARCheckWiseInc, dft_ARCheckWisePRE, dft_FacilityWiseInc, dft_FacilityWisePRE


    # EXPORT FILE
    def export(self):
        table_result1, table_result2, table_result3, table_result4, table_result5, table_result6  = self.summaryReport()

        # Rename orignal headers
        df.rename(res_dict , axis=1, inplace=True)

        # Remove column names month and year
        df.drop(['Month', 'Year'], axis = 1, inplace=True)

        new_list = [[""]]
        table_result_content = pd.DataFrame(new_list)

        # Save file dialog
        filename = QFileDialog.getSaveFileName(Dialog, "Save to Excel", "Summary_Sheet",
                                                "Excel Spreadsheet (*.xlsx);;"
                                                "All Files (*)")[0]

        # Taking transpose of data 
        table_result3 = table_result3.T
        table_result4 = table_result4.T
        table_result5 = table_result5.T
        table_result6 = table_result6.T

        # exporting to excel
        with pd.ExcelWriter(filename) as writer: 
            table_result_content.to_excel(writer, sheet_name='Description', engine='openpyxl')
            table_result2.to_excel(writer, sheet_name='Facility Guidance Sheet', engine='openpyxl')
            table_result1.to_excel(writer, sheet_name='Validation Summary Sheet', engine='openpyxl')
            table_result3.to_excel(writer, sheet_name='Attention Reqd (Inconsistent)', engine='openpyxl')
            table_result4.to_excel(writer, sheet_name='Attention Reqd (PRE)', engine='openpyxl')
            table_result5.to_excel(writer, sheet_name='Checks (Inconsistent)', engine='openpyxl')
            table_result6.to_excel(writer, sheet_name='Checks (PRE)', engine='openpyxl')
            df.to_excel(writer, sheet_name='Validated_Data')

        import openpyxl
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        from openpyxl.chart import BarChart, Reference, Series


        # PALETTES
        workbook = load_workbook(filename)
        sheet_0 = workbook['Description']
        sheet = workbook['Facility Guidance Sheet']
        sheet_1 = workbook['Validation Summary Sheet']
        sheet_2 = workbook['Attention Reqd (Inconsistent)']
        sheet_3 = workbook['Attention Reqd (PRE)']
        sheet_4 = workbook['Checks (Inconsistent)']
        sheet_5 = workbook['Checks (PRE)']

        # Handling contents sheet
        from openpyxl.styles.borders import Border, Side
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))

        workbook.active = sheet_0

        for i in range(10, 20):
            sheet_0.cell(row=i+1, column=10).border = thin_border
            sheet_0.cell(row=i+1, column=11).border = thin_border

        sheet_0.sheet_view.showGridLines = False
        sheet_0.cell(row=7, column=7).alignment = Alignment(horizontal='right')
        sheet_0['J10'] = "   Data Validation Check Tool"
        sheet_0['J11'] = "Facility Type:" + self.lineEdit_2.text()
        sheet_0['J12'] = "Sheet Number"
        sheet_0['J13'] = "Sheet 1"
        sheet_0['J14'] = "Sheet 2"
        sheet_0['J15'] = "Sheet 3"
        sheet_0['J16'] = "Sheet 4"
        sheet_0['J17'] = "Sheet 5"
        sheet_0['J18'] = "Sheet 6"
        sheet_0['J19'] = "Sheet 7"

        sheet_0['K10'] = ""
        sheet_0['K11'] = "Duration:" + self.lineEdit_3.text()
        sheet_0['K12'] = "Description"
        sheet_0['K13'] = "Facility Guidance Sheet"
        sheet_0['K14'] = "Validation Summary sheet"
        sheet_0['K15'] = "Attention Reqd(Inconsistent)"
        sheet_0['K16'] = "Attention Reqd(PRE)"
        sheet_0['K17'] = "Checks (Inconsistent)"
        sheet_0['K18'] = "Checks (PRE)"
        sheet_0['K19'] = "Validated Data"

        sheet_0['J10'].font = Font(size = 12, bold = True)
        sheet_0['J10'].fill = PatternFill(fgColor="FFFFCC", fill_type = "solid")
        sheet_0['K10'].fill = PatternFill(fgColor="FFFFCC", fill_type = "solid")
        sheet_0['J11'].fill = PatternFill(fgColor="00FF00", fill_type = "solid")
        sheet_0['K11'].fill = PatternFill(fgColor="00FF00", fill_type = "solid")
        
        workbook.save(filename=filename)


        # Activating sheets 
        workbook.active = sheet
        workbook.active = sheet_1

        sheet.sheet_view.showGridLines = False
        sheet_1.sheet_view.showGridLines = False

        # Coloring and palettes of Facility Guidance Sheet
        sheet['N4'] = 'Color Brackets'
        sheet['N5'].fill = PatternFill(fgColor="EF5350", fill_type = "solid")
        sheet['N6'].fill = PatternFill(fgColor="FFAF00", fill_type = "solid")
        sheet['N7'].fill = PatternFill(fgColor="C0C000", fill_type = "solid")
        sheet['N8'].fill = PatternFill(fgColor="00FF00", fill_type = "solid")
        sheet['N9'].fill = PatternFill(fgColor="00AF5F", fill_type = "solid")

        sheet["O4"] = "Range"
        sheet["O5"] = ">= 50%"
        sheet["O6"] = "25 - 50%"
        sheet["O7"] = "10 - 25%"
        sheet["O8"] = "5 - 10%"
        sheet["O9"] = "< 5%"
        sheet["O10"] = "Total Facilities"

        sheet["P4"] = "Inconsistent"
        sheet["P5"] = cnt1
        sheet["P6"] = cnt2
        sheet["P7"] = cnt3
        sheet["P8"] = cnt4
        sheet["P9"] = cnt5
        sheet["P10"] = cnt1 + cnt2 + cnt3 + cnt4 + cnt5

        sheet["Q4"] = "Probable Reporting Error"
        sheet["Q5"] = cnt6
        sheet["Q6"] = cnt7
        sheet["Q7"] = cnt8
        sheet["Q8"] = cnt9
        sheet["Q9"] = cnt10
        sheet["Q10"] = cnt6 + cnt7 + cnt8 + cnt9 + cnt10

        
        # Coloring of Validation Summary Sheet
        sheet_1['J10'] = "Color Brackets"
        sheet_1['J11'].fill = PatternFill(fgColor="EF5350", fill_type = "solid")
        sheet_1['J12'].fill = PatternFill(fgColor="FFAF00", fill_type = "solid")
        sheet_1['J13'].fill = PatternFill(fgColor="C0C000", fill_type = "solid")
        sheet_1['J14'].fill = PatternFill(fgColor="00AF5F", fill_type = "solid")

        sheet_1['K10'] = "Range"
        sheet_1['K11'] = ">= 25%"
        sheet_1['K12'] = "10 - 25%"
        sheet_1['K13'] = "5 - 10%"
        sheet_1['K14'] = "< 5%"
        sheet_1['K15'] = "Total Indicators"

        sheet_1['L10'] = "Inconsistent"
        sheet_1['L11'] = cnt21
        sheet_1['L12'] = cnt22
        sheet_1['L13'] = cnt23
        sheet_1['L14'] = cnt24
        sheet_1['L15'] = cnt21 + cnt22 + cnt23 + cnt24

        sheet_1['M10'] = "Probable Reporting Error"
        sheet_1['M11'] = cnt25
        sheet_1['M12'] = cnt26
        sheet_1['M13'] = cnt27
        sheet_1['M14'] = cnt28
        sheet_1['M15'] = cnt25 + cnt26 + cnt27 + cnt28

        ''' 
        GRAPH PLOTS
        '''
        
        Ranges = ['< 5%', '5 - 10%' , '10 - 25%', '25 - 50%', '>= 50%']
        Numbers_Inc = [cnt5, cnt4, cnt3, cnt2, cnt1]
        Numbers_PRE = [cnt10, cnt9, cnt8, cnt7, cnt6]
        
        # plotting a bar graph
        X_axis = np.arange(len(Ranges))

        plt.bar(X_axis - 0.2, Numbers_Inc, 0.4, label = 'Number of Inconsistents')
        plt.bar(X_axis + 0.2, Numbers_PRE, 0.4, label = 'Number of Probable Reporting Errors')
        # figure(figsize=(8, 8), dpi=50)
        plt.xticks(X_axis, Ranges)
        plt.xlabel("Condition")
        plt.ylabel("Number")
        plt.legend()
        plt.title('Error Summary Facility Wise')
        plt.savefig("myplot2.png", dpi = 80)

        img = openpyxl.drawing.image.Image('myplot2.png')
        img.anchor='N13'

        sheet.add_image(img)
        workbook.save(filename=filename)
        workbook.save(filename=filename)

        # Attention Required Sheet (Inconsistent)
        # =======================================

        workbook.active = sheet_2
        sheet_2.sheet_view.showGridLines = False
        for i in range(len(FList1)):

            for j in range(len(FList1[i])):
                sheet_2.cell(row=j+5,column=i+2).value = FList1[i][j]
                
            # Colors
            for k in range(1, len(FList1[i])+5):
                sheet_2.cell(row=4, column=i+1).alignment = Alignment(horizontal='center')
                sheet.cell(row=k, column=3).alignment = Alignment(horizontal='center')
                sheet.cell(row=k, column=4).alignment = Alignment(horizontal='center')
                sheet.cell(row=k, column=5).alignment = Alignment(horizontal='center')
                sheet_1.cell(row=k, column=4).alignment = Alignment(horizontal='center')
                sheet_1.cell(row=k, column=6).alignment = Alignment(horizontal='center')
                sheet_2.cell(row=2, column=i+1).fill = PatternFill(fgColor="fff5be", fill_type = "solid")
                # Top 10 Facility Checks not working well
                sheet_2.cell(row=k+2, column=2).fill = PatternFill(fgColor="EF9A9A", fill_type = "solid")
                sheet_2.cell(row=k+2, column=3).fill = PatternFill(fgColor="EF9A9A", fill_type = "solid")
                sheet_2.cell(row=k+2, column=4).fill = PatternFill(fgColor="EF9A9A", fill_type = "solid")
                sheet_2.cell(row=k+2, column=5).fill = PatternFill(fgColor="EF9A9A", fill_type = "solid")
                sheet_2.cell(row=k+2, column=6).fill = PatternFill(fgColor="EF9A9A", fill_type = "solid")
                sheet_2.cell(row=k+2, column=7).fill = PatternFill(fgColor="EF9A9A", fill_type = "solid")
                sheet_2.cell(row=k+2, column=8).fill = PatternFill(fgColor="EF9A9A", fill_type = "solid")
                sheet_2.cell(row=k+2, column=9).fill = PatternFill(fgColor="EF9A9A", fill_type = "solid")
                sheet_2.cell(row=k+2, column=10).fill = PatternFill(fgColor="EF9A9A", fill_type = "solid")
                sheet_2.cell(row=k+2, column=11).fill = PatternFill(fgColor="EF9A9A", fill_type = "solid")
        
        workbook.save(filename=filename)


        # Attention Required Sheet (PRE)
        # =======================================
        workbook.active = sheet_3

        sheet_3.sheet_view.showGridLines = False
        for i in range(len(FList2)):

            for j in range(len(FList2[i])):
                sheet_3.cell(row=j+5,column=i+2).value = FList2[i][j]
                
            # Colors
            for k in range(1, len(FList2[i])+5):
                sheet_3.cell(row=4, column=i+1).alignment = Alignment(horizontal='center')
                sheet_3.cell(row=2, column=i+1).fill = PatternFill(fgColor="fff5be", fill_type = "solid")
                # Top 10 Facility Checks not working well
                sheet_3.cell(row=k+2, column=2).fill = PatternFill(fgColor="EF9A9A", fill_type = "solid")
                sheet_3.cell(row=k+2, column=3).fill = PatternFill(fgColor="EF9A9A", fill_type = "solid")
                sheet_3.cell(row=k+2, column=4).fill = PatternFill(fgColor="EF9A9A", fill_type = "solid")
                sheet_3.cell(row=k+2, column=5).fill = PatternFill(fgColor="EF9A9A", fill_type = "solid")
                sheet_3.cell(row=k+2, column=6).fill = PatternFill(fgColor="EF9A9A", fill_type = "solid")
                sheet_3.cell(row=k+2, column=7).fill = PatternFill(fgColor="EF9A9A", fill_type = "solid")
                sheet_3.cell(row=k+2, column=8).fill = PatternFill(fgColor="EF9A9A", fill_type = "solid")
                sheet_3.cell(row=k+2, column=9).fill = PatternFill(fgColor="EF9A9A", fill_type = "solid")
                sheet_3.cell(row=k+2, column=10).fill = PatternFill(fgColor="EF9A9A", fill_type = "solid")
                sheet_3.cell(row=k+2, column=11).fill = PatternFill(fgColor="EF9A9A", fill_type = "solid")
        
        workbook.save(filename=filename)


        # Checks Sheet (Inconsistent)
        # =======================================
        workbook.active = sheet_4

        sheet_4.sheet_view.showGridLines = False
        for i in range(len(FList3)):
            for j in range(len(FList3[i])):
                sheet_4.cell(row=j+4,column=i+2).value = FList3[i][j]

        workbook.save(filename=filename)

        # Checks Sheet (PRE)
        # =======================================
        workbook.active = sheet_5

        sheet_5.sheet_view.showGridLines = False
        for i in range(len(FList4)):
            for j in range(len(FList4[i])):
                sheet_5.cell(row=j+4,column=i+2).value = FList4[i][j]

        workbook.save(filename=filename)


        # Create the messagebox object
        self.msg = QMessageBox()
        # Set the information icon
        self.msg.setWindowIcon(QtGui.QIcon('checked.png'))
        self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
        # Set the main message
        self.msg.setText("Excel file exported to the file location you have chosen")
        # Set the title of the window
        self.msg.setWindowTitle("Export Successful")
        # Display the message box
        self.msg.show()

    # RESET FUNCTION
    def reset(self):
        self.lineEdit.clear()
        self.lineEdit_2.clear()
        self.lineEdit_3.clear()
        self.pushButton.setEnabled(True)
        self.pushButton_2.setEnabled(True)
        self.pushButton_3.setEnabled(True)
        self.pushButton_4.setEnabled(True)
        self.pushButton_5.setEnabled(True)
        self.pushButton_6.setEnabled(True)
        self.pushButton_7.setEnabled(True)
        self.pushButton_3.setText('-- All Selected --')
        self.pushButton_4.setText('-- All Selected --')
        self.pushButton_5.setText('-- All Selected --')
        self.pushButton_6.setText('-- All Selected --')


        # Create the messagebox object
        self.msg = QMessageBox()
        # Set the information icon
        self.msg.setWindowIcon(QtGui.QIcon('checked.png'))
        self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
        # Set the main message
        self.msg.setText("Reset Successful , Now you can upload data again.")
        # Set the title of the window
        self.msg.setWindowTitle("Reset Successful Message")
        # Display the message box
        self.msg.show()

    # Display methodology pdf in browser
    def methodology(self):
        import os
        os.system('start Steps.pdf')

    # Display methodology pdf in browser
    def UserManual(self):
        import os
        os.system('start Steps.pdf')


################################### Main Function ############################
if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Dialog = QtWidgets.QDialog()
    ui = Ui_Dialog()
    ui.setupUi(Dialog)
    Dialog.show()
    sys.exit(app.exec_())